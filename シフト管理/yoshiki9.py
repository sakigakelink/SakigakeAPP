"""
Sakigake Shift - 様式9 Excel Generator
入院基本料等の施設基準に係る届出書添付書類を生成する。

3つのファイルを出力:
1. 様式9（精神療養 1病棟）- Ward 1 as 療養病棟 (特定入院料シート)
2. 様式9（特殊疾患 3病棟）- Ward 3 as 特殊疾患病棟 (特定入院料シート)
3. 様式9（15対1 2.3病棟）- Wards 2+3 combined (一般病棟シート)

テンプレートExcelの行・列位置に完全一致するよう出力する。
"""
import os
import json
import math
import calendar
import logging
from io import BytesIO
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# ============================================================
# 書式定数 (テンプレート準拠)
# ============================================================
# 数値フォーマット
FMT_1DEC = '0.0'       # ②③⑥⑦⑧⑨ 等 (小数1位)
FMT_2DEC = '0.00'      # 月延べ時間, 勤務時間 (小数2位)
FMT_3DEC = '0.000'     # ⑩ (小数3位)
FMT_INT = '0'          # 整数 (人数, 患者数等)

# フォント
FONT_TITLE = Font(name='ＭＳ Ｐゴシック', size=14, bold=True)
FONT_SECTION = Font(name='ＭＳ Ｐゴシック', size=11)
FONT_HEADER = Font(name='ＭＳ Ｐゴシック', size=9)
FONT_VALUE = Font(name='ＭＳ Ｐゴシック', size=10, bold=True)
FONT_LABEL = Font(name='ＭＳ Ｐゴシック', size=8)
FONT_TABLE = Font(name='ＭＳ Ｐゴシック', size=7)
FONT_SMALL = Font(name='ＭＳ Ｐゴシック', size=6)

# アライメント
ALIGN_WRAP = Alignment(wrap_text=True, vertical='top')
ALIGN_WRAP_CENTER = Alignment(wrap_text=True, vertical='center', horizontal='center')
ALIGN_CENTER = Alignment(vertical='center', horizontal='center')

# ============================================================
# テンプレート行定義
# ward23 = 一般病棟・療養病棟以外の入院基本料 (778 rows)
# ward1/ward3 = 特定入院料 (看護職員＋看護補助者) (786/747 rows)
# ============================================================
# 0-indexed row numbers (matching the template analysis output)
# openpyxl uses 1-indexed, so add 1 when writing

LAYOUT_IPPAN = {
    # 一般病棟・療養病棟以外の入院基本料 (ward23)
    "sheet_name": "一般病棟・療養病棟以外の入院基本料",
    "header": {
        "title_row": 0,
        "hospital_row": 2,    # col 9
        "todokede_row": 5,    # col 13=ward_type, col 27=ratio
        "bcount_row": 9,      # col 13=ward_count
        "beds_row": 10,       # col 13=bed_count
    },
    "summary": {
        "section_row": 50,    # "3．入院患者の数及び看護要員の数"
        "patients_row": 51,   # ① col 23=人数
        "nurse_config_row": 53,  # ② col 23=配置数, col 45=月延べ
        "nurse_ref_row": 54,     # (参考) col 23=必要数, col 45=基準値
        "ratio_row": 56,      # ③ col 23=%, col 45=看護師時間
        "ratio_sub_row": 57,  # col 45=准看護師時間
        "ratio_nurse_row": 58, # 看護師配置数 col 23, col 45=看護補助者時間
        "avg_stay_row": 60,   # ④
        "night_band_row": 62, # ⑤
        "avg_night_row": 64,  # ⑥ col 23
        "min_nurse_row": 67,  # ⑦ col 23, col 45
        "min_nurse_sub_row": 68,  # ⑦※小数点 + 基準時間(Row 69)
        "min_nurse_note_row": 69, # ⑦届出注記
        "min_nurse_ref_row": 70,
        "aide_config_row": 72, # ⑧ col 23, col 45
        "aide_note_rows": [73, 74, 75, 76],  # ⑧届出注記(4行) + 基準時間
        "aide_adjusted_row": 77,
        "aide_ref_row": 78,
        "aide_night_row": 81, # ⑨ col 23, col 45
        "aide_night_note_rows": [82, 83, 84],  # ⑨届出注記(3行) + 基準時間
        "aide_night_ref_row": 85,
        "clerical_row": 87,   # ⑩ col 23
        "clerical_ref_row": 89,
        "all_staff_row": 92,  # ⑪ col 23 (row 92-93)
        "all_staff_ref_row": 95,
        "yearmonth_row": 98,  # col 2=year, col 5=month, col 23=days
    },
    "staff": {
        "section_row": 100,   # "4．勤務実績表"
        "nurse_label_row": 101,
        "header_row": 102,    # 列ヘッダー
        "date_row": 103,      # 日付
        "dow_row": 104,       # 曜日
        "data_start_row": 105, # 看護職員データ開始
        "max_slots": 100,
    },
    # 固定行位置 (テンプレートから取得)
    "aide_label_row": 409,       # 《看護補助者表》
    "aide_header_row": 410,      # 種別, 番号, ...
    "aide_date_row": 411,        # 日付
    "aide_dow_row": 412,         # 曜日
    "aide_data_start_row": 413,  # 看護補助者データ開始
    "other_staff_label_row": 717, # 《その他職員表》
    "other_staff_header_row": 718,
    "other_staff_date_row": 719,
    "other_staff_dow_row": 720,
    "other_staff_data_start_row": 721,
    "final_b_row": 728,          # [B] 夜勤従事者数
    "section5_row": 736,         # "5．勤務体制及び申し送り時間"
    "schedule_row": 738,         # 3交代制
    "moushiokuri_row": 741,
    "ghijk_row": 746,            # [G][H][I][J][K] section
    "l_row": 757,                # [L] section
    "dist_row": 764,             # 夜勤時間分布
    "has_two_part_nurse": False,  # ②が単一行
    "has_other_staff_table": True,
    "has_schedule_section": True,
    "summary_value_col": 23,     # 統計値のメイン列
    "summary_hours_col": 45,     # 月延べ時間数の列
    "ghijk_value_col": 33,       # [G][H]..の値列
    "ghijk_formula_col": 19,     # 算式の列
}

LAYOUT_TOKUTEI = {
    # 特定入院料 (看護職員＋看護補助者) - ward1 & ward3
    "sheet_name": "特定入院料 (看護職員＋看護補助者)",
    "header": {
        "title_row": 0,
        "hospital_row": 2,
        "todokede_row": 5,    # col 13=ward_type, col 30=ratio
        "bcount_row": 9,
        "beds_row": 10,
    },
    "has_two_part_nurse": True,  # ②-1, ②-2 が分かれる
    "summary_value_col": 23,
    "summary_hours_col": 45,
    "ghijk_value_col": 33,
    "ghijk_formula_col": 19,
}

# ward1 (精神療養 1病棟) - 786 rows
LAYOUT_WARD1 = {
    **LAYOUT_TOKUTEI,
    "summary": {
        "section_row": 50,
        "patients_row": 51,
        "nurse_all_row": 53,    # ②-1 (看護職員+看護補助者) col 23, col 45
        "nurse_all_sub_row": 54, # ※小数点第２位以下切り捨て + 基準時間
        "nurse_all_adj_row": 55, # ⑩調整後
        "nurse_all_ref_row": 56, # (参考)
        "nurse_only_row": 58,   # ②-2 (看護職員のみ) col 23, col 45
        "nurse_only_ref_row": 59,
        "ratio1_row": 61,       # ③-1 看護要員中の看護職員比率
        "ratio1_base_row": 62,
        "ratio2_row": 64,       # ③-2 看護職員中の看護師比率
        "ratio2_desc_row": 65,  # ③-2 数式説明
        "ratio2_base_row": 66,  # ③-2 基準値
        "avg_stay_row": 68,
        "night_band_row": 70,   # ⑤
        "avg_night_row": 72,    # ⑥
        "min_nurse_row": 75,    # ⑦
        "min_nurse_sub_row": 76,  # ⑦※小数点+基準時間
        "min_nurse_note_row": 77, # ⑦届出注記
        "min_nurse_ref_row": 78,
        "aide_config_row": 80,  # ⑧
        "aide_note_rows": [81, 82, 83, 84],  # ⑧届出注記(4行)
        "aide_adjusted_row": 85,
        "aide_ref_row": 86,
        "aide_ref_sub_row": 87,   # ⑧参考※小数点以下切り上げ
        "aide_night_row": 89,   # ⑨
        "aide_night_note_rows": [90, 91, 92],  # ⑨届出注記(3行)
        "aide_night_ref_row": 93,
        "clerical_row": 95,     # ⑩
        "clerical_sub_row": 96,   # ⑩※小数点第３位
        "clerical_ref_row": 97,
        "clerical_ref_sub_row": 98,  # ⑩参考※小数点第３位
        "all_staff_row": 100,   # ⑪
        "all_staff_note_row": 102,  # ⑪届出注記
        "all_staff_ref_row": 103,
        "all_staff_ref_sub_row": 104,  # ⑪参考※小数点
        "yearmonth_row": 106,
    },
    "staff": {
        "section_row": 108,
        "nurse_label_row": 109,
        "header_row": 110,
        "date_row": 111,
        "dow_row": 112,
        "data_start_row": 113,
        "max_slots": 100,
    },
    # 固定行位置 (テンプレートから取得)
    "aide_label_row": 417,
    "aide_header_row": 418,
    "aide_date_row": 419,
    "aide_dow_row": 420,
    "aide_data_start_row": 421,
    "other_staff_label_row": 725,
    "other_staff_header_row": 726,
    "other_staff_date_row": 727,
    "other_staff_dow_row": 728,
    "other_staff_data_start_row": 729,
    "final_b_row": 736,
    "section5_row": 744,
    "schedule_row": 746,
    "moushiokuri_row": 749,
    "ghijk_row": 754,
    "l_row": 765,
    "dist_row": 772,
    "has_other_staff_table": True,
    "has_schedule_section": True,
}

# ward3 (特殊疾患 3病棟) - 747 rows
LAYOUT_WARD3 = {
    **LAYOUT_TOKUTEI,
    "summary": {
        "section_row": 41,
        "patients_row": 42,
        "nurse_all_row": 44,    # ②-1
        "nurse_all_sub_row": 45, # ※小数点第２位以下切り捨て + 基準時間
        "nurse_all_adj_row": 46,
        "nurse_all_ref_row": 47,
        "nurse_only_row": 49,   # ②-2
        "nurse_only_ref_row": 50,
        "ratio1_row": 52,       # ③-1
        "ratio1_base_row": 53,
        "ratio2_row": 55,       # ③-2
        "ratio2_desc_row": 56,  # ③-2 数式説明
        "ratio2_base_row": 57,  # ③-2 基準値
        "avg_stay_row": 59,
        "night_band_row": 61,   # ⑤
        "avg_night_row": 63,    # ⑥
        "min_nurse_row": 66,    # ⑦
        "min_nurse_sub_row": 67,
        "min_nurse_note_row": 68,
        "min_nurse_ref_row": 69,
        "aide_config_row": 71,  # ⑧
        "aide_note_rows": [72, 73],
        "aide_adjusted_row": 74,
        "aide_ref_row": 75,
        "aide_ref_sub_row": 76,
        "aide_night_row": 78,   # ⑨
        "aide_night_note_rows": [79, 80],
        "aide_night_ref_row": 81,
        "clerical_row": 83,     # ⑩
        "clerical_sub_row": 84,
        "clerical_ref_row": 85,
        "clerical_ref_sub_row": 86,
        "yearmonth_row": 88,
    },
    "staff": {
        "section_row": 90,
        "nurse_label_row": 91,
        "header_row": 92,
        "date_row": 93,
        "dow_row": 94,
        "data_start_row": 95,
        "max_slots": 100,
    },
    # 固定行位置 (テンプレートから取得)
    "aide_label_row": 399,
    "aide_header_row": 400,
    "aide_date_row": 401,
    "aide_dow_row": 402,
    "aide_data_start_row": 403,
    "final_b_row": 707,
    "ghijk_row": 715,
    "l_row": 726,
    "dist_row": 733,
    "has_other_staff_table": False,
    "has_schedule_section": False,
}

LAYOUT_MAP = {
    "ward1": LAYOUT_WARD1,
    "ward3": LAYOUT_WARD3,
    "ward23": LAYOUT_IPPAN,
}


def R(row_0indexed):
    """0-indexed row to openpyxl 1-indexed"""
    return row_0indexed + 1


def C(col_0indexed):
    """0-indexed col to openpyxl 1-indexed"""
    return col_0indexed + 1


# --- Excel数式ヘルパー ---
def cell_ref(row_0, col_0):
    """0-indexed → Excel '$A$1' 形式の絶対参照"""
    return f"${get_column_letter(col_0 + 1)}${row_0 + 1}"


def cell_ref_rel(row_0, col_0):
    """0-indexed → Excel 'A1' 形式の相対参照"""
    return f"{get_column_letter(col_0 + 1)}{row_0 + 1}"


def col_letter(col_0):
    """0-indexed → 列文字 ('A', 'B', ...)"""
    return get_column_letter(col_0 + 1)


# --- 設定読み込み ---
def load_config():
    config_path = os.path.join(BASE_DIR, "yoshiki9_config.json")
    with open(config_path, "r", encoding="utf-8") as f:
        return json.load(f)


def load_employees():
    path = os.path.join(BASE_DIR, "shared", "employees.json")
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def load_actual_shifts(ward_id, year, month):
    path = os.path.join(BASE_DIR, "shifts", ward_id, f"{year}-{month:02d}.json")
    if not os.path.exists(path):
        return None
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    actual = data.get("actual")
    if actual and actual.get("shifts"):
        return actual["shifts"]
    confirmed = data.get("confirmed")
    if confirmed and confirmed.get("shifts"):
        return confirmed["shifts"]
    selected = data.get("selectedDraft")
    drafts = data.get("drafts", {})
    if selected and selected in drafts:
        draft_shifts = drafts[selected].get("shifts")
        if draft_shifts:
            return draft_shifts
    return None


def load_day_hours(ward_id, year, month):
    """dayHoursデータを読み込む（confirmed/actualから）"""
    path = os.path.join(BASE_DIR, "shifts", ward_id, f"{year}-{month:02d}.json")
    if not os.path.exists(path):
        return {}
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    actual = data.get("actual")
    if actual and actual.get("dayHours"):
        return actual["dayHours"]
    confirmed = data.get("confirmed")
    if confirmed and confirmed.get("dayHours"):
        return confirmed["dayHours"]
    return {}


# --- 勤務時間計算 ---
def calculate_staff_hours(staff_shifts, num_days, shift_hours_config,
                          all_day_shifts_by_day=None, staff_id=None,
                          day_hours_override=None):
    """
    day_hours_override: dict of {"staffId-day": hours} for custom day shift hours.
      デフォルト(7.5h): 0830始業 → config通り（日勤8.0、夜勤0.5）
      カスタム: 0900始業、1630以降が夜勤帯（日勤帯=7.5h）
      休憩: 勤務 <= 6.0 → 0、勤務 > 6.0 → 1.0h
      拘束時間 = 勤務時間 + 休憩
      - 拘束 <= 7.5 → 日勤=勤務時間、夜勤=0
      - 拘束 > 7.5  → 日勤=勤務時間、夜勤=拘束-7.5
      例: 日3→日勤3,夜勤0  日6→日勤6,夜勤0  日7→日勤7,夜勤0.5
    """
    result = []
    for d in range(1, num_days + 1):
        shift = staff_shifts.get(str(d), "")
        cfg = shift_hours_config.get(shift, {"day": 0.0, "night": 0.0})
        day_h = cfg.get("day", 0.0)
        night_h = cfg.get("night", 0.0)

        # 日勤時間数のカスタマイズ適用
        if shift == "day" and day_hours_override and staff_id:
            override_key = f"{staff_id}-{d}"
            if override_key in day_hours_override:
                custom_h = float(day_hours_override[override_key])
                if custom_h != 7.5:
                    # 休憩: 6時間以内は0、6時間超は1.0h
                    break_h = 0.0 if custom_h <= 6.0 else 1.0
                    # 拘束時間 = 勤務 + 休憩 (0900始業)
                    bound_h = custom_h + break_h
                    if bound_h > 7.5:
                        # 16:30超え → 日勤=7.5(日勤帯フル)、夜勤=拘束-7.5
                        day_h = 7.5
                        night_h = bound_h - 7.5
                    else:
                        # 16:30以前に終了 → 日勤=勤務時間、夜勤なし
                        day_h = custom_h
                        night_h = 0.0

        if shift == "day" and all_day_shifts_by_day and staff_id:
            day_staff_list = all_day_shifts_by_day.get(d, [])
            # 申し送り対象は通常8.5h勤務者のみ（カスタム勤務時間の職員は除外）
            override_key_m = f"{staff_id}-{d}"
            has_override = (day_hours_override and override_key_m in day_hours_override)
            if not has_override:
                # 通常勤務者の中で1番手を探す
                for candidate in day_staff_list:
                    cand_key = f"{candidate}-{d}"
                    if not (day_hours_override and cand_key in day_hours_override):
                        # この候補が通常勤務の1番手
                        if candidate == staff_id:
                            moushiokuri = shift_hours_config.get("day_moushiokuri", {})
                            night_h = moushiokuri.get("night", 0.25)
                        break

        result.append({
            "day": d,
            "day_h": day_h,
            "night_h": night_h,
            "total_night_h": night_h,
        })
    return result


def build_day_shift_roster(all_shifts, staff_list, num_days):
    roster = {}
    for d in range(1, num_days + 1):
        day_staff = []
        for emp in staff_list:
            sid = emp["id"]
            shifts = all_shifts.get(sid, {})
            if shifts.get(str(d)) == "day":
                day_staff.append(sid)
        roster[d] = day_staff
    return roster


def truncate(value, decimals):
    """小数点以下切り捨て（0方向への切り捨て）"""
    factor = 10 ** decimals
    return math.trunc(value * factor) / factor


# ============================================================
# Excel生成メインクラス
# ============================================================
class Yoshiki9Generator:
    def __init__(self, year, month, patients=None, day_hours=None):
        self.year = year
        self.month = month
        self.num_days = calendar.monthrange(year, month)[1]
        self.config = load_config()
        self.employees = load_employees()
        self.shift_hours = self.config["shift_hours"]
        self.patients = patients or {}
        self.day_hours = day_hours or {}

    def generate(self, ward_key):
        """単一の様式9ファイルを生成"""
        wc = self.config["ward_configs"][ward_key]
        layout = LAYOUT_MAP[ward_key]
        ward_ids = wc["wards"]
        ward_names = wc.get("ward_names", {})

        staff_list = [e for e in self.employees if e["ward"] in ward_ids]
        nurses = [e for e in staff_list if e["type"] in ("nurse", "junkango")]
        aides = [e for e in staff_list if e["type"] == "nurseaide"]

        all_shifts = {}
        for wid in ward_ids:
            shifts = load_actual_shifts(wid, self.year, self.month)
            if shifts:
                all_shifts.update(shifts)

        # フロントエンドから送られたdayHoursを優先、なければJSONファイルから読み込み
        all_day_hours = {}
        if self.day_hours:
            all_day_hours = self.day_hours
        else:
            for wid in ward_ids:
                dh = load_day_hours(wid, self.year, self.month)
                if dh:
                    all_day_hours.update(dh)

        day_roster = build_day_shift_roster(all_shifts, staff_list, self.num_days)

        nurse_hours = {}
        for emp in nurses:
            sid = emp["id"]
            nurse_hours[sid] = calculate_staff_hours(
                all_shifts.get(sid, {}), self.num_days, self.shift_hours,
                all_day_shifts_by_day=day_roster, staff_id=sid,
                day_hours_override=all_day_hours
            )

        aide_hours = {}
        for emp in aides:
            sid = emp["id"]
            aide_hours[sid] = calculate_staff_hours(
                all_shifts.get(sid, {}), self.num_days, self.shift_hours,
                all_day_shifts_by_day=day_roster, staff_id=sid,
                day_hours_override=all_day_hours
            )

        wb = Workbook()
        ws = wb.active
        ws.title = layout.get("sheet_name", wc.get("sheet_name", "様式9"))

        # ============ ヘッダー ============
        self._write_header(ws, wc, layout)

        # ============ 集計統計 (Python側で分類データを計算) ============
        stats = self._calc_stats(wc, ward_key, layout, nurses, aides,
                                 nurse_hours, aide_hours, all_shifts)

        # ============ 年月・稼働日数 ============
        ym = layout["summary"]["yearmonth_row"]
        ws.cell(R(ym), C(2), self.year)
        ws.cell(R(ym), C(4), "年")
        ws.cell(R(ym), C(5), self.month)
        ws.cell(R(ym), C(7), "月")
        ws.cell(R(ym), C(18), "※今月の稼働日数")
        ws.cell(R(ym), C(23), self.num_days)
        ws.cell(R(ym), C(25), "日")

        # 稼働日数セルの参照 (数式から参照する)
        days_cell = cell_ref_rel(ym, 23)

        # ============ 勤務実績表ヘッダー ============
        sl = layout["staff"]
        ws.cell(R(sl["section_row"]), C(1), "4．勤務実績表")
        ws.cell(R(sl["nurse_label_row"]), C(2), "《看護職員表》")
        self._write_table_header(ws, sl["header_row"], sl["date_row"], sl["dow_row"])

        # ============ 看護職員データ ============
        nurse_result = self._write_staff_data(
            ws, nurses, nurse_hours, all_shifts, ward_names,
            sl["data_start_row"], is_nurse=True, layout=layout
        )

        # 未使用スロットにラベルだけ書き込む (看護職員表)
        self._write_empty_slot_labels(
            ws, sl["data_start_row"], nurse_result["staff_count"],
            sl["max_slots"], is_nurse=True)

        # 看護職員合計行 (data_start + 100 slots × 3 rows = total row area)
        nurse_total_row = sl["data_start_row"] + sl["max_slots"] * 3
        self._write_totals(ws, nurse_result, nurse_total_row, sl["data_start_row"])

        # 未使用行を非表示 (看護職員表)
        self._hide_unused_rows(ws, sl["data_start_row"],
                               nurse_result["staff_count"], sl["max_slots"])

        # ============ 看護補助者表 (固定位置) ============
        ws.cell(R(layout["aide_label_row"]), C(2), "《看護補助者表》")
        self._write_table_header(
            ws, layout["aide_header_row"],
            layout["aide_date_row"], layout["aide_dow_row"]
        )

        aide_result = self._write_staff_data(
            ws, aides, aide_hours, all_shifts, ward_names,
            layout["aide_data_start_row"], is_nurse=False, layout=layout
        )

        # 未使用スロットにラベルだけ書き込む (看護補助者表)
        self._write_empty_slot_labels(
            ws, layout["aide_data_start_row"], aide_result["staff_count"],
            sl["max_slots"], is_nurse=False)

        aide_total_row = layout["aide_data_start_row"] + sl["max_slots"] * 3
        self._write_totals(ws, aide_result, aide_total_row, layout["aide_data_start_row"])

        # 未使用行を非表示 (看護補助者表)
        self._hide_unused_rows(ws, layout["aide_data_start_row"],
                               aide_result["staff_count"], sl["max_slots"])

        # ============ その他職員表 (ward1, ward23のみ) ============
        if layout.get("has_other_staff_table"):
            self._write_other_staff_table(ws, layout)

        # ============ 最終集計セクション [B][C][D][E][F] → Excel数式 ============
        refs = self._write_final_summary(
            ws, layout["final_b_row"], layout, wc,
            nurse_result, aide_result, stats,
            nurse_total_row, aide_total_row,
            sl["data_start_row"], days_cell)

        # ============ 勤務体制 ============
        if layout.get("has_schedule_section"):
            if "section5_row" in layout:
                ws.cell(R(layout["section5_row"]), C(1),
                        "5．勤務体制及び申し送り時間")
            self._write_schedule(ws, layout)

        # ============ [G][H][I][J][K] → Excel数式 ============
        ghijk_refs = self._write_ghijk(
            ws, layout, stats, refs, aide_total_row, days_cell)

        # ============ [L] → Excel数式 ============
        self._write_l_section(ws, layout, stats, refs, days_cell)

        # ============ 集計統計セクション → Excel数式 (最後に書く) ============
        # _write_summary は refs, ghijk_refs を参照するため最後に呼ぶ
        self._write_summary(ws, layout, wc, ward_key, stats,
                           refs, ghijk_refs, days_cell,
                           nurse_total_row, aide_total_row,
                           sl["data_start_row"], layout["aide_data_start_row"])

        # ============ ward1: 不要な空白行を非表示 ============
        if ward_key == "ward1":
            s = layout["summary"]
            # ①②③(rows 50-66)と⑤(row 70)とyearmonth(row 106)のみ表示
            # それ以外の空白行・不要項目行を非表示にする
            hide_rows = set()
            hide_rows.add(67)   # ③と④の間の空白
            hide_rows.add(68)   # ④
            hide_rows.add(69)   # ④と⑤の間の空白
            hide_rows.add(71)   # ⑤と⑥の間の空白
            for r in range(72, 106):  # ⑥⑦⑧⑨⑩⑪ + 空白行
                hide_rows.add(r)
            for r in hide_rows:
                ws.row_dimensions[R(r)].hidden = True
        elif ward_key == "ward3":
            # ward3: ①②-1(rows 41-47)と⑤(row 61)とyearmonth(row 88)のみ表示
            for r in range(48, 61):  # ②-2, ③-1, ③-2, ④ + 空白行
                ws.row_dimensions[R(r)].hidden = True
            for r in range(62, 88):  # ⑥⑦⑧⑨⑩ + 空白行
                ws.row_dimensions[R(r)].hidden = True
        elif ward_key == "ward23":
            # ward23: ①②③④⑤⑥⑧のみ表示、⑦⑨⑩⑪は非表示
            for r in range(65, 72):  # ⑥後の空白 + ⑦ + ⑦と⑧の間の空白
                ws.row_dimensions[R(r)].hidden = True
            for r in range(79, 98):  # ⑧後の空白 + ⑨⑩⑪ + 空白行
                ws.row_dimensions[R(r)].hidden = True

        # ============ 夜勤時間分布 ============
        self._write_night_distribution(ws, layout, nurses, aides,
                                       nurse_hours, aide_hours)

        # 列幅 (テンプレート準拠)
        self._apply_column_widths(ws)

        # セル結合 (統計セクションの値セル)
        self._apply_merges(ws, layout)

        # 書式 (フォント・数値フォーマット)
        self._apply_formats(ws, layout)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def generate_all_zip(self):
        import zipfile
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
            for key in ["ward1", "ward3", "ward23"]:
                wc = self.config["ward_configs"][key]
                xlsx_data = self.generate(key)
                filename = f"{wc['file_label']}_{self.year}年{self.month}月.xlsx"
                zf.writestr(filename, xlsx_data.read())
        zip_buffer.seek(0)
        return zip_buffer

    # ============================================================
    # ヘッダー書き込み
    # ============================================================
    def _write_header(self, ws, wc, layout):
        h = layout["header"]
        cell = ws.cell(R(h["title_row"]), C(2),
                       "(様式９)入院基本料等の施設基準に係る届出書添付書類")
        cell.font = FONT_TITLE
        ws.cell(R(h["title_row"]), C(38), "作成年月日").font = FONT_LABEL
        ws.cell(R(h["title_row"]), C(41), self.year)
        ws.cell(R(h["title_row"]), C(44), "年")
        ws.cell(R(h["title_row"]), C(46), self.month)
        ws.cell(R(h["title_row"]), C(48), "月")

        ws.cell(R(h["hospital_row"]), C(2), "保険医療機関名").font = FONT_LABEL
        ws.cell(R(h["hospital_row"]), C(9),
                self.config["hospital_name"]).font = FONT_SECTION

        ws.cell(R(h["todokede_row"]), C(2),
                "届出入院基本料・特定入院料（届出区分）").font = FONT_LABEL
        ws.cell(R(h["todokede_row"]), C(13), wc["ward_type"]).font = FONT_HEADER

        # ratio (ward23: col 27, ward1/3: col 30)
        ratio = wc.get("ratio", "")
        if ratio:
            parts = ratio.split("対")
            if len(parts) == 2:
                if layout is LAYOUT_IPPAN:
                    ws.cell(R(h["todokede_row"]), C(27), int(parts[0]))
                    ws.cell(R(h["todokede_row"]), C(29), "対１")
                else:
                    ws.cell(R(h["todokede_row"]), C(29), "看護配置・看護補助配置")
                    ws.cell(R(h["todokede_row"]), C(30), int(parts[0]))
                    ws.cell(R(h["todokede_row"]), C(32), "対１")

        ws.cell(R(h["bcount_row"]), C(2), "本届出の病棟数")
        ws.cell(R(h["bcount_row"]), C(13), wc["ward_count"])
        ws.cell(R(h["beds_row"]), C(2), "本届出の病床数")
        ws.cell(R(h["beds_row"]), C(13), wc["bed_count"])

    # ============================================================
    # 統計計算
    # ============================================================
    def _calc_stats(self, wc, ward_key, layout, nurses, aides,
                    nurse_hours, aide_hours, all_shifts):
        patient_count = self._get_patient_count(ward_key)
        divisor = self.num_days * 8

        # 看護職員の合計時間
        nurse_day_total = 0.0
        nurse_night_total = 0.0
        nurse_type_hours = {"nurse": 0.0, "junkango": 0.0}
        for emp in nurses:
            sid = emp["id"]
            hours = nurse_hours.get(sid, [])
            emp_day = sum(h["day_h"] for h in hours)
            emp_night = sum(h["night_h"] for h in hours)
            emp_total = emp_day + emp_night
            nurse_day_total += emp_day
            nurse_night_total += emp_night
            nurse_type_hours[emp["type"]] += emp_total

        total_nurse_hours = nurse_day_total + nurse_night_total  # [C] 看護職員
        nurse_total_night = nurse_night_total  # 看護職員の夜勤時間

        # 看護補助者の合計時間
        aide_day_total = 0.0
        aide_night_total = 0.0
        aide_clerical_hours = 0.0  # 事務的業務の看護補助者の時間
        for emp in aides:
            sid = emp["id"]
            hours = aide_hours.get(sid, [])
            emp_day = sum(h["day_h"] for h in hours)
            emp_night = sum(h["night_h"] for h in hours)
            emp_total = emp_day + emp_night
            aide_day_total += emp_day
            aide_night_total += emp_night
            # 事務的業務フラグ（将来対応用、現在は0）

        total_aide_hours = aide_day_total + aide_night_total  # [G] 看護補助者のみ
        aide_total_night = aide_night_total

        # 夜勤従事者: 夜勤シフトがある人(夜専除く、月8時間以上)
        # まず夜勤有/無/夜専を分類
        def classify_night(emp_list, hours_dict):
            result = {"has_night": [], "no_night": [], "night_only": [],
                      "night_juuji": []}
            for emp in emp_list:
                sid = emp["id"]
                shifts = all_shifts.get(sid, {})
                cat = emp.get("shiftCategory", "twoShift")
                is_night_only = (cat == "nightOnly")
                has_night = any(
                    shifts.get(str(d)) in ("night2", "junnya", "shinya")
                    for d in range(1, self.num_days + 1)
                )
                hours = hours_dict.get(sid, [])
                total_night = sum(h["night_h"] for h in hours)

                if is_night_only:
                    result["night_only"].append(emp)
                elif has_night:
                    result["has_night"].append(emp)
                else:
                    result["no_night"].append(emp)

                # 夜勤従事者数への計上: has_night かつ not night_only かつ月夜勤8時間以上
                if has_night and not is_night_only and total_night >= 8:
                    result["night_juuji"].append(emp)
            return result

        nurse_class = classify_night(nurses, nurse_hours)
        aide_class = classify_night(aides, aide_hours)

        # [B] 夜勤従事者数 (夜専を除く) — 看護職員＋看護補助者
        night_workers_b = len(nurse_class["night_juuji"]) + len(aide_class["night_juuji"])
        # 看護職員のみの夜勤従事者数 (⑥の計算用)
        nurse_night_workers_b = len(nurse_class["night_juuji"])

        # [E] 月平均夜勤時間の計算に含まない者の夜勤時間数
        # = 夜専の夜勤時間 + 夜勤無の人の夜勤時間(微量) + 月8時間未満の者
        exclude_night_e = 0.0
        for emp in nurses:
            sid = emp["id"]
            hours = nurse_hours.get(sid, [])
            total_night = sum(h["night_h"] for h in hours)
            cat = emp.get("shiftCategory", "twoShift")
            is_night_only = (cat == "nightOnly")
            shifts = all_shifts.get(sid, {})
            has_night = any(
                shifts.get(str(d)) in ("night2", "junnya", "shinya")
                for d in range(1, self.num_days + 1)
            )
            # 含まない人: 夜専、夜勤無し、または月夜勤8時間未満の者
            if is_night_only or not has_night or total_night < 8:
                exclude_night_e += total_night

        # [D] 月延べ夜勤時間数 (看護職員のみ)
        total_d = nurse_total_night  # 看護職員の夜勤時間
        # ⑥ = (D - E) / B
        # ward1/ward3: D,Eが看護職員のみ → Bも看護職員のみの夜勤従事者数
        # ward23: Bは全夜勤従事者数（看護職員＋看護補助者）
        de = total_d - exclude_night_e
        if layout.get("has_two_part_nurse"):
            avg_night_6 = truncate(de / nurse_night_workers_b, 1) if nurse_night_workers_b > 0 else 0
        else:
            avg_night_6 = truncate(de / night_workers_b, 1) if night_workers_b > 0 else 0

        # 配置数計算
        # [C] = 看護職員の月延べ勤務時間数
        c_value = total_nurse_hours
        # 看護補助者の [G]
        g_value = total_aide_hours

        # ② 看護職員配置数 = C / (日数 × 8)
        nurse_config = truncate(c_value / divisor, 1) if divisor > 0 else 0
        # ②-1 (C+G) / (日数 × 8)
        all_config = truncate((c_value + g_value) / divisor, 1) if divisor > 0 else 0

        # 看護師配置数
        total_nurse_type_hours = nurse_type_hours["nurse"]
        total_junkango_type_hours = nurse_type_hours["junkango"]
        nurse_haichi = truncate(total_nurse_type_hours / divisor, 1) if divisor > 0 else 0

        # 配置区分の数 (ratio: "15対1" -> 15)
        ratio_str = wc.get("ratio", "15対1")
        try:
            ratio_num = int(ratio_str.split("対")[0])
        except (ValueError, IndexError):
            ratio_num = 15
        # 1日看護配置数(必要数) = ceil(A / ratio_num) * 3
        required_config = math.ceil(patient_count / ratio_num) * 3

        # ward1/ward3: ②-2の必要看護職員数 = required_config × (nurse_ratio_base / 100)
        nurse_ratio_base = wc.get("nurse_ratio_base", 0)
        if layout.get("has_two_part_nurse") and nurse_ratio_base > 0:
            required_nurse = required_config * (nurse_ratio_base / 100)
        else:
            required_nurse = 0  # ward23では使わない

        # ③ 看護師比率
        ratio_3 = 0
        ratio_31 = 0
        ratio_32 = 0
        if layout.get("has_two_part_nurse"):
            # ③-1 看護要員中の看護職員比率
            #   = (C / divisor) / required_config × 100
            ratio_31 = truncate((c_value / divisor) / required_config * 100, 1) if (required_config > 0 and divisor > 0) else 0
            # ③-2 看護職員中の看護師比率
            #   = (nurse_h / divisor) / required_nurse × 100
            ratio_32 = truncate((total_nurse_type_hours / divisor) / required_nurse * 100, 1) if (required_nurse > 0 and divisor > 0) else 0
        else:
            # ward23: ③ 看護職員中の看護師比率
            #   = nurse_hours / (必要数 × 8 × 日数) × 100
            #   PDF p.11: 看護師月総勤務時間数 ÷ 「1日平均看護配置数を満たす月延べ勤務時間数」
            base_hours = required_config * 8 * self.num_days
            ratio_3 = truncate(total_nurse_type_hours / base_hours * 100, 1) if base_hours > 0 else 0

        # [H] みなし看護補助者 = C - (1日看護配置数 × 8 × 日数)
        # ward23: 1日看護配置数 = required_config (②の必要数)
        # ward1/ward3: 1日看護配置数 = required_nurse (②-2の必要数)
        if layout.get("has_two_part_nurse") and required_nurse > 0:
            h_required = required_nurse
        else:
            h_required = required_config
        h_value = c_value - (h_required * 8 * self.num_days)
        if h_value < 0:
            h_value = 0

        # [I] 看護補助者のみの月延べ夜勤時間数
        i_value = aide_total_night

        # ⑧ 看護補助者配置数
        if layout.get("has_two_part_nurse"):
            # ward1/ward3(特定入院料): (G + H) / (日数 × 8)
            aide_config_8 = truncate((g_value + h_value) / divisor, 1) if divisor > 0 else 0
        else:
            # ward23: G / (日数 × 8)
            aide_config_8 = truncate(g_value / divisor, 1) if divisor > 0 else 0

        # ⑨ 夜間看護補助者配置数 = I / (日数 × 16)
        aide_night_config_9 = truncate(i_value / (self.num_days * 16), 1) if self.num_days > 0 else 0

        # ⑩ 事務的業務配置数 = F / (日数 × 8) ... F=0 (現在)
        f_value = aide_clerical_hours
        clerical_config_10 = truncate(f_value / divisor, 3) if divisor > 0 else 0
        clerical_ref = truncate((patient_count / 200) * 3, 3)

        # ⑦ 最小必要人数以上の看護職員配置数
        # = {C - (L × 日数 × 8)} / (日数 × 8)
        # L = ward23: ceil(A/13)*3, ward1/ward3(特定入院料): ceil(A/50)*3
        if layout.get("has_two_part_nurse"):
            l_divisor = 50
        else:
            l_divisor = 13
        l_value = math.ceil(patient_count / l_divisor) * 3
        min_nurse_hours = c_value - (l_value * 8 * self.num_days)
        min_nurse_7 = truncate(min_nurse_hours / divisor, 1) if divisor > 0 else 0
        if min_nurse_7 < 0:
            min_nurse_7 = 0

        # [K] = G / (日数 × 8)
        k_value = truncate(g_value / divisor, 1) if divisor > 0 else 0

        # （⑩の実績値-⑩の上限値）を減じた月平均１日当たり看護補助者配置数
        # = ⑧ - (⑩実績 - ⑩上限) = aide_config_8 - (clerical_config_10 - clerical_ref)
        aide_adjusted_raw = aide_config_8 - (clerical_config_10 - clerical_ref)
        aide_adjusted = truncate(aide_adjusted_raw, 1) if aide_adjusted_raw > 0 else 0

        # (G+H) / (日数 × 8)
        gh_haichi = truncate((g_value + h_value) / divisor, 1) if divisor > 0 else 0

        # K/J ratio
        j_value = required_config  # ward23用
        kj_ratio = truncate((k_value / j_value) * 100, 3) if j_value > 0 else 0

        # ⑪ = nurse_config (= ②の値、看護職員のみ)
        all_staff_11 = nurse_config

        return {
            "patient_count": patient_count,
            "divisor": divisor,
            "c_value": c_value,  # 看護職員月延べ
            "g_value": g_value,  # 看護補助者月延べ
            "h_value": h_value,  # みなし看護補助者月延べ
            "i_value": i_value,  # 看護補助者夜勤時間
            "f_value": f_value,  # 事務的業務
            "total_d": total_d,  # 看護職員月延べ夜勤
            "exclude_night_e": exclude_night_e,
            "night_workers_b": night_workers_b,
            "avg_night_6": avg_night_6,
            "nurse_config": nurse_config,  # ②
            "all_config": all_config,      # ②-1
            "nurse_haichi": nurse_haichi,   # 看護師配置数
            "total_nurse_type_hours": total_nurse_type_hours,
            "total_junkango_type_hours": total_junkango_type_hours,
            "total_aide_hours": total_aide_hours,
            "ratio_3": ratio_3 if not layout.get("has_two_part_nurse") else 0,
            "ratio_31": ratio_31 if layout.get("has_two_part_nurse") else 0,
            "ratio_32": ratio_32 if layout.get("has_two_part_nurse") else 0,
            "required_config": required_config,
            "required_nurse": required_nurse,
            "min_nurse_7": min_nurse_7,
            "min_nurse_hours": max(min_nurse_hours, 0),
            "aide_config_8": aide_config_8,
            "aide_night_config_9": aide_night_config_9,
            "clerical_config_10": clerical_config_10,
            "clerical_ref": clerical_ref,
            "aide_adjusted": aide_adjusted,
            "gh_haichi": gh_haichi,
            "k_value": k_value,
            "j_value": j_value,
            "kj_ratio": kj_ratio,
            "all_staff_11": all_staff_11,
            "l_value": l_value,
            "nurse_class": nurse_class,
            "aide_class": aide_class,
            "ratio_num": ratio_num,
            "nurse_ratio_base": nurse_ratio_base,
        }

    # ============================================================
    # 集計統計書き込み
    # ============================================================
    def _write_summary(self, ws, layout, wc, ward_key, stats,
                       refs, ghijk_refs, days_cell,
                       nurse_total_row, aide_total_row,
                       nurse_data_start, aide_data_start):
        """
        集計統計セクション → Excel数式。
        refs: _write_final_summary() から返されたセル位置dict。
        ghijk_refs: _write_ghijk() から返されたセル位置dict。
        days_cell: 稼働日数セル参照文字列。
        nurse/aide_total_row: 合計行位置 (0-indexed)。
        nurse/aide_data_start: データ開始行 (0-indexed)。
        """
        s = layout["summary"]
        vc = C(layout["summary_value_col"])  # col 24 (0-indexed 23)
        vc_0 = layout["summary_value_col"]   # 0-indexed
        hc = C(layout["summary_hours_col"])  # col 46 (0-indexed 45)
        hc_0 = layout["summary_hours_col"]   # 0-indexed

        c_ref = cell_ref_rel(*refs["C"])  # [C] セル
        g_ref = cell_ref_rel(*ghijk_refs["G"])  # [G] セル
        a_cell = cell_ref_rel(s["patients_row"], vc_0)  # ①のセル

        ratio_num = stats["ratio_num"]

        # Section header
        ws.cell(R(s["section_row"]), C(1),
                "3．入院患者の数及び看護要員の数").font = FONT_SECTION

        # ① 1日平均入院患者数 (静的値 — ユーザーが手動変更可能)
        ws.cell(R(s["patients_row"]), C(2),
                "①１日平均入院患者数〔A〕 ※小数点以下切り上げ")
        ws.cell(R(s["patients_row"]), vc, stats["patient_count"])
        ws.cell(R(s["patients_row"]), C(25), "人")
        # 算出期間 (当月の前月末までの12ヶ月間)
        import datetime
        end_dt = datetime.date(self.year, self.month, 1) - datetime.timedelta(days=1)
        start_dt = datetime.date(end_dt.year - 1, end_dt.month, 1)
        pr = s["patients_row"]
        ws.cell(R(pr), C(29), "(算出期間）")
        ws.cell(R(pr), C(30), start_dt.year)
        ws.cell(R(pr), C(32), "年")
        ws.cell(R(pr), C(33), start_dt.month)
        ws.cell(R(pr), C(34), "月")
        ws.cell(R(pr), C(35), start_dt.day)
        ws.cell(R(pr), C(36), "日")
        ws.cell(R(pr), C(37), "～")
        ws.cell(R(pr), C(38), end_dt.year)
        ws.cell(R(pr), C(40), "年")
        ws.cell(R(pr), C(41), end_dt.month)
        ws.cell(R(pr), C(42), "月")
        ws.cell(R(pr), C(43), end_dt.day)
        ws.cell(R(pr), C(44), "日")

        # 看護師時間数のSUMPRODUCT用範囲 (nurseデータのcol62/col63)
        nds = nurse_data_start
        nde = nurse_data_start + 300 - 1
        nurse_z = f"${col_letter(25)}${nds+1}:${col_letter(25)}${nde+1}"
        # col62=看護師時間, col63=准看護師時間 (全3行にデータあり)
        nurse_c62 = f"{col_letter(62)}{nds+1}:{col_letter(62)}{nde+1}"
        nurse_c63 = f"{col_letter(63)}{nds+1}:{col_letter(63)}{nde+1}"

        # aide SUMPRODUCT用
        ads = aide_data_start
        ade = aide_data_start + 300 - 1
        aide_z = f"${col_letter(25)}${ads+1}:${col_letter(25)}${ade+1}"
        aide_c60 = f"{col_letter(60)}{ads+1}:{col_letter(60)}{ade+1}"

        if layout.get("has_two_part_nurse"):
            # ②-1 = TRUNC((C+G)/(days*8), 1)
            ws.cell(R(s["nurse_all_row"]), C(2),
                    "②-1 月平均１日当たり看護職員配置数（看護職員+看護補助者）")
            ws.cell(R(s["nurse_all_row"]), vc).value = (
                f"=TRUNC(({c_ref}+{g_ref})/({days_cell}*8),1)")
            ws.cell(R(s["nurse_all_row"]), C(25), "人(実績値)")
            ws.cell(R(s["nurse_all_row"]), C(28), "〔(Ｃ＋Ｇ)／(日数×８)〕")
            ws.cell(R(s["nurse_all_row"]), hc).value = f"={c_ref}+{g_ref}"
            ws.cell(R(s["nurse_all_row"]), C(48), "時間(実績値)")
            ws.cell(R(s["nurse_all_row"]), C(44), "『月延べ勤務時間数』(看護職員＋看護補助者）")

            # ②-1 サブ行: ※小数点 + 基準時間
            sub_row = s.get("nurse_all_sub_row")
            if sub_row is not None:
                ws.cell(R(sub_row), C(2), "※小数点第２位以下切り捨て")
            # 必要数 (参考行)
            req_row = s.get("nurse_all_ref_row", s.get("nurse_only_ref_row"))
            if req_row is not None:
                ws.cell(R(req_row), C(2),
                        "（参考）1日看護職員配置数（必要数）※小数点以下切り上げ")
                ws.cell(R(req_row), vc).value = f"=CEILING({a_cell}/{ratio_num},1)*3"
                ws.cell(R(req_row), C(25), "人(基準値)")
                ws.cell(R(req_row), C(28), "〔(Ａ／配置区分の数)×３〕")
                # 基準時間 → サブ行に配置
                all_ref_cell = cell_ref_rel(req_row, vc_0)
                target = sub_row if sub_row is not None else req_row
                ws.cell(R(target), hc).value = (
                    f"={all_ref_cell}*8*{days_cell}")
                ws.cell(R(target), C(44),
                        "※「1日平均看護配置数」を満たす「月延べ勤務時間数」")
                ws.cell(R(target), C(48), "時間(基準値)")

            # ②-1調整 (⑩の実績値-⑩の上限値を減じた看護配置数)
            if s.get("nurse_all_adj_row") is not None:
                ws.cell(R(s["nurse_all_adj_row"]), C(2),
                        "（⑩の実績値-⑩の上限値）を減じた月平均１日当たり"
                        "看護配置数（看護職員+看護補助者）")
                nurse_all_cell = cell_ref_rel(s["nurse_all_row"], vc_0)
                clerical_cell = cell_ref_rel(s["clerical_row"], vc_0)
                clerical_ref_cell = cell_ref_rel(s["clerical_ref_row"], vc_0)
                ws.cell(R(s["nurse_all_adj_row"]), vc).value = (
                    f"=TRUNC(MAX({nurse_all_cell}-({clerical_cell}-{clerical_ref_cell}),0),1)")
                ws.cell(R(s["nurse_all_adj_row"]), C(25), "人(実績値)")

            # ②-2 = TRUNC(C/(days*8), 1) (ward3は出力不要)
            if ward_key != "ward3":
                ws.cell(R(s["nurse_only_row"]), C(2),
                        "②-2 月平均１日当たり看護職員配置数（看護職員）※小数点第２位以下切り捨て")
                ws.cell(R(s["nurse_only_row"]), vc).value = (
                    f"=TRUNC({c_ref}/({days_cell}*8),1)")
                ws.cell(R(s["nurse_only_row"]), C(25), "人(実績値)")
                ws.cell(R(s["nurse_only_row"]), C(28), "〔Ｃ／(日数×８)〕")
                ws.cell(R(s["nurse_only_row"]), hc).value = f"={c_ref}"
                ws.cell(R(s["nurse_only_row"]), C(44), "『月延べ勤務時間数』(看護職員）")
                ws.cell(R(s["nurse_only_row"]), C(48), "時間(実績値)")

                # required_config セル (②-2の参考行: 看護職員のみの必要数)
                req_ref_row = s.get("nurse_only_ref_row")
                if req_ref_row is not None:
                    ws.cell(R(req_ref_row), C(2),
                            "（参考）1日看護職員配置数（必要数）")
                    nurse_ratio_base = wc.get("nurse_ratio_base", 0)
                    if nurse_ratio_base > 0:
                        ws.cell(R(req_ref_row), vc).value = (
                            f"=CEILING({a_cell}/{ratio_num},1)*3"
                            f"*{nurse_ratio_base/100}")
                    else:
                        ws.cell(R(req_ref_row), vc).value = (
                            f"=CEILING({a_cell}/{ratio_num},1)*3")
                    ws.cell(R(req_ref_row), C(25), "人(基準値)")
                    only_ref_cell = cell_ref_rel(req_ref_row, vc_0)
                    ws.cell(R(req_ref_row), hc).value = (
                        f"={only_ref_cell}*8*{days_cell}")
                    ws.cell(R(req_ref_row), C(44),
                            "※「1日平均看護配置数」を満たす「月延べ勤務時間数」")
                    ws.cell(R(req_ref_row), C(48), "時間(基準値)")

            # required_nurse for ward1/ward3
            nurse_ratio_base = wc.get("nurse_ratio_base", 0)

            # ③-1, ③-2 (ward3は出力不要)
            if ward_key != "ward3":
                # ③-1 = TRUNC((C/(days*8)) / required_config * 100, 1)
                ws.cell(R(s["ratio1_row"]), C(2), "③-1 看護要員中の看護職員の比率")
                req_config_expr = f"CEILING({a_cell}/{ratio_num},1)*3"
                ws.cell(R(s["ratio1_row"]), vc).value = (
                    f"=IF({req_config_expr}>0,"
                    f"TRUNC(({c_ref}/({days_cell}*8))/({req_config_expr})*100,1),0)")
                ws.cell(R(s["ratio1_row"]), C(25), "％")
                ws.cell(R(s["ratio1_row"]), C(34), "看護要員の内訳")
                ws.cell(R(s["ratio1_row"]), C(39), "看護師")
                ws.cell(R(s["ratio1_row"]), C(44), "月間総勤務時間数")
                ws.cell(R(s["ratio1_row"]), hc).value = (
                    f'=SUMPRODUCT(({nurse_z}="病棟日勤")*{nurse_c62})'
                    f'+SUMPRODUCT(({nurse_z}="病棟夜勤")*{nurse_c62})')
                ws.cell(R(s["ratio1_row"]), C(48), "時間")

                # ③-1 基準値 + 准看護師
                if nurse_ratio_base > 0:
                    ws.cell(R(s["ratio1_base_row"]), C(2),
                            "\u3000※看護要員中に必要な看護職員の比率")
                    ws.cell(R(s["ratio1_base_row"]), vc, nurse_ratio_base)
                    ws.cell(R(s["ratio1_base_row"]), C(25), "％(基準値)")
                ws.cell(R(s["ratio1_base_row"]), C(39), "准看護師")
                ws.cell(R(s["ratio1_base_row"]), C(44), "月間総勤務時間数")
                ws.cell(R(s["ratio1_base_row"]), hc).value = (
                    f'=SUMPRODUCT(({nurse_z}="病棟日勤")*{nurse_c63})'
                    f'+SUMPRODUCT(({nurse_z}="病棟夜勤")*{nurse_c63})')
                ws.cell(R(s["ratio1_base_row"]), C(48), "時間")

                # 看護補助者
                ws.cell(R(s["ratio1_base_row"] + 1), C(39), "看護補助者")
                ws.cell(R(s["ratio1_base_row"] + 1), C(44), "月間総勤務時間数")
                ws.cell(R(s["ratio1_base_row"] + 1), hc).value = f"={g_ref}"
                ws.cell(R(s["ratio1_base_row"] + 1), C(48), "時間")

                # ③-2 = TRUNC((nurse_h/(days*8)) / required_nurse * 100, 1)
                ws.cell(R(s["ratio2_row"]), C(2), "③-2 看護職員中の看護師の比率")
                nurse_h_ref = cell_ref_rel(s["ratio1_row"], hc_0)  # 看護師時間セル
                if nurse_ratio_base > 0:
                    req_nurse_expr = f"CEILING({a_cell}/{ratio_num},1)*3*{nurse_ratio_base/100}"
                    ws.cell(R(s["ratio2_row"]), vc).value = (
                        f"=IF({req_nurse_expr}>0,"
                        f"TRUNC(({nurse_h_ref}/({days_cell}*8))/({req_nurse_expr})*100,1),0)")
                else:
                    ws.cell(R(s["ratio2_row"]), vc, 0)
                ws.cell(R(s["ratio2_row"]), C(25), "％")

                # ③-2 数式説明行
                if s.get("ratio2_desc_row") is not None:
                    ws.cell(R(s["ratio2_desc_row"]), C(2),
                            "\u3000 〔月平均１日当たり看護職員配置数のうちの"
                            "看護師数／1 日看護職員配置数〕")

                # ③-2 基準値行
                if s.get("ratio2_base_row") is not None and nurse_ratio_base > 0:
                    nurse_base = wc.get("nurse_nurse_ratio_base", 20)
                    ws.cell(R(s["ratio2_base_row"]), C(2),
                            "\u3000※看護職員中に必要な看護師の比率")
                    ws.cell(R(s["ratio2_base_row"]), vc, nurse_base)
                    ws.cell(R(s["ratio2_base_row"]), C(25), "％(基準値)")
        else:
            # ward23: ② = TRUNC(C/(days*8), 1)
            ws.cell(R(s["nurse_config_row"]), C(2),
                    "②月平均１日当たり看護職員配置数 ※小数点第２位以下切り捨て")
            ws.cell(R(s["nurse_config_row"]), vc).value = (
                f"=TRUNC({c_ref}/({days_cell}*8),1)")
            ws.cell(R(s["nurse_config_row"]), C(25), "人(実績値)")
            ws.cell(R(s["nurse_config_row"]), C(28), "〔Ｃ／(日数×８)〕")
            ws.cell(R(s["nurse_config_row"]), C(44), "『月延べ勤務時間数』")
            ws.cell(R(s["nurse_config_row"]), hc).value = f"={c_ref}"
            ws.cell(R(s["nurse_config_row"]), C(48), "時間(実績値)")

            # (参考) required_config = CEILING(A/ratio, 1)*3
            ws.cell(R(s["nurse_ref_row"]), C(2),
                    "（参考）1日看護職員配置数（必要数）※小数点以下切り上げ")
            ws.cell(R(s["nurse_ref_row"]), vc).value = (
                f"=CEILING({a_cell}/{ratio_num},1)*3")
            ws.cell(R(s["nurse_ref_row"]), C(25), "人(基準値)")
            ws.cell(R(s["nurse_ref_row"]), C(28), "〔(Ａ／配置区分の数)×３〕")
            req_cell = cell_ref_rel(s["nurse_ref_row"], vc_0)
            # 基準時間 = required_config * 8 * days
            ws.cell(R(s["nurse_ref_row"]), C(44),
                    "※「1日平均看護配置数」を満たす「月延べ勤務時間数」")
            ws.cell(R(s["nurse_ref_row"]), hc).value = (
                f"={req_cell}*8*{days_cell}")
            ws.cell(R(s["nurse_ref_row"]), C(48), "時間(基準値)")

            # ③ 看護師比率 = TRUNC(nurse_h / (required*8*days) * 100, 1)
            ws.cell(R(s["ratio_row"]), C(2), "③看護職員中の看護師の比率")
            # 看護師時間 = SUMPRODUCT
            nurse_h_formula = (
                f'SUMPRODUCT(({nurse_z}="病棟日勤")*{nurse_c62})'
                f'+SUMPRODUCT(({nurse_z}="病棟夜勤")*{nurse_c62})')
            ws.cell(R(s["ratio_row"]), vc).value = (
                f"=IF({req_cell}*8*{days_cell}>0,"
                f"MIN(TRUNC(({nurse_h_formula})/({req_cell}*8*{days_cell})*100,1),100),0)")
            ws.cell(R(s["ratio_row"]), C(25), "％")
            ws.cell(R(s["ratio_row"]), C(34), "看護要員の内訳")
            ws.cell(R(s["ratio_row"]), C(39), "看護師")
            ws.cell(R(s["ratio_row"]), C(44), "月間総勤務時間数")
            ws.cell(R(s["ratio_row"]), hc).value = f"={nurse_h_formula}"
            ws.cell(R(s["ratio_row"]), C(48), "時間")

            ws.cell(R(s["ratio_sub_row"]), C(39), "准看護師")
            ws.cell(R(s["ratio_sub_row"]), C(44), "月間総勤務時間数")
            junkango_formula = (
                f'SUMPRODUCT(({nurse_z}="病棟日勤")*{nurse_c63})'
                f'+SUMPRODUCT(({nurse_z}="病棟夜勤")*{nurse_c63})')
            ws.cell(R(s["ratio_sub_row"]), hc).value = f"={junkango_formula}"
            ws.cell(R(s["ratio_sub_row"]), C(48), "時間")

            # 看護師配置数 = TRUNC(nurse_h_sum/(days*8), 1)
            ws.cell(R(s["ratio_nurse_row"]), C(2), "　 月平均1日当たり配置数")
            ws.cell(R(s["ratio_nurse_row"]), C(22), "看護師")
            nurse_h_cell = cell_ref_rel(s["ratio_row"], hc_0)
            ws.cell(R(s["ratio_nurse_row"]), vc).value = (
                f"=TRUNC({nurse_h_cell}/({days_cell}*8),1)")
            ws.cell(R(s["ratio_nurse_row"]), C(25), "人")
            ws.cell(R(s["ratio_nurse_row"]), C(39), "看護補助者")
            ws.cell(R(s["ratio_nurse_row"]), C(44), "月間総勤務時間数")
            ws.cell(R(s["ratio_nurse_row"]), hc).value = f"={g_ref}"
            ws.cell(R(s["ratio_nurse_row"]), C(48), "時間")

        # ④ 平均在院日数 (ward1/ward3は出力不要)
        if s.get("avg_stay_row") is not None and ward_key not in ("ward1", "ward3"):
            ws.cell(R(s["avg_stay_row"]), C(2),
                    "④平均在院日数 ※小数点以下切り上げ")
            ws.cell(R(s["avg_stay_row"]), C(25), "日")
            # 算出期間 (②-1と同じ)
            import datetime
            end_dt = datetime.date(self.year, self.month, 1) \
                     - datetime.timedelta(days=1)
            start_dt = datetime.date(end_dt.year - 1, end_dt.month, 1)
            asr = s["avg_stay_row"]
            ws.cell(R(asr), C(29), "(算出期間）")
            ws.cell(R(asr), C(30), start_dt.year)
            ws.cell(R(asr), C(32), "年")
            ws.cell(R(asr), C(33), start_dt.month)
            ws.cell(R(asr), C(34), "月")
            ws.cell(R(asr), C(35), start_dt.day)
            ws.cell(R(asr), C(36), "日")
            ws.cell(R(asr), C(37), "～")
            ws.cell(R(asr), C(38), end_dt.year)
            ws.cell(R(asr), C(40), "年")
            ws.cell(R(asr), C(41), end_dt.month)
            ws.cell(R(asr), C(42), "月")
            ws.cell(R(asr), C(43), end_dt.day)
            ws.cell(R(asr), C(44), "日")

        # ⑤ 夜勤時間帯
        ws.cell(R(s["night_band_row"]), C(2), "⑤夜勤時間帯(16時間)")
        ws.cell(R(s["night_band_row"]), vc, 16)
        ws.cell(R(s["night_band_row"]), C(24), "時")
        ws.cell(R(s["night_band_row"]), C(25), 30)
        ws.cell(R(s["night_band_row"]), C(26), "分")
        ws.cell(R(s["night_band_row"]), C(27), "～")
        ws.cell(R(s["night_band_row"]), C(28), 8)
        ws.cell(R(s["night_band_row"]), C(29), "時")
        ws.cell(R(s["night_band_row"]), C(30), 30)
        ws.cell(R(s["night_band_row"]), C(31), "分")

        # ⑥ = TRUNC((D-E)/B, 1) (ward1/ward3は出力不要)
        de_ref = cell_ref_rel(*refs["DE"])
        b_ref = cell_ref_rel(*refs["B"])
        if ward_key not in ("ward1", "ward3"):
            label_6 = "⑥月平均夜勤時間数"
            if layout.get("has_two_part_nurse"):
                label_6 += "（看護職員）"
            label_6 += " ※小数点第２位以下切り捨て"
            ws.cell(R(s["avg_night_row"]), C(2), label_6)
            if layout.get("has_two_part_nurse"):
                # ward3: B=nurse only → nurse合計行col23
                b_nurse_only = cell_ref_rel(nurse_total_row + 1, 23)
                ws.cell(R(s["avg_night_row"]), vc).value = (
                    f"=IF({b_nurse_only}>0,TRUNC({de_ref}/{b_nurse_only},1),0)")
            else:
                ws.cell(R(s["avg_night_row"]), vc).value = (
                    f"=IF({b_ref}>0,TRUNC({de_ref}/{b_ref},1),0)")
            ws.cell(R(s["avg_night_row"]), C(25), "時間")
            ws.cell(R(s["avg_night_row"]), C(28), "〔(Ｄ－Ｅ)／Ｂ〕")

            # （参考）看護補助者（みなしは除く）の月平均夜勤時間数
            # ward23 テンプレート: AJ64 ラベル, AT65 計算式, AW65 時間
            if not layout.get("has_two_part_nurse"):
                # ラベルは avg_night_row の1行前
                ws.cell(R(s["avg_night_row"] - 1), C(35),
                        "（参考）看護補助者（みなしは除く）の月平均夜勤時間数")
                # AT (hc) = TRUNC((I - aide_E) / aide_B, 1)
                i_ghijk = cell_ref_rel(*ghijk_refs["I"])
                aide_night_e = cell_ref_rel(aide_total_row + 1, 61)
                aide_night_b = cell_ref_rel(aide_total_row + 1, 23)
                ws.cell(R(s["avg_night_row"]), hc).value = (
                    f"=IF(ISNUMBER(TRUNC(({i_ghijk}-{aide_night_e})"
                    f"/{aide_night_b},1)),"
                    f"TRUNC(({i_ghijk}-{aide_night_e})"
                    f"/{aide_night_b},1),0)")
                ws.cell(R(s["avg_night_row"]), C(48), "時間")

        # ⑦ = L section lr+3 を参照 (テンプレート準拠) (全病棟出力不要)
        lr = layout.get("l_row", 757)
        if False:  # ⑦は全病棟で出力不要
            ws.cell(R(s["min_nurse_row"]), C(2),
                    "⑦月平均１日当たり当該入院料の施設基準の最小必要人数以上の看護職員配置数")
            l_min_cell = cell_ref_rel(lr + 3, 33)
            ws.cell(R(s["min_nurse_row"]), vc).value = f"=TRUNC({l_min_cell},1)"
            ws.cell(R(s["min_nurse_row"]), C(25), "人(実績値)")
            l_ref = cell_ref_rel(lr + 1, 33)
            ws.cell(R(s["min_nurse_row"]), C(44), "『月延べ勤務時間数』")
            ws.cell(R(s["min_nurse_row"]), hc).value = (
                f"=MAX({c_ref}-{l_ref}*{days_cell}*8,0)")
            ws.cell(R(s["min_nurse_row"]), C(48), "時間(実績値)")

            # ⑦ サブ行: ※小数点 + 基準時間
            if s.get("min_nurse_sub_row") is not None:
                ws.cell(R(s["min_nurse_sub_row"]), C(2),
                        "※小数点第２位以下切り捨て")

            # ⑦ 届出注記
            if s.get("min_nurse_note_row") is not None:
                ws.cell(R(s["min_nurse_note_row"]), C(2),
                        "≪看護職員配置加算（A308-3 地域包括ケア病棟入院料の注3）"
                        "を届け出る場合に記載≫")

            # ⑦ (参考) 最小必要数以上の看護職員配置数（必要数）
            if s.get("min_nurse_ref_row") is not None:
                ws.cell(R(s["min_nurse_ref_row"]), C(2),
                        "（参考）最小必要数以上の看護職員配置数（必要数）"
                        " ※小数点以下切り上げ")
                ws.cell(R(s["min_nurse_ref_row"]), vc).value = (
                    f"=CEILING({a_cell}/{ratio_num},1)")
                ws.cell(R(s["min_nurse_ref_row"]), C(25), "人(基準値)")
                nrb = wc.get("nurse_ratio_base", 50)
                nrb_zen = str(nrb).translate(
                    str.maketrans("0123456789", "０１２３４５６７８９"))
                ws.cell(R(s["min_nurse_ref_row"]), C(28),
                        f"〔(Ａ／{nrb_zen})×３〕")
                if s.get("min_nurse_sub_row") is not None:
                    min_ref_cell = cell_ref_rel(s["min_nurse_ref_row"], vc_0)
                    target_row = s["min_nurse_sub_row"]
                    ws.cell(R(target_row), hc).value = (
                        f"={min_ref_cell}*8*{days_cell}")
                    ws.cell(R(target_row), C(44),
                            "※「1日平均看護職員配置数」を満たす「月延べ勤務時間数」")
                    ws.cell(R(target_row), C(48), "時間(基準値)")

        # ⑧ (ward1/ward3は出力不要、ward23は出力)
        h_ref = cell_ref_rel(*ghijk_refs["H"])
        i_ref = cell_ref_rel(*ghijk_refs["I"])
        f_ref = cell_ref_rel(*refs["F"])
        if ward_key not in ("ward1", "ward3"):
            # ⑧ 看護補助者配置数
            ws.cell(R(s["aide_config_row"]), C(2),
                    "⑧月平均１日当たり看護補助者配置数 ※小数点第２位以下切り捨て")
            if layout.get("has_two_part_nurse"):
                ws.cell(R(s["aide_config_row"]), vc).value = (
                    f"=TRUNC(({g_ref}+{h_ref})/({days_cell}*8),1)")
                ws.cell(R(s["aide_config_row"]), hc).value = f"={g_ref}+{h_ref}"
            else:
                # ward23: GHIJK の (G+H)/(days*8) 行を参照 (テンプレート準拠)
                gh_haichi_cell = cell_ref_rel(*ghijk_refs["GH_haichi"])
                ws.cell(R(s["aide_config_row"]), vc).value = (
                    f"=TRUNC({gh_haichi_cell},1)")
                ws.cell(R(s["aide_config_row"]), hc).value = f"={g_ref}+{h_ref}"
            ws.cell(R(s["aide_config_row"]), C(25), "人(実績値)")
            ws.cell(R(s["aide_config_row"]), C(44), "『月延べ勤務時間数』")
            ws.cell(R(s["aide_config_row"]), C(48), "時間(実績値)")

            # ⑧調整 (⑩の実績値-⑩の上限値を減じた看護補助者配置数)
            if s.get("aide_adjusted_row") is not None:
                ws.cell(R(s["aide_adjusted_row"]), C(2),
                        "（⑩の実績値-⑩の上限値）を減じた月平均１日当たり看護補助者配置数")
                aide_8_cell = cell_ref_rel(s["aide_config_row"], vc_0)
                cler_config_ref = cell_ref_rel(s["clerical_row"], vc_0)
                cler_ref_ref = cell_ref_rel(s["clerical_ref_row"], vc_0)
                ws.cell(R(s["aide_adjusted_row"]), vc).value = (
                    f"=TRUNC(MAX({aide_8_cell}-({cler_config_ref}-{cler_ref_ref}),0),1)")
                ws.cell(R(s["aide_adjusted_row"]), C(25), "人(実績値)")

            # ⑧ 届出注記 + 基準時間
            aide_notes = s.get("aide_note_rows", [])
            aide_note_texts = [
                "≪看護補助加算・看護補助体制充実加算"
                "（A106 障害者施設等入院基本料の注9・注10）、"
                "A207-3 急性期看護補助体制加算、",
                "\u3000A214 看護補助加算、看護補助体制加算"
                "（A304 地域包括医療病棟入院料の注5）、"
                "看護補助加算・看護補助体制充実加算",
                "（A307 小児入院医療管理料の注9・注10）、"
                "看護補助者配置加算・看護補助体制充実加算"
                "（A308-3 地域包括ケア病棟入院料の",
                "\u3000注4・注5）を届け出る場合に記載≫",
            ]
            for i, nr in enumerate(aide_notes):
                if i < len(aide_note_texts):
                    ws.cell(R(nr), C(2), aide_note_texts[i])
            if aide_notes:
                aide_config_ref = cell_ref_rel(s.get("aide_ref_row",
                                                      s["aide_config_row"]), vc_0)
                ws.cell(R(aide_notes[0]), C(44),
                        "※「1日平均看護補助者配置数」を満たす「月延べ勤務時間数」")
                if s.get("aide_ref_row") is not None:
                    aide_ref_cell = cell_ref_rel(s["aide_ref_row"], vc_0)
                    ws.cell(R(aide_notes[0]), hc).value = (
                        f"={aide_ref_cell}*8*{days_cell}")
                ws.cell(R(aide_notes[0]), C(48), "時間(基準値)")

            # ⑧ (参考) 看護補助者配置数（必要数）
            if s.get("aide_ref_row") is not None:
                ws.cell(R(s["aide_ref_row"]), C(2),
                        "（参考）看護補助者配置数（必要数）")
                ws.cell(R(s["aide_ref_row"]), vc).value = (
                    f"=CEILING({a_cell}/{ratio_num},1)*3")
                ws.cell(R(s["aide_ref_row"]), C(25), "人(基準値)")
                ws.cell(R(s["aide_ref_row"]), C(28),
                        "〔(Ａ／配置区分の数)×３〕")

            # ⑧(参考) サブ行: ※小数点以下切り上げ
            if s.get("aide_ref_sub_row") is not None:
                ws.cell(R(s["aide_ref_sub_row"]), C(2),
                        " ※小数点以下切り上げ")

        else:
            # ward1/ward3: ⑧のメインは出力不要だが、
            # aide_ref_row の必要数はGHIJK[J]から参照されるため値を書く
            if s.get("aide_ref_row") is not None:
                ws.cell(R(s["aide_ref_row"]), vc).value = (
                    f"=CEILING({a_cell}/{ratio_num},1)*3")

        # ⑨ (全病棟出力不要)
        # ⑩ (全病棟出力不要)

        # ⑪ = ②の値 (全病棟出力不要)
        if False:  # ⑪は全病棟で出力不要
            ws.cell(R(s["all_staff_row"]), C(2),
                    "⑪月平均１日当たり看護職員、作業療法士、精神保健福祉士及び公認心理師配置数")
            ws.cell(R(s["all_staff_row"] + 1), C(2), "※小数点第２位以下切り捨て")
            # ⑪ = ②(nurse_config) = TRUNC(C/(days*8), 1)
            ws.cell(R(s["all_staff_row"] + 1), vc).value = (
                f"=TRUNC({c_ref}/({days_cell}*8),1)")
            ws.cell(R(s["all_staff_row"] + 1), C(25), "人(実績値)")

            # ⑪ 届出注記
            if s.get("all_staff_note_row") is not None:
                ws.cell(R(s["all_staff_note_row"]), C(2),
                        "≪A315 精神科地域包括ケア病棟入院料"
                        "を届け出る場合に記載≫")

            # ⑪ (参考) 必要配置数 (全スタッフ: 看護職員+OT+PSW+心理師)
            if s.get("all_staff_ref_row") is not None:
                all_staff_ratio = wc.get("all_staff_ratio", 13)
                ws.cell(R(s["all_staff_ref_row"]), C(2),
                        "（参考）１日看護職員、作業療法士、"
                        "精神保健福祉士及び公認心理師配置数（必要数）")
                ws.cell(R(s["all_staff_ref_row"]), vc).value = (
                    f"=CEILING({a_cell}*3/{all_staff_ratio},1)")
                ws.cell(R(s["all_staff_ref_row"]), C(25), "人(基準値)")
                ratio_zen = str(all_staff_ratio).translate(
                    str.maketrans("0123456789", "０１２３４５６７８９"))
                ws.cell(R(s["all_staff_ref_row"]), C(28),
                        f"〔(Ａ／{ratio_zen})×３〕")

            # ⑪参考 サブ行: ※小数点以下切り上げ
            if s.get("all_staff_ref_sub_row") is not None:
                ws.cell(R(s["all_staff_ref_sub_row"]), C(2),
                        " ※小数点以下切り上げ")

    # ============================================================
    # テーブルヘッダー
    # ============================================================
    def _write_table_header(self, ws, header_row, date_row, dow_row):
        """Write column headers for staff table (0-indexed rows)"""
        hdr_font = FONT_SMALL
        hdr_align = ALIGN_WRAP_CENTER

        def hdr(col, text):
            cell = ws.cell(R(header_row), C(col), text)
            cell.font = hdr_font
            cell.alignment = hdr_align

        hdr(2, "種別")
        hdr(4, "番号")
        hdr(6, "病棟名")
        hdr(8, "氏    名")
        hdr(14, "雇用・勤務形態")
        hdr(19, "看護補助者の業務")
        hdr(20, "夜勤の有無")
        hdr(23, "夜勤従事者数への計上")
        hdr(29, "日付別の勤務時間数")
        hdr(60, "月延べ勤務時間数")
        hdr(61, "再掲）月平均夜勤時間数の計算に含まない者の夜勤時間数")
        hdr(62, "職種別月\n勤務時間数")
        hdr(64, "職種別\n人数")

        # ヘッダー行高さ (wrap_text 対応)
        ws.row_dimensions[R(header_row)].height = 36

        # 日付ヘッダー
        for d in range(1, 32):
            ws.cell(R(date_row), C(29 + d - 1), f"{d}日")

        # 曜日
        dow_names = ["月", "火", "水", "木", "金", "土", "日"]
        for d in range(1, self.num_days + 1):
            wd = date(self.year, self.month, d).weekday()
            ws.cell(R(dow_row), C(29 + d - 1), dow_names[wd])

        # 追加ヘッダー
        for _c, _t in [(62, "看護師"), (63, "准看護師"),
                        (64, "看護師"), (65, "准看護師"),
                        (66, "全体"), (67, "夜専")]:
            ws.cell(R(dow_row), C(_c), _t).font = FONT_SMALL
        _cell = ws.cell(R(dow_row), C(68), "夜勤時間\nチェック")
        _cell.font = FONT_SMALL
        _cell.alignment = ALIGN_WRAP

    # ============================================================
    # 職員データ書き込み
    # ============================================================
    def _write_staff_data(self, ws, staff_list, hours_dict, all_shifts,
                          ward_names, start_row_0, is_nurse, layout):
        """
        3行/人のデータを書き込む。start_row_0 は 0-indexed。
        Returns result dict with totals.
        """
        result = {
            "day_total": [0.0] * (self.num_days + 1),
            "night_total": [0.0] * (self.num_days + 1),
            "total_night_total": [0.0] * (self.num_days + 1),
            "monthly_day_sum": 0.0,
            "monthly_night_sum": 0.0,
            "monthly_total_sum": 0.0,
            "nurse_hours_sum": 0.0,
            "junkango_hours_sum": 0.0,
            "nurse_night_sum": 0.0,
            "junkango_night_sum": 0.0,
            "nurse_count": 0,
            "junkango_count": 0,
            "jokin": 0,
            "hjokin": 0,
            "hijokin": 0,
            "night_ari": 0,
            "night_nashi": 0,
            "night_only": 0,
            "night_juuji": 0,
            "exclude_night_total": 0.0,
            "staff_nights": [],  # (emp, total_night) for dist
        }

        for slot, emp in enumerate(staff_list):
            if slot >= 100:
                break

            sid = emp["id"]
            r1 = start_row_0 + slot * 3      # Row 1: 病棟日勤
            r2 = start_row_0 + slot * 3 + 1  # Row 2: 病棟夜勤
            r3 = start_row_0 + slot * 3 + 2  # Row 3: 総夜勤

            staff_shifts = all_shifts.get(sid, {})
            hours = hours_dict.get(sid, [])

            # --- Row 1: 種別 + 番号 + 病棟名 + 氏名 + ヘッダーラベル + 病棟日勤 ---
            if is_nurse:
                stype = "看護師" if emp["type"] == "nurse" else "准看護師"
                ws.cell(R(r1), C(2), stype).font = FONT_SMALL
            # 種別セル結合 (col 2-3)
            self._safe_merge(ws, R(r1), C(2), R(r1), C(3))

            ws.cell(R(r1), C(4), slot + 1)
            # 番号セル結合 (col 4-5)
            self._safe_merge(ws, R(r1), C(4), R(r1), C(5))

            wid = emp.get("ward", "")
            wname = ward_names.get(wid, wid)
            ws.cell(R(r1), C(6), wname).font = FONT_SMALL
            # 病棟名セル結合 (col 6-7)
            self._safe_merge(ws, R(r1), C(6), R(r1), C(7))

            _name_cell = ws.cell(R(r1), C(8), emp.get("name", ""))
            _name_cell.font = FONT_SMALL
            # 氏名セル結合 (col 8-13)
            self._safe_merge(ws, R(r1), C(8), R(r1), C(13))

            # Row 1: ヘッダーラベル (text labels in col 14-17, 20-23)
            _sf = FONT_SMALL
            _sw = ALIGN_WRAP
            for _c, _t in [(14, "常勤"), (15, "短時間"),
                           (16, "非常勤"), (17, "他部署兼務"),
                           (20, "有"), (21, "無"), (22, "夜専"),
                           (23, "夜勤従事者"), (25, "病棟日勤")]:
                _cell = ws.cell(R(r1), C(_c), _t)
                _cell.font = _sf
                _cell.alignment = _sw

            if not is_nurse:
                _cell = ws.cell(R(r1), C(19), "事務的業務")
                _cell.font = _sf
                _cell.alignment = _sw

            # --- Row 2: 雇用形態フラグ (0/1) + 夜勤フラグ + 病棟夜勤 ---
            # 雇用形態 (全員常勤 = col14=1, col15=0, col16=0, col17=0)
            emp_type_flags = emp.get("employmentType", "fulltime")
            if emp_type_flags == "fulltime":
                ws.cell(R(r2), C(14), 1)
                ws.cell(R(r2), C(15), 0)
                ws.cell(R(r2), C(16), 0)
                ws.cell(R(r2), C(17), 0)
                result["jokin"] += 1
            elif emp_type_flags == "parttime":
                ws.cell(R(r2), C(14), 0)
                ws.cell(R(r2), C(15), 0)
                ws.cell(R(r2), C(16), 1)
                ws.cell(R(r2), C(17), 0)
                result["hijokin"] += 1
            else:
                ws.cell(R(r2), C(14), 1)
                ws.cell(R(r2), C(15), 0)
                ws.cell(R(r2), C(16), 0)
                ws.cell(R(r2), C(17), 0)
                result["jokin"] += 1

            # 事務的業務フラグ (看護補助者のみ)
            if not is_nurse:
                ws.cell(R(r2), C(19), 0)

            # 夜勤区分 — 夜専フラグのみ Python 固定、有/無/従事者は Excel 数式
            cat = emp.get("shiftCategory", "twoShift")
            is_night_only = (cat == "nightOnly")
            ws.cell(R(r2), C(22), 1 if is_night_only else 0)  # 夜専

            # セル参照
            _c14 = cell_ref_rel(r2, 14)  # O: 常勤
            _c15 = cell_ref_rel(r2, 15)  # P: 短時間
            _c16 = cell_ref_rel(r2, 16)  # Q: 非常勤
            _c17 = cell_ref_rel(r2, 17)  # R: 他部署兼務
            _c20 = cell_ref_rel(r2, 20)  # U: 有
            _c22 = cell_ref_rel(r2, 22)  # W: 夜専
            _c60_r2 = cell_ref_rel(r2, 60)  # BI r2: 月延べ夜勤
            _c60_r3 = cell_ref_rel(r3, 60)  # BI r3: 月延べ総夜勤

            # U (col 20, 有): 空スロット→0, BI>=8→1, else 0
            ws.cell(R(r2), C(20)).value = (
                f'=IF(AND({_c14}=0,{_c15}=0,{_c16}=0,{_c17}=0),0,'
                f'IF({_c60_r2}>=8,1,0))')

            # V (col 21, 無): 空スロット→0, U=1なら0, else 1
            ws.cell(R(r2), C(21)).value = (
                f'=IF(AND({_c14}=0,{_c15}=0,{_c16}=0,{_c17}=0),0,'
                f'IF({_c20}<>1,1,0))')

            # X (col 23, 夜勤従事者): 夜専=1 or U<>1 → 0, else MIN(1, BI_r2/BI_r3)
            ws.cell(R(r2), C(23)).value = (
                f'=IF(OR({_c22}=1,{_c20}<>1),0,IFERROR(MIN(1,{_c60_r2}/{_c60_r3}),0))')

            # Python 側カウント (集計用)
            total_night_hours = sum(h["night_h"] for h in hours)
            if is_night_only:
                result["night_only"] += 1
            elif total_night_hours >= 8:
                result["night_ari"] += 1
                result["night_juuji"] += 1
            else:
                result["night_nashi"] += 1

            ws.cell(R(r2), C(25), "病棟夜勤").font = _sf
            ws.cell(R(r2), C(25)).alignment = _sw
            ws.cell(R(r3), C(25), "総夜勤").font = _sf

            # --- 日別時間データ ---
            monthly_day = 0.0
            monthly_night = 0.0

            for h in hours:
                d = h["day"]
                col = C(29 + d - 1)  # col 29 = day 1 (0-indexed)
                day_h = h["day_h"]
                night_h = h["night_h"]

                if day_h > 0:
                    ws.cell(R(r1), col, day_h)
                    result["day_total"][d] += day_h
                if night_h > 0:
                    ws.cell(R(r2), col, night_h)
                    result["night_total"][d] += night_h
                    ws.cell(R(r3), col, night_h)  # 総夜勤 = 病棟夜勤 (単一病棟)
                    result["total_night_total"][d] += night_h

                monthly_day += day_h
                monthly_night += night_h

            monthly_total = monthly_day + monthly_night

            # 月延べ勤務時間数 (col 60) → Excel SUM数式
            last_day_col = 29 + self.num_days - 1  # 0-indexed
            day_range = f"{col_letter(29)}{r1+1}:{col_letter(last_day_col)}{r1+1}"
            night_range = f"{col_letter(29)}{r2+1}:{col_letter(last_day_col)}{r2+1}"
            total_range = f"{col_letter(29)}{r3+1}:{col_letter(last_day_col)}{r3+1}"
            ws.cell(R(r1), C(60)).value = f"=SUM({day_range})"
            ws.cell(R(r2), C(60)).value = f"=SUM({night_range})"
            ws.cell(R(r3), C(60)).value = f"=SUM({total_range})"

            # 再掲 (col 61): Excel IF数式
            # =IF(OR(col22=1, col20=0), col60_r2, 0)
            _c22 = cell_ref_rel(r2, 22)
            _c20 = cell_ref_rel(r2, 20)
            _c60_r2 = cell_ref_rel(r2, 60)
            ws.cell(R(r2), C(61)).value = f'=IF(OR({_c22}=1,{_c20}=0),{_c60_r2},0)'
            # Python側の集計も保持(夜勤分布用)
            if is_night_only or total_night_hours < 8:
                result["exclude_night_total"] += monthly_night

            # 職種別月勤務時間数 (col 62-63) → col60を参照する数式
            _c60_r1 = cell_ref_rel(r1, 60)  # 日勤月延べ
            _c60_r2 = cell_ref_rel(r2, 60)  # 夜勤月延べ
            _c60_r3 = cell_ref_rel(r3, 60)  # 総夜勤月延べ
            if is_nurse:
                if emp["type"] == "nurse":
                    ws.cell(R(r1), C(62)).value = f"={_c60_r1}"
                    ws.cell(R(r1), C(63), 0)
                    ws.cell(R(r2), C(62)).value = f"={_c60_r2}"
                    ws.cell(R(r2), C(63), 0)
                    ws.cell(R(r3), C(62)).value = f"={_c60_r3}"
                    ws.cell(R(r3), C(63), 0)
                    result["nurse_hours_sum"] += monthly_total
                    result["nurse_night_sum"] += monthly_night
                    result["nurse_count"] += 1
                else:
                    ws.cell(R(r1), C(62), 0)
                    ws.cell(R(r1), C(63)).value = f"={_c60_r1}"
                    ws.cell(R(r2), C(62), 0)
                    ws.cell(R(r2), C(63)).value = f"={_c60_r2}"
                    ws.cell(R(r3), C(62), 0)
                    ws.cell(R(r3), C(63)).value = f"={_c60_r3}"
                    result["junkango_hours_sum"] += monthly_total
                    result["junkango_night_sum"] += monthly_night
                    result["junkango_count"] += 1

            # 職種別人数 (col 64-65) - Row 2
            if is_nurse:
                if emp["type"] == "nurse":
                    ws.cell(R(r2), C(64), 1)
                    ws.cell(R(r2), C(65), 0)
                else:
                    ws.cell(R(r2), C(64), 0)
                    ws.cell(R(r2), C(65), 1)
            else:
                ws.cell(R(r2), C(64), 0)
                ws.cell(R(r2), C(65), 1)

            # 全体夜勤 (col 66) → col60 Row2参照
            ws.cell(R(r2), C(66)).value = f"={_c60_r2}"

            # 夜専 (col 67) → 夜専フラグ×col60
            _c22_r2 = cell_ref_rel(r2, 22)
            ws.cell(R(r2), C(67)).value = f'=IF({_c22_r2}=1,{_c60_r2},"-")'

            # 夜勤時間チェック (col 68) → col60 Row2参照
            ws.cell(R(r2), C(68)).value = f"={_c60_r2}"
            ws.cell(R(r1), C(68)).value = f"={_c60_r2}"
            ws.cell(R(r3), C(68)).value = f"={_c60_r3}"

            result["monthly_day_sum"] += monthly_day
            result["monthly_night_sum"] += monthly_night
            result["monthly_total_sum"] += monthly_total
            result["staff_nights"].append((emp, monthly_night))

        result["staff_count"] = min(len(staff_list), 100)
        return result

    def _write_empty_slot_labels(self, ws, start_row_0, staff_count,
                                 max_slots, is_nurse):
        """
        未使用スロット (staff_count+1 ~ max_slots) にラベルのみ書き込む。
        テンプレートでは全100スロットにラベル (番号 + 常勤/短時間/... + 病棟日勤/夜勤/総夜勤) がある。
        """
        _sf = FONT_SMALL
        _sw = ALIGN_WRAP
        for slot in range(staff_count, max_slots):
            r1 = start_row_0 + slot * 3
            r2 = r1 + 1
            r3 = r1 + 2
            ws.cell(R(r1), C(4), slot + 1)
            self._safe_merge(ws, R(r1), C(4), R(r1), C(5))
            for _c, _t in [(14, "常勤"), (15, "短時間"),
                           (16, "非常勤"), (17, "他部署兼務"),
                           (20, "有"), (21, "無"), (22, "夜専"),
                           (23, "夜勤従事者"), (25, "病棟日勤")]:
                _cell = ws.cell(R(r1), C(_c), _t)
                _cell.font = _sf
                _cell.alignment = _sw
            if not is_nurse:
                _cell = ws.cell(R(r1), C(19), "事務的業務")
                _cell.font = _sf
                _cell.alignment = _sw
            ws.cell(R(r2), C(25), "病棟夜勤").font = _sf
            ws.cell(R(r3), C(25), "総夜勤").font = _sf

    # ============================================================
    # 合計行 (テンプレートの「計」行: 3行構成)
    # ============================================================
    def _write_totals(self, ws, result, total_row_0, data_start_row):
        """
        Write totals row (0-indexed) using SUMPRODUCT formulas.
        data_start_row: 0-indexed start of staff data (100 slots × 3 rows).
        """
        ds = data_start_row  # 0-indexed
        de = data_start_row + 300 - 1  # 0-indexed, last row of 100 slots

        # Z列(col25) のラベル範囲
        z_range = f"${col_letter(25)}${ds+1}:${col_letter(25)}${de+1}"

        # Row 0: 計 + 分類ラベル + 病棟日勤
        _sf = FONT_SMALL
        _sw = ALIGN_WRAP
        ws.cell(R(total_row_0), C(4), "計")
        for _c, _t in [(14, "常勤"), (15, "短時間"),
                        (16, "非常勤"), (17, "他部署兼務"),
                        (19, "事務的業務"),
                        (20, "有"), (21, "無"), (22, "夜専"),
                        (23, "夜勤従事者")]:
            _cell = ws.cell(R(total_row_0), C(_c), _t)
            _cell.font = _sf
            _cell.alignment = _sw

        # Row 1: 分類カウント → SUMPRODUCT数式 (col25="病棟夜勤"の行=Row2)
        # (col=1)で判定: R1行にヘッダーテキストが入るため、N()ではなく=1比較を使用
        # (N()はSUMPRODUCT配列コンテキストでテキスト混在時に正しく動作しない)
        for col_idx in [14, 15, 16, 17, 19, 20, 21, 22, 23]:
            val_range = f"{col_letter(col_idx)}{ds+1}:{col_letter(col_idx)}{de+1}"
            ws.cell(R(total_row_0 + 1), C(col_idx)).value = (
                f'=SUMPRODUCT(({z_range}="病棟夜勤")*({val_range}=1))')

        # 時間データ行ラベル
        for _r, _t in [(total_row_0, "病棟日勤"),
                        (total_row_0 + 1, "病棟夜勤"),
                        (total_row_0 + 2, "総夜勤")]:
            _cell = ws.cell(R(_r), C(25), _t)
            _cell.font = _sf
            _cell.alignment = _sw

        # 日別合計 → SUMPRODUCT数式
        for d in range(1, self.num_days + 1):
            col_d = 29 + d - 1  # 0-indexed
            d_range = f"{col_letter(col_d)}{ds+1}:{col_letter(col_d)}{de+1}"
            ws.cell(R(total_row_0), C(col_d)).value = (
                f'=SUMPRODUCT(({z_range}="病棟日勤")*{d_range})')
            ws.cell(R(total_row_0 + 1), C(col_d)).value = (
                f'=SUMPRODUCT(({z_range}="病棟夜勤")*{d_range})')
            ws.cell(R(total_row_0 + 2), C(col_d)).value = (
                f'=SUMPRODUCT(({z_range}="総夜勤")*{d_range})')

        # 月延べ合計 → SUMPRODUCT数式 (col 60, 61)
        for cc in [60, 61]:
            cc_range = f"{col_letter(cc)}{ds+1}:{col_letter(cc)}{de+1}"
            ws.cell(R(total_row_0), C(cc)).value = (
                f'=SUMPRODUCT(({z_range}="病棟日勤")*{cc_range})')
            ws.cell(R(total_row_0 + 1), C(cc)).value = (
                f'=SUMPRODUCT(({z_range}="病棟夜勤")*{cc_range})')
            ws.cell(R(total_row_0 + 2), C(cc)).value = (
                f'=SUMPRODUCT(({z_range}="総夜勤")*{cc_range})')

    # ============================================================
    # その他職員表 (ward1, ward23のみ、テンプレートに存在)
    # ============================================================
    def _write_other_staff_table(self, ws, layout):
        """その他職員表（作業療法士、精神保健福祉士等）のヘッダーと空枠を出力"""
        osr = layout.get("other_staff_label_row")
        if osr is None:
            return

        ws.cell(R(osr), C(2), "《その他職員表》")
        # ヘッダー行 (column headers)
        hr = layout["other_staff_header_row"]
        hdr_font = FONT_SMALL
        hdr_align = ALIGN_WRAP_CENTER

        def ohdr(col, text):
            cell = ws.cell(R(hr), C(col), text)
            cell.font = hdr_font
            cell.alignment = hdr_align

        ohdr(4, "番号")
        ohdr(6, "病棟名")
        ohdr(8, "氏    名")
        ohdr(14, "雇用・勤務形態")
        ohdr(19, "看護補助者の業務")
        ohdr(20, "夜勤の有無")
        ohdr(23, "夜勤従事者数への計上")
        ohdr(29, "日付別の勤務時間数")
        ohdr(60, "月延べ勤務時間数")
        ohdr(61, "再掲）月平均夜勤時間数の計算に含まない者の夜勤時間数")
        ws.row_dimensions[R(hr)].height = 36

        # 日付ヘッダー
        dr = layout["other_staff_date_row"]
        for d in range(1, 32):
            ws.cell(R(dr), C(29 + d - 1), f"{d}日")

        # 曜日（テンプレートは「　　曜」形式）
        dw = layout["other_staff_dow_row"]
        for d in range(1, self.num_days + 1):
            ws.cell(R(dw), C(29 + d - 1), "\u3000\u3000曜")

        # データ部分は空（現在、その他職員のデータは管理していない）
        # 1スロット分 (R729-R731 for ward1) のラベルのみ書く
        _sf = FONT_SMALL
        _sw = ALIGN_WRAP
        dsr = layout["other_staff_data_start_row"]
        ws.cell(R(dsr), C(4), 1)  # スロット番号
        self._safe_merge(ws, R(dsr), C(4), R(dsr), C(5))
        for _c, _t in [(14, "常勤"), (15, "短時間"),
                        (16, "非常勤"), (17, "他部署兼務"),
                        (20, "有"), (21, "無"), (22, "夜専"),
                        (23, "夜勤従事者"), (25, "病棟日勤")]:
            _cell = ws.cell(R(dsr), C(_c), _t)
            _cell.font = _sf
            _cell.alignment = _sw
        ws.cell(R(dsr + 1), C(25), "病棟夜勤").font = _sf
        ws.cell(R(dsr + 1), C(25)).alignment = _sw
        ws.cell(R(dsr + 2), C(25), "総夜勤").font = _sf

        # 合計行 (計)
        total_row = dsr + 3
        ws.cell(R(total_row), C(4), "計")
        for _c, _t in [(14, "常勤"), (15, "短時間"),
                        (16, "非常勤"), (17, "他部署兼務"),
                        (19, "事務的業務"),
                        (20, "有"), (21, "無"), (22, "夜専"),
                        (23, "夜勤従事者"), (25, "病棟日勤")]:
            _cell = ws.cell(R(total_row), C(_c), _t)
            _cell.font = _sf
            _cell.alignment = _sw
        # 合計値は全て0
        for cc in [14, 15, 16, 17, 19, 20, 21, 22, 23]:
            ws.cell(R(total_row + 1), C(cc), 0)
        ws.cell(R(total_row + 1), C(25), "病棟夜勤").font = _sf
        ws.cell(R(total_row + 1), C(25)).alignment = _sw
        ws.cell(R(total_row + 2), C(25), "総夜勤").font = _sf
        # 日別合計 = 全て0 (3行×日数 + col60, col61)
        for dr in range(3):
            for d in range(1, self.num_days + 1):
                ws.cell(R(total_row + dr), C(29 + d - 1), 0)
            ws.cell(R(total_row + dr), C(60), 0)
            ws.cell(R(total_row + dr), C(61), 0)

    # ============================================================
    # 最終集計セクション (テーブル下部)
    # ============================================================
    def _write_final_summary(self, ws, row_0, layout, wc,
                             nurse_result, aide_result, stats,
                             nurse_total_row, aide_total_row,
                             nurse_data_start, days_cell):
        """
        テーブル下のB/C/D/E/F集計 → Excel数式。
        nurse_total_row, aide_total_row: 0-indexed total row positions.
        nurse_data_start: 0-indexed nurse data start row.
        days_cell: 稼働日数セルの参照文字列.
        Returns refs dict with cell positions for downstream formulas.
        """
        refs = {}
        _wrap = ALIGN_WRAP

        # [B] 夜勤従事者数
        ws.cell(R(row_0), C(2),
                "夜勤従事者数(夜勤ありの職員数)〔Ｂ〕").alignment = _wrap
        b_nurse = cell_ref_rel(nurse_total_row + 1, 23)
        # B = 看護職員のみ（全ward共通、テンプレート準拠）
        ws.cell(R(row_0), C(18)).value = f"={b_nurse}"
        refs["B"] = (row_0, 18)

        # [C] 月延べ勤務時間数 = nurse日勤月延べ + nurse夜勤月延べ
        ws.cell(R(row_0), C(20),
                "月延べ勤務時間数〔Ｃ〕\n（上段と中段の計）").alignment = _wrap
        c_day = cell_ref_rel(nurse_total_row, 60)
        c_night = cell_ref_rel(nurse_total_row + 1, 60)
        ws.cell(R(row_0), C(33)).value = f"={c_day}+{c_night}"
        refs["C"] = (row_0, 33)

        # [D] 月延べ夜勤時間数 = nurse夜勤月延べ
        ws.cell(R(row_0 + 1), C(20),
                "月延べ夜勤時間数〔Ｄ〕\n（中段の計）").alignment = _wrap
        ws.cell(R(row_0 + 1), C(33)).value = f"={c_night}"
        refs["D"] = (row_0 + 1, 33)

        # [E] = SUMPRODUCT((nurse z_range="病棟夜勤") * nurse col61_range)
        ws.cell(R(row_0 + 1), C(35),
                "月平均夜勤時間数の計算に含まない者の夜勤時間数〔E〕").alignment = _wrap
        nds = nurse_data_start
        nde = nurse_data_start + 300 - 1
        nurse_z = f"${col_letter(25)}${nds+1}:${col_letter(25)}${nde+1}"
        nurse_c61 = f"{col_letter(61)}{nds+1}:{col_letter(61)}{nde+1}"
        ws.cell(R(row_0 + 1), C(46)).value = (
            f'=SUMPRODUCT(({nurse_z}="病棟夜勤")*{nurse_c61})')
        refs["E"] = (row_0 + 1, 46)

        # [D-E]
        ws.cell(R(row_0 + 1), C(2),
                "月延べ夜勤時間数\u3000〔D－E〕").alignment = _wrap
        d_ref = cell_ref_rel(*refs["D"])
        e_ref = cell_ref_rel(*refs["E"])
        ws.cell(R(row_0 + 1), C(18)).value = f"={d_ref}-{e_ref}"
        refs["DE"] = (row_0 + 1, 18)

        # [F] 事務的業務 (現在0、将来対応)
        ws.cell(R(row_0 + 2), C(2),
                "（再掲） 主として事務的業務を行う"
                "看護補助者の月延べ勤務時間数の計 〔Ｆ〕").alignment = _wrap
        refs["F"] = (row_0 + 2, 33)

        # 1日看護配置数（必要数）+ 月平均配置数 (同一行 row_0+3)
        # テンプレートではsummary sectionの参考行セルを参照
        s = layout["summary"]
        a_cell = cell_ref_rel(s["patients_row"], layout["summary_value_col"])
        ws.cell(R(row_0 + 3), C(2),
                "1日看護配置数（必要数）\n"
                "\u3000〔(A ／配置区分の数)×３〕").alignment = _wrap
        if layout.get("has_two_part_nurse"):
            # ward1/ward3: nurse_only_ref_row (②-2参考) を参照
            ref_row = s.get("nurse_only_ref_row")
        else:
            # ward23: nurse_ref_row (②参考) を参照
            ref_row = s.get("nurse_ref_row")
        if ref_row is not None:
            req_ref_cell = cell_ref_rel(ref_row, layout["summary_value_col"])
            ws.cell(R(row_0 + 3), C(18)).value = f"={req_ref_cell}"
        else:
            ratio_str = wc.get("ratio", "15対1")
            try:
                ratio_num_val = int(ratio_str.split("対")[0]) if isinstance(ratio_str, str) else ratio_str
            except (ValueError, IndexError):
                ratio_num_val = 15
            ws.cell(R(row_0 + 3), C(18)).value = (
                f"=CEILING({a_cell}/{ratio_num_val},1)*3")

        c_ref = cell_ref_rel(*refs["C"])
        ws.cell(R(row_0 + 3), C(20),
                "月平均１日当たり看護配置数\n"
                "\u3000〔C／(日数×８）〕").alignment = _wrap
        ws.cell(R(row_0 + 3), C(33)).value = (
            f"=TRUNC({c_ref}/({days_cell}*8),1)")
        refs["nurse_config_final"] = (row_0 + 3, 33)

        # 事務的業務上限 = TRUNC((A/200)*3, 3)
        ws.cell(R(row_0 + 4), C(2),
                "主として事務的業務を行う"
                "看護補助者配置数（上限）\u3000\n"
                "\u3000〔(A ／２００)×３〕").alignment = _wrap
        ws.cell(R(row_0 + 4), C(18)).value = f"=TRUNC(({a_cell}/200)*3,3)"
        refs["clerical_ref"] = (row_0 + 4, 18)

        # 事務的配置数 = TRUNC(F/(days*8), 3)
        ws.cell(R(row_0 + 4), C(20),
                "月平均１日当たりの主として事務的\n"
                "業務を行う看護補助配置数\u3000"
                "〔Ｆ／(日数×８）〕").alignment = _wrap
        f_ref = cell_ref_rel(*refs["F"])
        ws.cell(R(row_0 + 4), C(33)).value = (
            f"=TRUNC({f_ref}/({days_cell}*8),3)")
        refs["clerical_config"] = (row_0 + 4, 33)

        # 注記行
        ws.cell(R(row_0 + 5), C(2),
                "注１）１日看護配置数\u3000≦\u3000"
                "月平均１日当り看護配置数")
        ws.cell(R(row_0 + 6), C(2),
                "注２）主として事務的業務を行う看護補助者配置数\u3000≧\u3000"
                "月平均１日当りの主として事務的業務を行う"
                "看護補助者配置数")

        return refs

    # ============================================================
    # 勤務体制
    # ============================================================
    def _write_schedule(self, ws, layout):
        sr = layout.get("schedule_row", 738)
        sched = self.config.get("shift_schedules", {})

        def _time_bracket(row, sc, sh, sm, eh, em):
            """時間帯の括弧装飾を書き込む: （HH：MM～HH：MM）"""
            ws.cell(R(row), C(sc), "（")
            ws.cell(R(row), C(sc + 1), sh)
            ws.cell(R(row), C(sc + 2), "：")
            ws.cell(R(row), C(sc + 3), sm)
            ws.cell(R(row), C(sc + 4), "～")
            ws.cell(R(row), C(sc + 5), eh)
            ws.cell(R(row), C(sc + 6), "：")
            ws.cell(R(row), C(sc + 7), em)
            ws.cell(R(row), C(sc + 8), "）")

        # 勤務体制ラベル (schedule_row の1行上)
        ws.cell(R(sr - 1), C(2), "勤務体制")

        def _sched_merge(row):
            """勤務体制行の共通セル結合"""
            self._safe_merge(ws, R(row), C(2), R(row), C(5))
            self._safe_merge(ws, R(row), C(6), R(row), C(8))
            self._safe_merge(ws, R(row), C(21), R(row), C(23))
            self._safe_merge(ws, R(row), C(36), R(row), C(38))

        # 3交代制
        ws.cell(R(sr), C(2), "３交代制")
        ws.cell(R(sr), C(6), "日勤")
        _sched_merge(sr)
        k3 = sched.get("3kohtai", {})
        d3 = k3.get("day", {})
        sh, sm = self._parse_time(d3.get("start", "")) if d3 else (None, None)
        eh, em = self._parse_time(d3.get("end", "")) if d3 else (None, None)
        _time_bracket(sr, 9, sh, sm, eh, em)
        ws.cell(R(sr), C(21), "準夜勤")
        j3 = k3.get("junnya", {})
        sh, sm = self._parse_time(j3.get("start", "")) if j3 else (None, None)
        eh, em = self._parse_time(j3.get("end", "")) if j3 else (None, None)
        _time_bracket(sr, 24, sh, sm, eh, em)
        ws.cell(R(sr), C(36), "深夜勤")
        s3 = k3.get("shinya", {})
        sh, sm = self._parse_time(s3.get("start", "")) if s3 else (None, None)
        eh, em = self._parse_time(s3.get("end", "")) if s3 else (None, None)
        _time_bracket(sr, 39, sh, sm, eh, em)

        # 2交代制
        ws.cell(R(sr + 1), C(2), "２交代制")
        ws.cell(R(sr + 1), C(6), "日勤")
        _sched_merge(sr + 1)
        k2 = sched.get("2kohtai", {})
        d2 = k2.get("day", {})
        sh, sm = self._parse_time(d2.get("start", "")) if d2 else (None, None)
        eh, em = self._parse_time(d2.get("end", "")) if d2 else (None, None)
        _time_bracket(sr + 1, 9, sh, sm, eh, em)
        ws.cell(R(sr + 1), C(21), "夜勤")
        n2 = k2.get("night2", {})
        sh, sm = self._parse_time(n2.get("start", "")) if n2 else (None, None)
        eh, em = self._parse_time(n2.get("end", "")) if n2 else (None, None)
        _time_bracket(sr + 1, 24, sh, sm, eh, em)

        # その他
        ws.cell(R(sr + 2), C(2), "その他")
        ws.cell(R(sr + 2), C(6), "日勤")
        _sched_merge(sr + 2)
        _time_bracket(sr + 2, 9, None, None, None, None)
        _time_bracket(sr + 2, 24, None, None, None, None)
        _time_bracket(sr + 2, 39, None, None, None, None)

        # 申し送り時間
        mr = layout.get("moushiokuri_row", sr + 3)
        ws.cell(R(mr), C(2), "申し送り時間")
        self._safe_merge(ws, R(mr), C(2), R(mr), C(50))

        # 申し送り時間: 3交代制
        ws.cell(R(mr + 1), C(2), "３交代制")
        ws.cell(R(mr + 1), C(6), "日勤")
        _sched_merge(mr + 1)
        _time_bracket(mr + 1, 9, None, None, None, None)
        ws.cell(R(mr + 1), C(21), "準夜勤")
        _time_bracket(mr + 1, 24, None, None, None, None)
        ws.cell(R(mr + 1), C(36), "深夜勤")
        _time_bracket(mr + 1, 39, None, None, None, None)

        # 申し送り時間: 2交代制
        ws.cell(R(mr + 2), C(2), "２交代制")
        ws.cell(R(mr + 2), C(6), "日勤")
        _sched_merge(mr + 2)
        _time_bracket(mr + 2, 9, None, None, None, None)
        ws.cell(R(mr + 2), C(21), "夜勤")
        _time_bracket(mr + 2, 24, None, None, None, None)

        # 申し送り時間: その他
        ws.cell(R(mr + 3), C(2), "その他")
        ws.cell(R(mr + 3), C(6), "日勤")
        _sched_merge(mr + 3)
        _time_bracket(mr + 3, 9, None, None, None, None)
        _time_bracket(mr + 3, 24, None, None, None, None)
        _time_bracket(mr + 3, 39, None, None, None, None)

    def _parse_time(self, time_str):
        """Parse "HH:MM" to (hour, minute)"""
        if not time_str:
            return (0, 0)
        parts = time_str.split(":")
        return (int(parts[0]), int(parts[1])) if len(parts) == 2 else (0, 0)

    # ============================================================
    # [G][H][I][J][K] セクション
    # ============================================================
    def _write_ghijk(self, ws, layout, stats, refs,
                     aide_total_row, days_cell):
        """
        [G][H][I][J][K] セクション → Excel数式。
        refs: _write_final_summary() から返されたセル位置dict。
        aide_total_row: 0-indexed aide total row.
        days_cell: 稼働日数セル参照文字列.
        Returns ghijk_refs dict with cell positions.
        """
        gr = layout.get("ghijk_row", 746)
        vc = C(layout["ghijk_value_col"])  # col 34 (0-indexed 33)
        vc_0 = layout["ghijk_value_col"]   # 0-indexed for cell_ref

        ghijk_refs = {}

        _wrap = ALIGN_WRAP

        ws.cell(R(gr), C(2),
                "〔急性期看護補助体制加算・看護補助加算等を届け出る場合の看護補助者の算出方法〕").alignment = _wrap

        # [G] = aide日勤月延べ + aide夜勤月延べ
        ws.cell(R(gr + 1), C(2),
                "看護補助者のみの月延べ勤務時間数の計〔G〕").alignment = _wrap
        a_day = cell_ref_rel(aide_total_row, 60)
        a_night = cell_ref_rel(aide_total_row + 1, 60)
        ws.cell(R(gr + 1), vc).value = f"={a_day}+{a_night}"
        ghijk_refs["G"] = (gr + 1, vc_0)

        # [H] = MAX(C - required_nurse*8*days, 0)
        ws.cell(R(gr + 2), C(2),
                "みなし看護補助者の月延べ勤務時間数の計〔H〕\u3000").alignment = _wrap
        ws.cell(R(gr + 2), C(19),
                "〔C 〕- 〔1日看護配置数×８×日数〕").alignment = _wrap
        c_ref = cell_ref_rel(*refs["C"])
        # required_nurse: ward1/ward3 = required_config * (nurse_ratio_base/100)
        #                 ward23 = required_config
        if layout.get("has_two_part_nurse"):
            nurse_ratio_base = stats.get("nurse_ratio_base", 0)
            if nurse_ratio_base > 0:
                # H = MAX(C - CEILING(A/r,1)*3*(base/100)*8*days, 0)
                a_cell = cell_ref_rel(layout["summary"]["patients_row"],
                                      layout["summary_value_col"])
                ratio_num = stats["ratio_num"]
                ws.cell(R(gr + 2), vc).value = (
                    f"=MAX({c_ref}-CEILING({a_cell}/{ratio_num},1)*3"
                    f"*{nurse_ratio_base/100}*8*{days_cell},0)")
            else:
                ws.cell(R(gr + 2), vc, 0)
        else:
            # ward23: H = MAX(C - CEILING(A/r,1)*3*8*days, 0)
            a_cell = cell_ref_rel(layout["summary"]["patients_row"],
                                  layout["summary_value_col"])
            ratio_num = stats["ratio_num"]
            ws.cell(R(gr + 2), vc).value = (
                f"=MAX({c_ref}-CEILING({a_cell}/{ratio_num},1)*3*8*{days_cell},0)")
        ghijk_refs["H"] = (gr + 2, vc_0)

        # [I] = aide夜勤月延べ
        ws.cell(R(gr + 3), C(2),
                "看護補助者のみの月延べ夜勤時間数の計〔I〕 ").alignment = _wrap
        ws.cell(R(gr + 3), C(19),
                "看護補助者（みなしを除く）のみの〔D〕").alignment = _wrap
        ws.cell(R(gr + 3), vc).value = f"={a_night}"
        ghijk_refs["I"] = (gr + 3, vc_0)

        # [J] = aide_ref_row を参照 (テンプレート準拠)
        ws.cell(R(gr + 4), C(2),
                "1日看護補助配置数（必要数） 〔J〕").alignment = _wrap
        ws.cell(R(gr + 4), C(19),
                "\u3000〔(A ／配置区分の数)×３〕").alignment = _wrap
        s_ref = layout["summary"]
        ratio_num = stats["ratio_num"]
        aide_ref_row = s_ref.get("aide_ref_row")
        if aide_ref_row is not None:
            aide_ref_cell = cell_ref_rel(aide_ref_row, layout["summary_value_col"])
            ws.cell(R(gr + 4), vc).value = f"={aide_ref_cell}"
        else:
            a_cell = cell_ref_rel(s_ref["patients_row"], layout["summary_value_col"])
            ws.cell(R(gr + 4), vc).value = f"=CEILING({a_cell}/{ratio_num},1)*3"
        ws.cell(R(gr + 4), C(35), "（基準値）")
        ghijk_refs["J"] = (gr + 4, vc_0)

        # (G+H) / (日数×8) = TRUNC((G+H)/(days*8), 1)
        ws.cell(R(gr + 5), C(2),
                "月平均1日当たり看護補助者"
                "配置数(みなし看護補助者含む）\u3000").alignment = _wrap
        ws.cell(R(gr + 5), C(19),
                "〔(Ｇ+H) ／ (日数×８) 〕").alignment = _wrap
        g_ref = cell_ref_rel(*ghijk_refs["G"])
        h_ref = cell_ref_rel(*ghijk_refs["H"])
        ws.cell(R(gr + 5), vc).value = (
            f"=TRUNC(({g_ref}+{h_ref})/({days_cell}*8),1)")
        ws.cell(R(gr + 5), C(35), "（実績値）")
        ghijk_refs["GH_haichi"] = (gr + 5, vc_0)

        # [K] = TRUNC(G/(days*8), 1)
        ws.cell(R(gr + 6), C(2),
                "月平均1日当たり看護補助者配置数（みなし看護補助者除く）〔K〕\u3000").alignment = _wrap
        ws.cell(R(gr + 6), C(19),
                "\u3000〔G ／ (日数×８) 〕").alignment = _wrap
        ws.cell(R(gr + 6), vc).value = (
            f"=TRUNC({g_ref}/({days_cell}*8),1)")
        ghijk_refs["K"] = (gr + 6, vc_0)

        # 夜間必要数 = CEILING(A/配置区分, 1)
        ws.cell(R(gr + 7), C(2),
                "夜間看護補助配置数（必要数）").alignment = _wrap
        ws.cell(R(gr + 7), C(19),
                "〔Ａ／配置区分の数〕").alignment = _wrap
        a_cell = cell_ref_rel(layout["summary"]["patients_row"],
                              layout["summary_value_col"])
        ws.cell(R(gr + 7), vc).value = f"=CEILING({a_cell}/{ratio_num},1)"
        ws.cell(R(gr + 7), C(35), "（基準値）")

        # 夜間実績 = TRUNC(I/(days*16), 1)
        ws.cell(R(gr + 8), C(2),
                "月平均1日当たり夜間看護補助者配置数").alignment = _wrap
        ws.cell(R(gr + 8), C(19),
                "（Ｉ／(日数×１６）〕").alignment = _wrap
        i_ref = cell_ref_rel(*ghijk_refs["I"])
        ws.cell(R(gr + 8), vc).value = (
            f"=TRUNC({i_ref}/({days_cell}*16),1)")
        ws.cell(R(gr + 8), C(35), "（実績値）")

        # K/J ratio = TRUNC((K/J)*100, 3)
        ws.cell(R(gr + 9), C(2),
                "看護補助者（みなし看護補助者を含む）の最小必要数"
                "に対する看護補助者（みなし看護補助者を除く）"
                "の割合（％）").alignment = _wrap
        ws.cell(R(gr + 9), C(19),
                " 〔（Ｋ／Ｊ）×１００〕").alignment = _wrap
        k_ref = cell_ref_rel(*ghijk_refs["K"])
        j_ref = cell_ref_rel(*ghijk_refs["J"])
        ws.cell(R(gr + 9), vc).value = (
            f"=IF({j_ref}>0,TRUNC(({k_ref}/{j_ref})*100,3),0)")
        ws.cell(R(gr + 9), C(35), "（実績値）")

        return ghijk_refs

    # ============================================================
    # [L] セクション
    # ============================================================
    def _write_l_section(self, ws, layout, stats, refs, days_cell):
        """
        [L] セクション → Excel数式。
        refs: _write_final_summary() から返されたセル位置dict。
        days_cell: 稼働日数セル参照文字列.
        """
        lr = layout.get("l_row", 757)
        a_cell = cell_ref_rel(layout["summary"]["patients_row"],
                              layout["summary_value_col"])
        c_ref = cell_ref_rel(*refs["C"])

        _wrap = ALIGN_WRAP
        ws.cell(R(lr), C(2),
                "〔看護職員配置加算（地域包括ケア病棟入院料の注３）を届け出る場合の看護職員数の算出方法〕").alignment = _wrap

        # [L] = summary section の参考行を参照
        # テンプレートでは =X60 (nurse_only_ref_row) を参照
        s = layout["summary"]
        if layout.get("has_two_part_nurse"):
            l_src_row = s.get("nurse_only_ref_row")
        else:
            l_src_row = s.get("nurse_ref_row")
        ws.cell(R(lr + 1), C(2),
                "1日看護職員配置数（必要数）〔Ｌ〕").alignment = _wrap
        ws.cell(R(lr + 1), C(19),
                "\u3000〔（A ／配置区分の数）×３〕")
        if l_src_row is not None:
            l_src_cell = cell_ref_rel(l_src_row, layout["summary_value_col"])
            ws.cell(R(lr + 1), C(33)).value = f"={l_src_cell}"
        else:
            l_div = 13
            ws.cell(R(lr + 1), C(33)).value = f"=CEILING({a_cell}/{l_div},1)*3"
        ws.cell(R(lr + 1), C(35), "（基準値）")
        l_ref = cell_ref_rel(lr + 1, 33)

        # 月平均配置数 = TRUNC(C/(days*8), 1)
        ws.cell(R(lr + 2), C(2),
                "月平均１日当たり看護職員配置数").alignment = _wrap
        ws.cell(R(lr + 2), C(19),
                "〔C／（日数×８ （時間））〕 ")
        ws.cell(R(lr + 2), C(33)).value = f"=TRUNC({c_ref}/({days_cell}*8),1)"
        ws.cell(R(lr + 2), C(35), "（実績値）")

        # 最小必要人数以上 = TRUNC(MAX(C - L*days*8, 0)/(days*8), 1)
        ws.cell(R(lr + 3), C(2),
                "月平均１日当たり当該入院料の施設基準の"
                "最小必要人数以上の看護職員配置数").alignment = _wrap
        ws.cell(R(lr + 3), C(19),
                "\u3000〔｛Ｃ - (Ｌ×日数×８ （時間））｝"
                "／（日数×８ （時間））〕").alignment = _wrap
        ws.cell(R(lr + 3), C(33)).value = (
            f"=TRUNC(MAX({c_ref}-{l_ref}*{days_cell}*8,0)/({days_cell}*8),1)")
        ws.cell(R(lr + 3), C(35), "（実績値）")

    # ============================================================
    # 夜勤時間分布
    # ============================================================
    def _write_night_distribution(self, ws, layout, nurses, aides,
                                  nurse_hours, aide_hours):
        dr = layout.get("dist_row", 764)

        ws.cell(R(dr), C(3), "参考")
        ws.cell(R(dr), C(4),
                "月総夜勤時間別の看護職員数（夜勤従事者数）の分布　　…「入院基本料等に関する実施状況報告書」用")
        ws.cell(R(dr + 1), C(4),
                "※夜勤専従者数は再掲として（　）に人数を記入のこと")

        # バンド定義 (col 6,8,10,12,14,16,18,20,22,24,26,28)
        bands = [
            (0, 16), (16, 32), (32, 48), (48, 64), (64, 72), (72, 80),
            (80, 88), (88, 96), (96, 112), (112, 144), (144, 152), (152, 99999)
        ]
        band_labels = [
            "16時間未満", "16～32時間", "32～48時間", "48～64時間",
            "64～72時間", "72～80時間", "80～88時間", "88～96時間",
            "96～112時間", "112～144時間", "144～152時間", "152時間以上"
        ]
        band_cols = [6, 8, 10, 12, 14, 16, 18, 20, 22, 24, 26, 28]

        def calc_dist(emp_list, hours_dict):
            """Calculate night hour distribution, returning (counts, night_only_counts)"""
            counts = [0] * len(bands)
            yasen_counts = [0] * len(bands)
            for emp in emp_list:
                sid = emp["id"]
                hours = hours_dict.get(sid, [])
                total = sum(h["night_h"] for h in hours)
                if total <= 0:
                    continue
                cat = emp.get("shiftCategory", "twoShift")
                is_yasen = (cat == "nightOnly")
                for i, (lo, hi) in enumerate(bands):
                    if lo <= total < hi:
                        counts[i] += 1
                        if is_yasen:
                            yasen_counts[i] += 1
                        break
            return counts, yasen_counts

        def write_dist_section(label, start_row, emp_list, hours_dict):
            ws.cell(R(start_row), C(3), label)
            ws.cell(R(start_row + 1), C(3), "夜勤時間数")
            for i, lbl in enumerate(band_labels):
                _cell = ws.cell(R(start_row + 1), C(band_cols[i]), lbl)
                _cell.font = FONT_SMALL
                _cell.alignment = ALIGN_WRAP
            ws.cell(R(start_row + 2), C(3), "夜勤従事者数")
            counts, yasen = calc_dist(emp_list, hours_dict)
            for i, bc in enumerate(band_cols):
                ws.cell(R(start_row + 2), C(bc), counts[i])
                # 夜専は隣の列(odd col)
                ws.cell(R(start_row + 2), C(bc + 1), yasen[i])
            return counts, yasen

        # 【看護職員】
        nc, ny = write_dist_section("【看護職員】", dr + 3, nurses, nurse_hours)
        # 【看護補助者】
        ac, ay = write_dist_section("【看護補助者】", dr + 7, aides, aide_hours)
        # 【看護職員＋看護補助者】
        ws.cell(R(dr + 11), C(3), "【看護職員＋看護補助者】")
        ws.cell(R(dr + 12), C(3), "夜勤時間数")
        for i, lbl in enumerate(band_labels):
            _cell = ws.cell(R(dr + 12), C(band_cols[i]), lbl)
            _cell.font = FONT_SMALL
            _cell.alignment = ALIGN_WRAP
        ws.cell(R(dr + 13), C(3), "夜勤従事者数")
        for i, bc in enumerate(band_cols):
            ws.cell(R(dr + 13), C(bc), nc[i] + ac[i])
            ws.cell(R(dr + 13), C(bc + 1), ny[i] + ay[i])

    # ============================================================
    # 未使用行の非表示
    # ============================================================
    def _hide_unused_rows(self, ws, data_start_row, staff_count, max_slots):
        """
        職員表の未使用スロット行を非表示にする。
        data_start_row: 0-indexed データ開始行
        staff_count: 実際のスタッフ数
        max_slots: 最大スロット数 (100)
        """
        if staff_count >= max_slots:
            return
        # 未使用スロットの開始 (0-indexed)
        unused_start = data_start_row + staff_count * 3
        # 合計行の直前まで (合計行 = data_start + max_slots * 3)
        unused_end = data_start_row + max_slots * 3 - 1
        for row_0 in range(unused_start, unused_end + 1):
            ws.row_dimensions[R(row_0)].hidden = True

    # ============================================================
    # 列幅・セル結合・書式
    # ============================================================
    def _apply_column_widths(self, ws):
        """テンプレート準拠の列幅を設定"""
        # 基本幅: 全列 3.44 (テンプレート標準)
        for col in range(1, 70):
            ws.column_dimensions[get_column_letter(col)].width = 3.44

        # 特定列は幅広 (月延べ、再掲、職種別等)
        ws.column_dimensions[get_column_letter(61)].width = 7.0    # 月延べ勤務時間数
        ws.column_dimensions[get_column_letter(62)].width = 7.11   # 再掲
        ws.column_dimensions[get_column_letter(63)].width = 5.0    # 職種別
        ws.column_dimensions[get_column_letter(64)].width = 5.0    # 職種別
        ws.column_dimensions[get_column_letter(65)].width = 6.11   # 職種別人数
        ws.column_dimensions[get_column_letter(66)].width = 6.11
        ws.column_dimensions[get_column_letter(67)].width = 6.11
        ws.column_dimensions[get_column_letter(68)].width = 6.11
        ws.column_dimensions[get_column_letter(69)].width = 6.11

    def _safe_merge(self, ws, r1, c1, r2, c2):
        """安全なセル結合 (単一セル・完全重複はスキップ)"""
        if r1 == r2 and c1 == c2:
            return  # 単一セル結合は不要
        # 完全一致の重複チェック (同じ範囲の二重登録を防止)
        for mr in ws.merged_cells.ranges:
            if (mr.min_row == r1 and mr.max_row == r2 and
                    mr.min_col == c1 and mr.max_col == c2):
                return  # 既に同じ範囲が結合済み
        try:
            ws.merge_cells(start_row=r1, start_column=c1,
                           end_row=r2, end_column=c2)
        except Exception as e:
            logger.warning("セル結合失敗 (%d,%d)-(%d,%d): %s", r1, c1, r2, c2, e)

    def _apply_merges(self, ws, layout):
        """セル結合を一括適用 (ラベル・値・時間セル)"""
        s = layout["summary"]
        vc_0 = layout["summary_value_col"]   # 0-indexed (23)
        hc_0 = layout["summary_hours_col"]   # 0-indexed (45)

        # ========================================
        # 統計セクション: ラベルセル結合 (col 2 ~ vc_0-1)
        # テンプレートではラベルが col C ~ col W 程度を結合
        # ========================================
        label_end = vc_0 - 1   # ラベル結合の終端列 (0-indexed, 値列の1つ手前)
        label_rows = [
            s.get("patients_row"),
            s.get("nurse_config_row"), s.get("nurse_ref_row"),
            s.get("nurse_all_row"), s.get("nurse_all_sub_row"),
            s.get("nurse_all_ref_row"),
            s.get("nurse_only_row"), s.get("nurse_only_ref_row"),
            s.get("nurse_all_adj_row"),
            s.get("ratio_row"), s.get("ratio1_row"), s.get("ratio2_row"),
            s.get("ratio2_desc_row"), s.get("ratio2_base_row"),
            s.get("ratio_nurse_row"),
            s.get("avg_stay_row"),
            s.get("avg_night_row"),
            s.get("min_nurse_row"),
            s.get("min_nurse_sub_row"), s.get("min_nurse_note_row"),
            s.get("min_nurse_ref_row"),
            s.get("aide_config_row"), s.get("aide_adjusted_row"),
            s.get("aide_ref_row"), s.get("aide_ref_sub_row"),
            s.get("aide_night_row"), s.get("aide_night_ref_row"),
            s.get("clerical_row"), s.get("clerical_sub_row"),
            s.get("clerical_ref_row"), s.get("clerical_ref_sub_row"),
            s.get("all_staff_row"),
            s.get("all_staff_note_row"),
            s.get("all_staff_ref_sub_row"),
        ]
        # 届出注記行もラベル結合
        for note_key in ["aide_note_rows", "aide_night_note_rows"]:
            for nr in s.get(note_key, []):
                label_rows.append(nr)
        if s.get("all_staff_row") is not None:
            label_rows.append(s["all_staff_row"] + 1)
        if s.get("all_staff_ref_row") is not None:
            label_rows.append(s["all_staff_ref_row"])

        for row in label_rows:
            if row is not None:
                self._safe_merge(ws, R(row), C(2), R(row), C(label_end))

        # 単位ラベル結合 (col 25-28 = 4列結合: "人(実績値)" 等)
        unit_rows = [
            s.get("patients_row"),
            s.get("nurse_config_row"), s.get("nurse_ref_row"),
            s.get("nurse_all_row"), s.get("nurse_all_sub_row"),
            s.get("nurse_all_ref_row"),
            s.get("nurse_only_row"), s.get("nurse_only_ref_row"),
            s.get("min_nurse_row"), s.get("min_nurse_ref_row"),
            s.get("aide_config_row"), s.get("aide_adjusted_row"),
            s.get("aide_ref_row"),
            s.get("aide_night_row"), s.get("aide_night_ref_row"),
            s.get("clerical_row"), s.get("clerical_ref_row"),
            s.get("avg_night_row"),
            s.get("ratio_row"), s.get("ratio1_row"), s.get("ratio1_base_row"),
            s.get("ratio2_row"), s.get("ratio2_base_row"),
        ]
        if s.get("all_staff_row") is not None:
            unit_rows.append(s["all_staff_row"] + 1)
        if s.get("all_staff_ref_row") is not None:
            unit_rows.append(s["all_staff_ref_row"])

        for row in unit_rows:
            if row is not None:
                # col 25-27 結合 (単位ラベル) ※col28は算式ラベルが使用
                self._safe_merge(ws, R(row), C(25), R(row), C(27))
                # col 48-51 結合 (時間単位ラベル)
                self._safe_merge(ws, R(row), C(48), R(row), C(51))

        # 算式ラベル結合 (col 28-33 程度: 〔C／(日数×８)〕等)
        formula_label_rows = [
            s.get("nurse_config_row"), s.get("nurse_ref_row"),
            s.get("nurse_all_row"), s.get("nurse_all_ref_row"),
            s.get("nurse_only_row"),
            s.get("min_nurse_ref_row"),
            s.get("aide_ref_row"), s.get("aide_night_ref_row"),
            s.get("clerical_row"), s.get("clerical_ref_row"),
            s.get("avg_night_row"),
            s.get("all_staff_ref_row"),
        ]
        for row in formula_label_rows:
            if row is not None:
                self._safe_merge(ws, R(row), C(28), R(row), C(33))

        # 月延べ時間ラベル結合 (col 43-44: 『月延べ勤務時間数』等, 2列結合)
        hours_label_rows = [
            s.get("nurse_config_row"), s.get("nurse_ref_row"),
            s.get("nurse_all_row"), s.get("nurse_all_sub_row"),
            s.get("nurse_only_row"),
            s.get("nurse_only_ref_row"),
            s.get("min_nurse_sub_row"),
            s.get("ratio_row"), s.get("ratio_sub_row"),
            s.get("ratio_nurse_row"),
            s.get("ratio1_row"), s.get("ratio1_base_row"),
        ]
        if s.get("ratio1_base_row") is not None:
            hours_label_rows.append(s["ratio1_base_row"] + 1)
        # 基準時間ラベルがある届出注記の先頭行も追加
        for note_key in ["aide_note_rows", "aide_night_note_rows"]:
            notes = s.get(note_key, [])
            if notes:
                hours_label_rows.append(notes[0])
        # ※ col 43-44 (AR-AS) はテンプレートでは結合されていない
        # hours_label_rows は書式設定のみに使用

        # ========================================
        # 統計セクション: 値セル結合 (vc_0 ~ vc_0+1 = 2列)
        # ========================================
        value_rows = list(label_rows)  # same rows as labels
        for row in value_rows:
            if row is not None:
                self._safe_merge(ws, R(row), C(vc_0), R(row), C(vc_0 + 1))

        # 時間セル結合 (hc_0 ~ hc_0+2 = 3列)
        hours_rows = [
            s.get("nurse_config_row"), s.get("nurse_ref_row"),
            s.get("nurse_all_row"), s.get("nurse_all_sub_row"),
            s.get("nurse_only_row"),
            s.get("nurse_only_ref_row"),
            s.get("min_nurse_row"), s.get("min_nurse_sub_row"),
            s.get("aide_config_row"), s.get("aide_ref_row"),
            s.get("aide_night_row"),
            s.get("ratio_row"), s.get("ratio_sub_row"),
            s.get("ratio_nurse_row"),
            s.get("ratio1_row"), s.get("ratio1_base_row"),
        ]
        if s.get("ratio1_base_row") is not None:
            hours_rows.append(s["ratio1_base_row"] + 1)
        # 基準時間がある届出注記の先頭行も追加
        for note_key in ["aide_note_rows", "aide_night_note_rows"]:
            notes = s.get(note_key, [])
            if notes:
                hours_rows.append(notes[0])
        for row in hours_rows:
            if row is not None:
                self._safe_merge(ws, R(row), C(hc_0), R(row), C(hc_0 + 2))

        # ========================================
        # Final summary [B][C][D][E][F]
        # ========================================
        fb = layout.get("final_b_row")
        if fb is not None:
            # ラベル結合 (col 2-17)
            self._safe_merge(ws, R(fb), C(2), R(fb), C(17))       # [B]ラベル
            self._safe_merge(ws, R(fb + 1), C(2), R(fb + 1), C(17))  # [D-E]ラベル
            self._safe_merge(ws, R(fb + 2), C(2), R(fb + 2), C(17))  # [F]ラベル
            self._safe_merge(ws, R(fb + 3), C(2), R(fb + 3), C(17))  # 1日看護配置数ラベル
            self._safe_merge(ws, R(fb + 4), C(2), R(fb + 4), C(17))  # 事務上限ラベル
            self._safe_merge(ws, R(fb + 5), C(2), R(fb + 5), C(51))  # 注１
            self._safe_merge(ws, R(fb + 6), C(2), R(fb + 6), C(51))  # 注２
            # [C][D]ラベル (col 20-32)
            self._safe_merge(ws, R(fb), C(20), R(fb), C(32))
            self._safe_merge(ws, R(fb + 1), C(20), R(fb + 1), C(32))
            # [E] ラベル (col 35-45)
            self._safe_merge(ws, R(fb + 1), C(35), R(fb + 1), C(45))
            # 月平均配置数ラベル (col 20-32)
            self._safe_merge(ws, R(fb + 3), C(20), R(fb + 3), C(32))
            # 事務配置数ラベル (col 20-32)
            self._safe_merge(ws, R(fb + 4), C(20), R(fb + 4), C(32))
            # 値セル結合
            for offset in range(0, 5):
                self._safe_merge(ws, R(fb + offset), C(18), R(fb + offset), C(19))
                self._safe_merge(ws, R(fb + offset), C(33), R(fb + offset), C(34))
            # [E] 値セル結合 (col 46-48)
            self._safe_merge(ws, R(fb + 1), C(46), R(fb + 1), C(48))

        # ========================================
        # GHIJK セクション
        # ========================================
        gr = layout.get("ghijk_row")
        if gr is not None:
            gvc_0 = layout["ghijk_value_col"]  # 0-indexed (33)
            gfc_0 = layout["ghijk_formula_col"]  # 0-indexed (19)
            # タイトル行
            self._safe_merge(ws, R(gr), C(2), R(gr), C(gvc_0 - 1))
            for offset in range(1, 10):
                # ラベル (col 2 ~ gfc_0-1)
                self._safe_merge(ws, R(gr + offset), C(2),
                                 R(gr + offset), C(gfc_0 - 1))
                # 算式ラベル (col gfc_0 ~ gvc_0-1)
                self._safe_merge(ws, R(gr + offset), C(gfc_0),
                                 R(gr + offset), C(gvc_0 - 1))
                # 値セル (col gvc_0 ~ gvc_0+1)
                self._safe_merge(ws, R(gr + offset), C(gvc_0),
                                 R(gr + offset), C(gvc_0 + 1))
                # 基準/実績ラベル (col gvc_0+2 ~ gvc_0+4)
                self._safe_merge(ws, R(gr + offset), C(gvc_0 + 2),
                                 R(gr + offset), C(gvc_0 + 4))

        # ========================================
        # [L] セクション
        # ========================================
        lr = layout.get("l_row")
        if lr is not None:
            # タイトル行
            self._safe_merge(ws, R(lr), C(2), R(lr), C(40))
            for offset in range(1, 4):
                # ラベル (col 2 ~ 18)
                self._safe_merge(ws, R(lr + offset), C(2),
                                 R(lr + offset), C(18))
                # 算式ラベル (col 19 ~ 32)
                self._safe_merge(ws, R(lr + offset), C(19),
                                 R(lr + offset), C(32))
                # 値セル (col 33 ~ 34)
                self._safe_merge(ws, R(lr + offset), C(33),
                                 R(lr + offset), C(34))
                # 基準/実績ラベル (col 35 ~ 37)
                self._safe_merge(ws, R(lr + offset), C(35),
                                 R(lr + offset), C(37))

        # ========================================
        # テーブルヘッダー
        # ========================================
        for hr_key in ["header_row", "aide_header_row", "other_staff_header_row"]:
            if hr_key == "header_row":
                hr = layout["staff"].get(hr_key)
            else:
                hr = layout.get(hr_key)
            if hr is None:
                continue
            # 種別 (col 2-3)
            self._safe_merge(ws, R(hr), C(2), R(hr), C(3))
            # 番号 (col 4-5)
            self._safe_merge(ws, R(hr), C(4), R(hr), C(5))
            # 病棟名 (col 6-7)
            self._safe_merge(ws, R(hr), C(6), R(hr), C(7))
            # 氏名 (col 8-13)
            self._safe_merge(ws, R(hr), C(8), R(hr), C(13))
            # 雇用形態 (col 14-18)
            self._safe_merge(ws, R(hr), C(14), R(hr), C(18))
            # 夜勤有無 (col 20-22)
            self._safe_merge(ws, R(hr), C(20), R(hr), C(22))
            # 夜勤従事者計上 (col 23-28)
            self._safe_merge(ws, R(hr), C(23), R(hr), C(28))
            # 日付別 (col 29-59)
            self._safe_merge(ws, R(hr), C(29), R(hr), C(59))
            # 月延べ (col 61 単独で幅7.0なのでOK)
            # 再掲 (col 62-64: 3列結合で長いヘッダーを表示)
            self._safe_merge(ws, R(hr), C(62), R(hr), C(64))
            # 職種別勤務時間 (col 63-64)→既にcol62-64に含まれる
            # 職種別人数 (col 65-66)
            self._safe_merge(ws, R(hr), C(65), R(hr), C(66))

        # ========================================
        # セクション見出し・年月行等
        # ========================================
        # "3．入院患者の数及び看護要員の数" (col 1 → col 1-20)
        sec_row = s.get("section_row")
        if sec_row is not None:
            self._safe_merge(ws, R(sec_row), C(1), R(sec_row), C(20))

        # "4．勤務実績表" (staff section_row)
        staff_sec = layout["staff"].get("section_row")
        if staff_sec is not None:
            self._safe_merge(ws, R(staff_sec), C(1), R(staff_sec), C(15))

        # ⑤ 夜勤時間帯 (night_band_row)
        nb_row = s.get("night_band_row")
        if nb_row is not None:
            self._safe_merge(ws, R(nb_row), C(2), R(nb_row), C(label_end))

        # 年月行 (yearmonth_row)
        ym = s.get("yearmonth_row")
        if ym is not None:
            self._safe_merge(ws, R(ym), C(18), R(ym), C(22))

        # ③の内訳ラベル (看護要員の内訳 col 34-38, 看護師 col 39-42)
        # ※col 43は月延べ時間ラベル結合が使用するため col 42 まで
        for key in ["ratio_row", "ratio1_row"]:
            row = s.get(key)
            if row is not None:
                self._safe_merge(ws, R(row), C(34), R(row), C(38))
                self._safe_merge(ws, R(row), C(39), R(row), C(42))
        # 准看護師・看護補助者ラベル (col 39-42)
        for key in ["ratio_sub_row", "ratio_nurse_row",
                     "ratio1_base_row"]:
            row = s.get(key)
            if row is not None:
                self._safe_merge(ws, R(row), C(39), R(row), C(42))
        # ratio1_base_row + 1
        if s.get("ratio1_base_row") is not None:
            self._safe_merge(ws, R(s["ratio1_base_row"] + 1), C(39),
                             R(s["ratio1_base_row"] + 1), C(42))

        # ⑥の算式ラベル (col 28-33) → formula_label_rows で既に結合済み

        # "5．勤務体制..." セクション見出し
        s5 = layout.get("section5_row")
        if s5 is not None:
            self._safe_merge(ws, R(s5), C(1), R(s5), C(20))

        # ========================================
        # ヘッダーセクション
        # ========================================
        h = layout["header"]
        # タイトル (col 2-36)
        self._safe_merge(ws, R(h["title_row"]), C(2), R(h["title_row"]), C(36))
        # 作成年月日 (col 38-40)
        self._safe_merge(ws, R(h["title_row"]), C(38), R(h["title_row"]), C(40))
        # 保険医療機関名ラベル (col 2-8)
        self._safe_merge(ws, R(h["hospital_row"]), C(2), R(h["hospital_row"]), C(8))
        # 保険医療機関名 (col 9-25)
        self._safe_merge(ws, R(h["hospital_row"]), C(9), R(h["hospital_row"]), C(25))
        # 届出入院基本料ラベル (col 2-12)
        self._safe_merge(ws, R(h["todokede_row"]), C(2), R(h["todokede_row"]), C(12))
        # 届出入院基本料の値 (col 13-25)
        self._safe_merge(ws, R(h["todokede_row"]), C(13), R(h["todokede_row"]), C(25))
        # 看護配置・看護補助配置 ラベル (col 29-35, ward1/ward3 only)
        if layout.get("has_two_part_nurse"):
            self._safe_merge(ws, R(h["todokede_row"]), C(29),
                             R(h["todokede_row"]), C(35))
        # 病棟数・病床数ラベル (col 2-12)
        self._safe_merge(ws, R(h["bcount_row"]), C(2), R(h["bcount_row"]), C(12))
        self._safe_merge(ws, R(h["beds_row"]), C(2), R(h["beds_row"]), C(12))

        # ========================================
        # 合計行ラベル
        # ========================================
        sl = layout["staff"]
        nurse_total = sl["data_start_row"] + sl["max_slots"] * 3
        aide_total = layout["aide_data_start_row"] + sl["max_slots"] * 3
        for total_row in [nurse_total, aide_total]:
            # 「計」ラベル (col 4-5, 3行)
            self._safe_merge(ws, R(total_row), C(4),
                             R(total_row + 2), C(5))
            # 他部署兼務 (col 17-18, 各行)
            for dr in range(3):
                self._safe_merge(ws, R(total_row + dr), C(17),
                                 R(total_row + dr), C(18))
            # 夜勤従事者 (col 23-24, 各行)
            for dr in range(3):
                self._safe_merge(ws, R(total_row + dr), C(23),
                                 R(total_row + dr), C(24))
            # Z列ラベル (col 26-28, 各行)
            for dr in range(3):
                self._safe_merge(ws, R(total_row + dr), C(26),
                                 R(total_row + dr), C(28))
        # 看護職員計 col 2-3 (3行結合)
        self._safe_merge(ws, R(nurse_total), C(2),
                         R(nurse_total + 2), C(3))

        # 表ラベル (《看護職員表》等)
        nl = sl.get("nurse_label_row")
        if nl is not None:
            self._safe_merge(ws, R(nl), C(2), R(nl), C(10))
        al = layout.get("aide_label_row")
        if al is not None:
            self._safe_merge(ws, R(al), C(2), R(al), C(10))
        ol = layout.get("other_staff_label_row")
        if ol is not None:
            self._safe_merge(ws, R(ol), C(2), R(ol), C(10))

        # ========================================
        # 夜勤時間分布セクション
        # ========================================
        dr = layout.get("dist_row")
        if dr is not None:
            # タイトル (col 4-29)
            self._safe_merge(ws, R(dr), C(4), R(dr), C(29))
            self._safe_merge(ws, R(dr + 1), C(4), R(dr + 1), C(29))
            # 各バンド (band_cols 間隔: 6,8,10,12,14,16,18,20,22,24,26,28)
            # 各ラベルは2列結合
            band_cols = [6, 8, 10, 12, 14, 16, 18, 20, 22, 24, 26, 28]
            for section_offset in [3, 7, 11]:  # 看護職員, 補助者, 合計
                label_r = dr + section_offset
                # セクションラベル (col 3-5 → 合計は長いので col 3-14)
                if section_offset == 11:
                    # 【看護職員＋看護補助者】→ 12列結合
                    self._safe_merge(ws, R(label_r), C(3), R(label_r), C(14))
                else:
                    self._safe_merge(ws, R(label_r), C(3), R(label_r), C(9))
                # 夜勤時間数・夜勤従事者数ラベル (col 3-5)
                self._safe_merge(ws, R(label_r + 1), C(3), R(label_r + 1), C(5))
                self._safe_merge(ws, R(label_r + 2), C(3), R(label_r + 2), C(5))
                # バンドラベル (各2列結合)
                for bc in band_cols:
                    self._safe_merge(ws, R(label_r + 1), C(bc), R(label_r + 1), C(bc + 1))

    def _apply_formats(self, ws, layout):
        """フォント・数値フォーマットを一括適用"""
        s = layout["summary"]
        vc = C(layout["summary_value_col"])    # col 24
        hc = C(layout["summary_hours_col"])    # col 46

        # --- 統計セクション: ラベルにFONT_LABEL、値にFONT_VALUE+数値書式 ---
        def fmt_label_row(row, label_col=2):
            """統計行のラベルにフォント適用"""
            if row is None:
                return
            cell = ws.cell(R(row), C(label_col))
            if cell.value:
                cell.font = FONT_LABEL

        def fmt_value(row, col, font=FONT_VALUE, nfmt=None):
            """統計値セルにフォント+数値書式適用"""
            if row is None:
                return
            cell = ws.cell(R(row), col)
            if cell.value is not None:
                cell.font = font
                if nfmt:
                    cell.number_format = nfmt

        # ① 患者数
        fmt_label_row(s["patients_row"])
        fmt_value(s["patients_row"], vc, nfmt=FMT_INT)

        # ② (ward23: nurse_config_row, ward1/3: nurse_all_row, nurse_only_row)
        for key in ["nurse_config_row", "nurse_all_row", "nurse_only_row"]:
            row = s.get(key)
            fmt_label_row(row)
            fmt_value(row, vc, nfmt=FMT_1DEC)
            fmt_value(row, hc, nfmt=FMT_2DEC)

        # 参考行
        for key in ["nurse_ref_row", "nurse_all_ref_row", "nurse_only_ref_row",
                     "nurse_all_adj_row"]:
            row = s.get(key)
            fmt_label_row(row)
            fmt_value(row, vc, nfmt=FMT_1DEC)
            fmt_value(row, hc, nfmt=FMT_2DEC)

        # ③ 比率
        for key in ["ratio_row", "ratio1_row", "ratio2_row"]:
            row = s.get(key)
            fmt_label_row(row)
            fmt_value(row, vc, nfmt=FMT_1DEC)
            fmt_value(row, hc, nfmt=FMT_2DEC)

        # 看護師配置数
        fmt_label_row(s.get("ratio_nurse_row"))
        fmt_value(s.get("ratio_nurse_row"), vc, nfmt=FMT_1DEC)

        # ⑥ 平均夜勤
        fmt_label_row(s.get("avg_night_row"))
        fmt_value(s.get("avg_night_row"), vc, nfmt=FMT_1DEC)

        # ⑦
        fmt_label_row(s.get("min_nurse_row"))
        fmt_value(s.get("min_nurse_row"), vc, nfmt=FMT_1DEC)
        fmt_value(s.get("min_nurse_row"), hc, nfmt=FMT_2DEC)
        fmt_label_row(s.get("min_nurse_ref_row"))
        fmt_value(s.get("min_nurse_ref_row"), vc, nfmt=FMT_1DEC)
        fmt_value(s.get("min_nurse_ref_row"), hc, nfmt=FMT_2DEC)

        # ⑧
        fmt_label_row(s.get("aide_config_row"))
        fmt_value(s.get("aide_config_row"), vc, nfmt=FMT_1DEC)
        fmt_value(s.get("aide_config_row"), hc, nfmt=FMT_2DEC)
        fmt_label_row(s.get("aide_adjusted_row"))
        fmt_value(s.get("aide_adjusted_row"), vc, nfmt=FMT_1DEC)
        fmt_label_row(s.get("aide_ref_row"))
        fmt_value(s.get("aide_ref_row"), vc, nfmt=FMT_1DEC)
        fmt_value(s.get("aide_ref_row"), hc, nfmt=FMT_2DEC)

        # ⑨
        fmt_label_row(s.get("aide_night_row"))
        fmt_value(s.get("aide_night_row"), vc, nfmt=FMT_1DEC)
        fmt_value(s.get("aide_night_row"), hc, nfmt=FMT_2DEC)
        fmt_label_row(s.get("aide_night_ref_row"))
        fmt_value(s.get("aide_night_ref_row"), vc, nfmt=FMT_1DEC)

        # ⑩ (小数3位)
        fmt_label_row(s.get("clerical_row"))
        fmt_value(s.get("clerical_row"), vc, nfmt=FMT_3DEC)
        fmt_label_row(s.get("clerical_ref_row"))
        fmt_value(s.get("clerical_ref_row"), vc, nfmt=FMT_3DEC)

        # ⑪
        if s.get("all_staff_row") is not None:
            fmt_label_row(s["all_staff_row"])
            fmt_value(s["all_staff_row"] + 1, vc, nfmt=FMT_1DEC)
        if s.get("all_staff_ref_row") is not None:
            fmt_label_row(s["all_staff_ref_row"])
            fmt_value(s["all_staff_ref_row"], vc, nfmt=FMT_1DEC)

        # --- Final summary [B][C][D][E][F] ---
        fb = layout.get("final_b_row")
        if fb is not None:
            for offset in range(0, 5):
                fmt_label_row(fb + offset)
            # [B] = 整数
            fmt_value(fb, C(18), nfmt=FMT_INT)
            # [C] = 小数2位
            fmt_value(fb, C(33), nfmt=FMT_2DEC)
            # [D-E] = 小数2位
            fmt_value(fb + 1, C(18), nfmt=FMT_2DEC)
            # [D] = 小数2位
            fmt_value(fb + 1, C(33), nfmt=FMT_2DEC)
            # [E] = 小数2位
            fmt_value(fb + 1, C(46), nfmt=FMT_2DEC)
            # [F] = 小数2位
            fmt_value(fb + 2, C(33), nfmt=FMT_2DEC)
            # 月平均配置数
            fmt_value(fb + 3, C(33), nfmt=FMT_1DEC)
            # 事務配置数
            fmt_value(fb + 4, C(18), nfmt=FMT_3DEC)
            fmt_value(fb + 4, C(33), nfmt=FMT_3DEC)

        # --- GHIJK ---
        gr = layout.get("ghijk_row")
        if gr is not None:
            gvc = C(layout["ghijk_value_col"])
            fmt_label_row(gr)
            for offset in range(1, 10):
                fmt_label_row(gr + offset)
            # [G][H][I] = 小数2位
            fmt_value(gr + 1, gvc, nfmt=FMT_2DEC)
            fmt_value(gr + 2, gvc, nfmt=FMT_2DEC)
            fmt_value(gr + 3, gvc, nfmt=FMT_2DEC)
            # [J] = 整数
            fmt_value(gr + 4, gvc, nfmt=FMT_INT)
            # (G+H)/(日数×8) = 小数1位
            fmt_value(gr + 5, gvc, nfmt=FMT_1DEC)
            # [K] = 小数1位
            fmt_value(gr + 6, gvc, nfmt=FMT_1DEC)
            # 夜間必要数 = 整数
            fmt_value(gr + 7, gvc, nfmt=FMT_INT)
            # 夜間実績 = 小数1位
            fmt_value(gr + 8, gvc, nfmt=FMT_1DEC)
            # K/J% = 小数3位
            fmt_value(gr + 9, gvc, nfmt=FMT_3DEC)

        # --- [L] ---
        lr = layout.get("l_row")
        if lr is not None:
            fmt_label_row(lr)
            for offset in range(1, 4):
                fmt_label_row(lr + offset)
            fmt_value(lr + 1, C(33), nfmt=FMT_INT)   # [L] 必要数
            fmt_value(lr + 2, C(33), nfmt=FMT_1DEC)   # 月平均配置数
            fmt_value(lr + 3, C(33), nfmt=FMT_1DEC)   # 最小必要人数以上

        # --- 勤務実績テーブル: 日付別・月延べにFONT_TABLE+数値書式 ---
        sl = layout["staff"]
        data_sections = [
            (sl["data_start_row"], sl["max_slots"]),
        ]
        # aide table
        aide_start = layout.get("aide_data_start_row")
        if aide_start is not None:
            data_sections.append((aide_start, sl["max_slots"]))

        for data_start, max_slots in data_sections:
            for slot in range(max_slots):
                for row_offset in range(3):  # 3 rows per slot
                    row = data_start + slot * 3 + row_offset
                    # col 60 (月延べ) — フォント+数値書式
                    cell60 = ws.cell(R(row), C(60))
                    cell60.number_format = FMT_2DEC
                    cell60.font = FONT_TABLE
                    # 日付別 (col 29-59) — フォント+数値書式
                    for c in range(29, 29 + self.num_days):
                        cell = ws.cell(R(row), C(c))
                        cell.number_format = FMT_2DEC
                        cell.font = FONT_TABLE

            # 合計行 (data_start + max_slots * 3)
            total_row = data_start + max_slots * 3
            for row_offset in range(3):
                ws.cell(R(total_row + row_offset), C(60)).number_format = FMT_2DEC
                ws.cell(R(total_row + row_offset), C(60)).font = FONT_VALUE
                for c in range(29, 29 + self.num_days):
                    ws.cell(R(total_row + row_offset), C(c)).number_format = FMT_2DEC

        # --- 年月行 ---
        ym = s.get("yearmonth_row")
        if ym is not None:
            fmt_label_row(ym)

        # --- wrap_text: 月延べ勤務時間数ラベル (col 43, summary section) ---
        _wrap = ALIGN_WRAP
        hours_label_rows = [
            s.get("nurse_config_row"), s.get("nurse_all_row"),
            s.get("nurse_only_row"),
            s.get("ratio_row"), s.get("ratio_sub_row"),
            s.get("ratio_nurse_row"),
            s.get("ratio1_row"), s.get("ratio1_base_row"),
        ]
        if s.get("ratio1_base_row") is not None:
            hours_label_rows.append(s["ratio1_base_row"] + 1)
        for row in hours_label_rows:
            if row is not None:
                cell = ws.cell(R(row), C(44))
                if cell.value:
                    cell.alignment = _wrap
                    cell.font = FONT_SMALL

        # --- wrap_text: summary section ラベル (col 2) for very long labels ---
        wrap_label_rows = [
            s.get("nurse_all_row"), s.get("nurse_only_row"),
            s.get("nurse_config_row"), s.get("nurse_ref_row"),
            s.get("nurse_all_ref_row"), s.get("nurse_only_ref_row"),
            s.get("nurse_all_adj_row"),
            s.get("min_nurse_row"), s.get("min_nurse_ref_row"),
            s.get("aide_config_row"), s.get("aide_adjusted_row"),
            s.get("aide_ref_row"),
            s.get("aide_night_row"), s.get("aide_night_ref_row"),
            s.get("clerical_row"), s.get("clerical_ref_row"),
            s.get("ratio_row"), s.get("ratio1_row"), s.get("ratio2_row"),
            s.get("avg_night_row"),
        ]
        if s.get("all_staff_row") is not None:
            wrap_label_rows.append(s["all_staff_row"])
            wrap_label_rows.append(s["all_staff_row"] + 1)
        if s.get("all_staff_ref_row") is not None:
            wrap_label_rows.append(s["all_staff_ref_row"])
        for row in wrap_label_rows:
            if row is not None:
                cell = ws.cell(R(row), C(2))
                if cell.value:
                    cell.alignment = _wrap

        # --- wrap_text: final_summary labels ---
        fb = layout.get("final_b_row")
        if fb is not None:
            for offset in range(0, 5):
                for col in [2, 20, 35]:
                    cell = ws.cell(R(fb + offset), C(col))
                    if cell.value:
                        cell.alignment = _wrap

        # --- row heights: GHIJK / L section ---
        gr = layout.get("ghijk_row")
        if gr is not None:
            for offset in range(0, 10):
                ws.row_dimensions[R(gr + offset)].height = 24
        lr = layout.get("l_row")
        if lr is not None:
            for offset in range(0, 4):
                ws.row_dimensions[R(lr + offset)].height = 24

        # --- row heights: final summary ---
        if fb is not None:
            for offset in range(0, 5):
                ws.row_dimensions[R(fb + offset)].height = 24

        # --- row heights: header section ---
        h = layout["header"]
        ws.row_dimensions[R(h["todokede_row"])].height = 18

    # ============================================================
    # ユーティリティ
    # ============================================================
    def _get_patient_count(self, ward_key):
        mapping = {
            "ward1": self.patients.get("ward1", 58),
            "ward3": self.patients.get("ward3", 58),
            "ward23": (self.patients.get("ward2", 42) + self.patients.get("ward3", 58)),
        }
        return mapping.get(ward_key, 58)


# ============================================================
# 外部API用関数
# ============================================================
def get_finalization_status(year, month):
    wards = {
        "ichiboutou": {"name": "1病棟"},
        "nibyoutou": {"name": "2病棟"},
        "sanbyoutou": {"name": "3病棟"},
    }
    result = {}
    for wid, info in wards.items():
        path = os.path.join(BASE_DIR, "shifts", wid, f"{year}-{month:02d}.json")
        if os.path.exists(path):
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
            actual = data.get("actual")
            result[wid] = {
                "name": info["name"],
                "hasActual": actual is not None and bool(actual.get("shifts")),
                "finalized": actual is not None and actual.get("finalizedAt") is not None,
                "finalizedAt": actual.get("finalizedAt") if actual else None,
            }
        else:
            result[wid] = {
                "name": info["name"],
                "hasActual": False,
                "finalized": False,
                "finalizedAt": None,
            }
    return result
