#!/usr/bin/env python3
"""看護勤務表ExcelのデータをシフトアプリのJSONに取り込むスクリプト

使用方法:
    python import_kinmuhyo.py --year 2026 --month 3 [--xlsx 看護勤務表3月.xlsx]

引数:
    --year      対象年 (例: 2026)
    --month     対象月 (例: 3)
    --xlsx      Excelファイルパス (省略時: 看護勤務表{month}月.xlsx)
    --staff-hours  スタッフ別dayHours設定 (例: "10340440:3.0,99999999:5.0")
"""

import argparse
import json
import os
from datetime import datetime
import openpyxl

# --- 定数 ---
EMPLOYEES_PATH = os.path.join(os.path.dirname(__file__), "shared", "employees.json")
SHIFTS_DIR = os.path.join(os.path.dirname(__file__), "shifts")

SHEET_WARD_MAP = {
    "1病棟": "ichiboutou",
    "2病棟": "nibyoutou",
    "3病棟": "sanbyoutou",
}

# Excelシフト略称 → アプリコード
SHIFT_MAP = {
    "日": "day",
    "夜": "night2",
    "明": "ake",
    "休": "off",
    "深": "shinya",
    "準": "junnya",
    "リ": "refresh",
    "有": "paid",

    "遅２": "late",
    "AM": "day",
    "I": "day",
    "病": "off",
    "E": "day",
    "12": "day",
    "15": "day",
}

# dayHoursが必要なコード (Excel略称 → 時間数)
DAY_HOURS_MAP = {
    "AM": 3.0,
    "I": 7.0,
    "E": 5.0,
    "12": 3.0,
    "15": 6.0,
}


def parse_args():
    parser = argparse.ArgumentParser(description="看護勤務表ExcelをシフトアプリJSONに変換")
    parser.add_argument("--year", type=int, required=True, help="対象年 (例: 2026)")
    parser.add_argument("--month", type=int, required=True, help="対象月 (例: 3)")
    parser.add_argument("--xlsx", type=str, default=None,
                        help="Excelファイルパス (省略時: 看護勤務表{month}月.xlsx)")
    parser.add_argument("--staff-hours", type=str, default="",
                        help='スタッフ別dayHours設定 (例: "10340440:3.0,99999999:5.0")')
    return parser.parse_args()


def parse_staff_hours(staff_hours_str):
    """--staff-hours 引数をパースして {staff_id: hours} の辞書に変換"""
    result = {}
    if not staff_hours_str:
        return result
    for entry in staff_hours_str.split(","):
        entry = entry.strip()
        if ":" not in entry:
            continue
        sid, hours = entry.split(":", 1)
        try:
            result[sid.strip()] = float(hours.strip())
        except ValueError:
            print(f"  [WARN] --staff-hours の解析失敗: '{entry}'")
    return result


def load_employees():
    with open(EMPLOYEES_PATH, encoding="utf-8") as f:
        return json.load(f)


def build_name_to_id(employees, ward_id):
    """病棟ごとに名前→IDのマップを作成（全角スペースの揺れを吸収）"""
    mapping = {}
    for emp in employees:
        if emp["ward"] == ward_id:
            name = emp["name"]
            mapping[name] = emp["id"]
            mapping[name.replace("\u3000", "")] = emp["id"]
    return mapping


def parse_sheet(ws, name_to_id, staff_day_hours):
    """シートからスタッフのシフトデータを抽出

    Returns:
        tuple: (shifts, day_hours, skipped_names)
            - shifts: { staff_id: { "1": "day", ... } }
            - day_hours: { "staff_id-day": hours }
            - skipped_names: マッチしなかった名前のリスト
    """
    shifts = {}      # { staff_id: { "1": "day", ... } }
    day_hours = {}   # { "staff_id-day": hours }
    skipped_names = []

    row_idx = 9
    while row_idx <= ws.max_row:
        d_val = ws.cell(row=row_idx, column=4).value  # D列 = 予定/実施
        if d_val != "予定":
            row_idx += 1
            continue

        next_row = row_idx + 1
        if next_row > ws.max_row:
            break
        staff_name = ws.cell(row=next_row, column=3).value  # C列 = 氏名
        if not staff_name:
            row_idx += 1
            continue

        staff_name = str(staff_name).strip()

        staff_id = name_to_id.get(staff_name) or name_to_id.get(staff_name.replace("\u3000", ""))
        if not staff_id:
            print(f"  [SKIP] 名前マッチなし: {staff_name}")
            skipped_names.append(staff_name)
            row_idx += 2
            continue

        # 予定行のシフトデータ読み取り (E列=1日 ～ AI列=31日)
        staff_shifts = {}
        for col_idx in range(5, 36):  # E=5 to AI=35
            day = col_idx - 4  # 1-31
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val is None:
                continue
            # float型セル (例: 12.0, 15.0) を整数文字列に正規化
            if isinstance(cell_val, float) and cell_val == int(cell_val):
                cell_val = int(cell_val)
            code = str(cell_val).strip()
            if not code:
                continue

            app_code = SHIFT_MAP.get(code)
            if app_code:
                staff_shifts[str(day)] = app_code
                if code in DAY_HOURS_MAP:
                    day_hours[f"{staff_id}-{day}"] = DAY_HOURS_MAP[code]
                elif app_code == "day" and staff_id in staff_day_hours:
                    day_hours[f"{staff_id}-{day}"] = staff_day_hours[staff_id]

        if staff_shifts:
            shifts[staff_id] = staff_shifts
            print(f"  [OK] {staff_name} (ID:{staff_id}) → {len(staff_shifts)}日分")

        row_idx += 2  # 予定+実施の2行分進む

    return shifts, day_hours, skipped_names


def build_json(ward_id, year, month, shifts, day_hours):
    """アプリ用のJSON構造を構築"""
    now = datetime.now().isoformat()
    draft_name = "Excel取込"

    data = {
        "year": year,
        "month": month,
        "ward": ward_id,
        "status": "confirmed",
        "selectedDraft": draft_name,
        "confirmedAt": now,
        "drafts": {
            draft_name: {
                "createdAt": now,
                "score": 0,
                "shifts": shifts,
            }
        },
        "confirmed": {
            "shifts": shifts,
        },
        "changeHistory": [],
    }

    if day_hours:
        data["confirmed"]["dayHours"] = day_hours

    return data


def main():
    args = parse_args()
    year = args.year
    month = args.month

    if month < 1 or month > 12:
        print(f"[ERROR] month は 1-12 の範囲で指定してください: {month}")
        return

    # Excelファイルパス決定
    if args.xlsx:
        xlsx_path = args.xlsx
    else:
        xlsx_path = os.path.join(os.path.dirname(__file__), f"看護勤務表{month}月.xlsx")

    if not os.path.exists(xlsx_path):
        print(f"[ERROR] Excelファイルが見つかりません: {xlsx_path}")
        return

    staff_day_hours = parse_staff_hours(args.staff_hours)

    print(f"=== 看護勤務表取込 ({year}年{month}月) ===")
    print(f"  Excel: {xlsx_path}")
    if staff_day_hours:
        print(f"  スタッフ別dayHours: {staff_day_hours}")
    print()

    employees = load_employees()
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    # サマリレポート用
    summary = {"success": [], "skipped_sheets": [], "skipped_staff": [], "errors": []}

    for sheet_name, ward_id in SHEET_WARD_MAP.items():
        if sheet_name not in wb.sheetnames:
            print(f"--- {sheet_name} ({ward_id}) --- [SKIP: シートなし]")
            summary["skipped_sheets"].append(sheet_name)
            continue
        print(f"--- {sheet_name} ({ward_id}) ---")
        ws = wb[sheet_name]
        name_to_id = build_name_to_id(employees, ward_id)

        shifts, day_hours, skipped_names = parse_sheet(ws, name_to_id, staff_day_hours)
        if skipped_names:
            summary["skipped_staff"].extend(
                {"ward": sheet_name, "name": n} for n in skipped_names
            )

        data = build_json(ward_id, year, month, shifts, day_hours)
        out_path = os.path.join(SHIFTS_DIR, ward_id, f"{year}-{month:02d}.json")
        os.makedirs(os.path.dirname(out_path), exist_ok=True)

        try:
            with open(out_path, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            print(f"  → {out_path} に保存 ({len(shifts)}名)")
            if day_hours:
                print(f"  → dayHours: {len(day_hours)}件")
            summary["success"].append({"ward": sheet_name, "staff_count": len(shifts), "day_hours_count": len(day_hours)})
        except (OSError, IOError) as e:
            print(f"  [ERROR] ファイル保存失敗: {e}")
            summary["errors"].append({"ward": sheet_name, "error": str(e)})
        print()

    # サマリレポート出力
    print("=" * 40)
    print("  取込結果サマリ")
    print("=" * 40)
    if summary["success"]:
        for s in summary["success"]:
            print(f"  ✓ {s['ward']}: {s['staff_count']}名取込, dayHours {s['day_hours_count']}件")
    if summary["skipped_sheets"]:
        print(f"  - スキップ（シートなし）: {', '.join(summary['skipped_sheets'])}")
    if summary["skipped_staff"]:
        print(f"  - 名前マッチなし（{len(summary['skipped_staff'])}名）:")
        for s in summary["skipped_staff"]:
            print(f"      {s['ward']}: {s['name']}")
    if summary["errors"]:
        for e in summary["errors"]:
            print(f"  ✗ {e['ward']}: {e['error']}")
    if not summary["success"] and not summary["errors"]:
        print("  取込対象のデータがありませんでした")
    print()

    total_staff = sum(s["staff_count"] for s in summary["success"])
    print(f"完了！ (合計 {total_staff}名取込)")
    if summary["errors"]:
        print(f"[WARNING] {len(summary['errors'])}件のエラーが発生しました。上記を確認してください。")


if __name__ == "__main__":
    main()
