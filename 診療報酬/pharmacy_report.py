#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""薬剤月次レポート生成スクリプト

精神科病院の入院帳票PDF（薬剤.pdf）から薬剤データを自動抽出し、
薬効分類別に集計した月次Excelレポートを生成する。

使用法:
    python generate_drug_report.py

出力:
    薬剤月次レポート.xlsx（4シート構成）
"""

import os
import re
import sys
import datetime
import unicodedata
from collections import defaultdict

import pdfplumber
import openpyxl
from openpyxl.styles import (
    Font, Alignment, Border, Side, PatternFill, numbers,
)
from openpyxl.formatting.rule import CellIsRule, DataBarRule
from openpyxl.utils import get_column_letter

# ======================================================================
# パス定義
# ======================================================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_PATH = os.path.join(BASE_DIR, '薬剤月次レポート.xlsx')

import glob as _glob
import json

# ======================================================================
# スタイル定数（generate_excel_v2.py と同系統）
# ======================================================================
thin = Side(style='thin')
thick = Side(style='medium')
thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
sub_header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
light_blue_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
total_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
kpi_green = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
kpi_red = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')
kpi_blue = PatternFill(start_color='DAEEF3', end_color='DAEEF3', fill_type='solid')
kpi_gray = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
cf_red = PatternFill(start_color='FFCCCC', end_color='FFCCCC')
cf_blue = PatternFill(start_color='CCE5FF', end_color='CCE5FF')
cf_red_strong = PatternFill(start_color='FF9999', end_color='FF9999')
cf_blue_strong = PatternFill(start_color='99CCFF', end_color='99CCFF')
setting_fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')

title_font = Font(bold=True, size=16, color='2F5496')
subtitle_font = Font(bold=True, size=13, color='2F5496')
section_font = Font(bold=True, size=11, color='2F5496')
header_font_w = Font(bold=True, size=10, color='FFFFFF')
data_font = Font(size=10)
bold_font = Font(bold=True, size=10)
kpi_value_font = Font(bold=True, size=22, color='2F5496')
kpi_label_font = Font(size=9, color='666666')
kpi_delta_pos = Font(bold=True, size=14, color='0066CC')
kpi_delta_neg = Font(bold=True, size=14, color='CC0000')

NUM = '#,##0'
PCT = '0.0%'
YEN = '#,##0"円"'

# ======================================================================
# 薬効分類マッピング辞書
# ======================================================================
# 分類名 → キーワードリスト（薬名に含まれていればマッチ）
DRUG_CLASSIFICATION = {
    '抗精神病薬（非定型）': [
        'オランザピン', 'クエチアピン', 'リスペリドン', 'リスパダール',
        'アリピプラゾール', 'エビリファイ', 'インヴェガ', 'パリペリドン',
        'ブロナンセリン', 'レキサルティ', 'ジプレキサ', 'ラツーダ',
        'クロザリル', 'クロザピン', 'ロナセン', 'ロドピン',
        'セロクエル', 'ルーラン',
    ],
    '抗精神病薬（定型）': [
        'セレネース', 'ハロペリドール', 'ヒルナミン', 'レボメプロマジン',
        'コントミン', 'クロルプロマジン', 'フルフェナジン',
        'ブロムペリドール', 'インプロメン',
        'ピレチア', 'プロメタジン',
    ],
    '抗うつ薬': [
        'イフェクサー', 'ベンラファキシン', 'デュロキセチン', 'サインバルタ',
        'トリンテリックス', 'ミルタザピン', 'リフレックス', 'レメロン',
        'セルトラリン', 'ジェイゾロフト', 'トラゾドン', 'デジレル', 'レスリン',
        'ルボックス', 'フルボキサミン', 'アナフラニール', 'クロミプラミン',
        'パロキセチン', 'エスシタロプラム', 'レクサプロ',
    ],
    '睡眠薬': [
        'デエビゴ', 'レンボレキサント', 'ベルソムラ', 'スボレキサント',
        'エスゾピクロン', 'ルネスタ', 'ゾルピデム', 'マイスリー',
        'フルニトラゼパム', 'サイレース', 'ブロチゾラム', 'レンドルミン',
        'ベンザリン', 'ニトラゼパム', 'ラメルテオン', 'ロゼレム',
        'トリアゾラム', 'ハルシオン',
    ],
    '抗不安薬': [
        'エチゾラム', 'デパス', 'メイラックス', 'ロフラゼプ',
        'レキソタン', 'ブロマゼパム', 'ワイパックス', 'ロラゼパム',
        'リーゼ', 'クロチアゼパム', 'アタラックス', 'ヒドロキシジン',
        'アルプラゾラム', 'ソラナックス', 'コンスタン',
        'ホリゾン', 'ジアゼパム', 'セルシン',
        'セディール', 'タンドスピロン',
    ],
    '気分安定薬・抗てんかん薬': [
        'デパケン', 'バルプロ酸', 'テグレトール', 'カルバマゼピン',
        '炭酸リチウム', 'リーマス', 'ラモトリギン', 'ラミクタール',
        'レベチラセタム', 'イーケプラ', 'アレビアチン', 'フェニトイン',
        'ゾニサミド', 'エクセグラン', 'リボトリール', 'クロナゼパム',
        'フェノバール', 'フェノバルビタール', 'ダイアップ',
        'トピラマート', 'ガバペンチン', 'ラコサミド',
        'ホストイン', 'ホスフェニトイン',
    ],
    '抗パーキンソン薬': [
        'アーテン', 'トリヘキシフェニジル', 'アキネトン', 'ビペリデン',
        'ビ・シフロール', 'プラミペキソール', 'メネシット', 'レボドパ',
        'タスモリン',
    ],
    'ADHD治療薬': [
        'インチュニブ', 'グアンファシン', 'コンサータ', 'メチルフェニデート',
        'ストラテラ', 'アトモキセチン', 'ビバンセ',
    ],
    '認知症治療薬': [
        'アリセプト', 'ドネペジル', 'メマリー', 'メマンチン',
        'レミニール', 'ガランタミン', 'イクセロンパッチ', 'リバスチグミン',
    ],
    '消化器系薬': [
        'タケキャブ', 'ボノプラザン', 'オメプラール', 'オメプラゾール',
        'ランソプラゾール', 'タケプロン', 'ネキシウム', 'エソメプラゾール',
        'レバミピド', 'ムコスタ', 'プロマック', 'ポラプレジンク',
        'ファモチジン', 'ガスター', 'ガスコン', 'ジメチコン',
        'モサプリド', 'ガスモチン', 'ウルソ', 'ウルソデオキシコール',
        'アミティーザ', 'ルビプロストン', 'グーフィス', 'エロビキシバット',
        'センノシド', 'プルゼニド', 'ピコスルファート', 'ラキソベロン',
        'マグミット', '酸化マグネシウム', 'ラクツロース', 'モニラック',
        'ビオスリー', 'ビオフェルミン',
        'S・M配合散', 'エスエム', 'ミヤBM', 'ナウゼリン', 'ドンペリドン',
        'パンクレリパーゼ', 'リパクレオン', 'カロナール配合',
        'テレミンソフト', 'グリセリン浣腸',
    ],
    '循環器系薬': [
        'アムロジピン', 'ノルバスク', 'カンデサルタン', 'ブロプレス',
        'アジルバ', 'アジルサルタン', 'エンレスト', 'サクビトリル',
        'ラシックス', 'フロセミド', 'ダイアート', 'アゾセミド',
        'スピロノラクトン', 'アルダクトン', 'メインテート', 'ビソプロロール',
        'ニフェジピン', 'アダラート', 'シグマート', 'ニコランジル',
        'ニトロール', 'イソソルビド', 'エブランチル', 'ウラピジル',
        'バイアスピリン', 'アスピリン', 'プラビックス', 'クロピドグレル',
        'リクシアナ', 'エドキサバン', 'シロスタゾール', 'プレタール',
        'サムスカ', 'トルバプタン', 'テルミサルタン', 'ミカルディス',
        'カルベジロール', 'アーチスト', 'ドキサゾシン', 'カルデナリン',
        'ワーファリン', 'ワルファリン', 'エリキュース', 'アピキサバン',
        'リバロ', 'ピタバスタチン', 'ロスバスタチン', 'クレストール',
        'ゼチーア', 'エゼチミブ', 'アトルバスタチン', 'リピトール',
        'プラバスタチン', 'メバロチン', 'フェノフィブラート', 'リピディル',
        'ベザフィブラート', 'ベザトール',
    ],
    '糖尿病薬': [
        'ジャディアンス', 'エンパグリフロジン', 'ジャヌビア', 'シタグリプチン',
        'フォシーガ', 'ダパグリフロジン', 'グルファスト', 'ミチグリニド',
        'トラゼンタ', 'リナグリプチン', 'ボグリボース', 'ベイスン',
        'メトグルコ', 'メトホルミン', 'トルリシティ', 'デュラグルチド',
        'トレシーバ', 'インスリンデグルデク', 'ヒューマログ', 'インスリンリスプロ',
        'ノボラピッド', 'ランタス', 'グラルギン',
    ],
    '漢方薬': [
        '抑肝散', 'ツムラ抑肝散', '補中益気湯', '六君子湯',
        '加味帰脾湯', '半夏厚朴湯', '葛根湯', '芍薬甘草湯',
        '柴胡加竜骨', '桂枝加竜骨', '加味逍遙散', '当帰芍薬散',
        '五苓散', '麻子仁丸', '牛車腎気丸', '八味地黄丸',
        '人参養栄湯', '十全大補湯', '防風通聖散',
        '大建中湯', '大黄甘草湯',
    ],
    '輸液': [
        'ラクテック', 'KN', '生食', '生理食塩', '大塚生食', '大塚糖液',
        'アクチット', 'ソルデム', 'ヴィーン', 'フィジオ', 'ソリタ',
        'エルネオパ', 'ビーフリード', 'ソルラクト',
    ],
    '抗菌薬': [
        'スルバシリン', 'セフトリアキソン', 'オーグメンチン', 'サワシリン',
        'ケフレックス', 'セファレキシン', 'レボフロキサシン', 'クラビット',
        'バクトラミン', 'バルトレックス', 'バラシクロビル',
        'テルビナフィン',
        'メロペネム', 'メロペン', 'アモキシシリン',
    ],
    '皮膚科外用薬': [
        'アンテベート', 'ヒルドイド', 'ロコイド', 'マイザー',
        'リンデロン', 'オイラックス', 'ケラチナミン', 'ワセリン',
        'サリチル酸', '亜鉛華', 'ヘパリン類似物質', 'ダイアコート',
        'リドメックス', 'エキザルベ', 'レスタミン', 'ヘモレックス',
        'ゲーベン', 'デルモベート', 'プロペト', 'ベトネベート',
        'アラセナ', 'ビダラビン',
        'ニゾラール', 'ケトコナゾール', 'ラミシール',
    ],
}

# 分類の表示順序
CLASSIFICATION_ORDER = [
    '抗精神病薬（非定型）',
    '抗精神病薬（定型）',
    '抗うつ薬',
    '睡眠薬',
    '抗不安薬',
    '気分安定薬・抗てんかん薬',
    '抗パーキンソン薬',
    'ADHD治療薬',
    '認知症治療薬',
    '消化器系薬',
    '循環器系薬',
    '糖尿病薬',
    '漢方薬',
    '輸液',
    '抗菌薬',
    '皮膚科外用薬',
    'その他',
]

# 診区コード → 名称
SHINKU_MAP = {
    '21': '内服', '22': '屯服', '23': '外用',
    '31': '注射', '33': '点滴', '40': '処置薬',
}

# ======================================================================
# 1. PDF解析モジュール
# ======================================================================

def parse_number(text):
    """カンマ付き数値文字列をfloatに変換"""
    if text is None:
        return 0.0
    text = str(text).strip().replace(',', '').replace(' ', '')
    if text in ('', '-', '―'):
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def extract_drugs_from_pdf(pdf_path):
    """薬剤.pdfからデータを抽出しリストで返す。

    Returns:
        list[dict]: 各薬剤のデータ辞書。キー:
            shinku, code, name, unit_price,
            outpatient_qty, inpatient_qty, total_qty,
            outpatient_amt, inpatient_amt, total_amt, ratio
    """
    drugs = []

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables({
                'vertical_strategy': 'lines',
                'horizontal_strategy': 'lines',
                'snap_tolerance': 5,
            })
            if not tables:
                # lines ベースでテーブルが見つからない場合 text ベースにフォールバック
                tables = page.extract_tables({
                    'vertical_strategy': 'text',
                    'horizontal_strategy': 'text',
                    'snap_tolerance': 5,
                })
            for table in tables:
                for row in table:
                    if row is None or len(row) < 8:
                        continue

                    # ヘッダー行・小計行・合計行をスキップ
                    cell0 = str(row[0] or '').strip()
                    row_text = ' '.join(str(c or '') for c in row)
                    if '診区' in cell0 and 'コード' in row_text:
                        continue
                    if '小計' in row_text or '合計' in row_text:
                        continue
                    if '診療行為' in row_text or '集計表' in row_text:
                        continue
                    if '令和' in row_text and '月' in row_text and len(cell0) <= 2:
                        continue

                    # 診区コードが21,22,23,31,33,40のいずれかであること
                    if cell0 not in ('21', '22', '23', '31', '33', '40'):
                        continue

                    # コード列が数値的かチェック
                    code_str = str(row[1] or '').strip().replace(' ', '')
                    if not code_str or not any(c.isdigit() for c in code_str):
                        continue

                    # 薬名を取得
                    name = str(row[2] or '').strip()
                    if not name:
                        continue

                    # 列数に応じてパース
                    # 期待: [診区, コード, 薬名, 単価, 外来使用量, 入院使用量, 合計使用量,
                    #         外来金額, 入院金額, 合計金額, 全体比率]
                    # PDFのテーブル抽出では列数が変動する場合がある
                    if len(row) >= 11:
                        drug = {
                            'shinku': cell0,
                            'code': code_str,
                            'name': name,
                            'unit_price': parse_number(row[3]),
                            'outpatient_qty': parse_number(row[4]),
                            'inpatient_qty': parse_number(row[5]),
                            'total_qty': parse_number(row[6]),
                            'outpatient_amt': parse_number(row[7]),
                            'inpatient_amt': parse_number(row[8]),
                            'total_amt': parse_number(row[9]),
                            'ratio': parse_number(row[10]),
                        }
                    elif len(row) >= 10:
                        drug = {
                            'shinku': cell0,
                            'code': code_str,
                            'name': name,
                            'unit_price': parse_number(row[3]),
                            'outpatient_qty': parse_number(row[4]),
                            'inpatient_qty': parse_number(row[5]),
                            'total_qty': parse_number(row[6]),
                            'outpatient_amt': parse_number(row[7]),
                            'inpatient_amt': parse_number(row[8]),
                            'total_amt': parse_number(row[9]),
                            'ratio': 0.0,
                        }
                    else:
                        # 最低限の列だけ取れた場合
                        drug = {
                            'shinku': cell0,
                            'code': code_str,
                            'name': name,
                            'unit_price': parse_number(row[3]) if len(row) > 3 else 0,
                            'outpatient_qty': 0,
                            'inpatient_qty': parse_number(row[4]) if len(row) > 4 else 0,
                            'total_qty': parse_number(row[5]) if len(row) > 5 else 0,
                            'outpatient_amt': 0,
                            'inpatient_amt': parse_number(row[6]) if len(row) > 6 else 0,
                            'total_amt': parse_number(row[7]) if len(row) > 7 else 0,
                            'ratio': 0.0,
                        }

                    drugs.append(drug)

    # 重複除去（同一コード+診区の重複がページ跨ぎで発生する場合）
    seen = set()
    unique_drugs = []
    for d in drugs:
        key = (d['shinku'], d['code'])
        if key not in seen:
            seen.add(key)
            unique_drugs.append(d)

    return unique_drugs


# ======================================================================
# 2. 薬効分類モジュール
# ======================================================================

def normalize_text(text):
    """全角英数字を半角に正規化（Ｓ→S, ０→0等）"""
    return unicodedata.normalize('NFKC', text)


def strip_maker_name(name):
    """後発品の社名（「メーカー名」「メーカー名）を除去"""
    import re
    return re.sub(r'[「\uFF62][^」\uFF63]*[」\uFF63]?\s*$', '', name).rstrip()


def classify_drug(drug_name):
    """薬名から薬効分類を返す。マッチしなければ'その他'。"""
    normalized = normalize_text(drug_name)
    for category, keywords in DRUG_CLASSIFICATION.items():
        for kw in keywords:
            if kw in drug_name or kw in normalized:
                return category
    return 'その他'


def classify_all_drugs(drugs):
    """全薬剤に分類を付与してdictに'classification'キーを追加。"""
    for d in drugs:
        d['classification'] = classify_drug(d['name'])
    return drugs


# ======================================================================
# 3. 集計ロジック
# ======================================================================

def aggregate_by_classification(drugs):
    """薬効分類ごとに入院金額を集計。

    Returns:
        dict: {分類名: {'total_amt': 金額合計, 'drugs': [drug, ...], 'count': 件数}}
    """
    result = defaultdict(lambda: {'total_amt': 0, 'drugs': [], 'count': 0})
    for d in drugs:
        cat = d['classification']
        result[cat]['total_amt'] += d['inpatient_amt']
        result[cat]['drugs'].append(d)
        result[cat]['count'] += 1
    return dict(result)


def get_report_month_from_pdf(pdf_path):
    """PDFの表紙から対象月を取得。失敗時は現在月-1。"""
    try:
        with pdfplumber.open(pdf_path) as pdf:
            text = pdf.pages[0].extract_text() or ''
            # 「令和X年Y月」パターンを探す
            m = re.search(r'令和(\d+)年(\d+)月', text)
            if m:
                reiwa_year = int(m.group(1))
                month = int(m.group(2))
                western_year = reiwa_year + 2018
                return western_year, month
    except Exception:
        pass
    # フォールバック: 先月
    today = datetime.date.today()
    first = today.replace(day=1)
    prev = first - datetime.timedelta(days=1)
    return prev.year, prev.month


def load_history(excel_path):
    """既存の薬剤月次レポート.xlsxから推移データを読み込む。

    Returns:
        list[dict]: [{'年月': 'YYYY-MM', '薬効分類': str, '金額': int}, ...]
    """
    if not os.path.exists(excel_path):
        return []
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        if '推移グラフ用データ' not in wb.sheetnames:
            wb.close()
            return []
        ws = wb['推移グラフ用データ']
        history = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1] is not None:
                history.append({
                    '年月': str(row[0]),
                    '薬効分類': str(row[1]),
                    '金額': int(row[2] or 0),
                })
        wb.close()
        return history
    except Exception:
        return []


# ======================================================================
# 4. Excelヘルパー関数
# ======================================================================

def set_outer_border(ws, min_row, min_col, max_row, max_col):
    """外枠太線・内枠細線を設定"""
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            cell = ws.cell(row=r, column=c)
            left = thick if c == min_col else thin
            right = thick if c == max_col else thin
            top = thick if r == min_row else thin
            bottom = thick if r == max_row else thin
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)


def write_header_row(ws, row, col_start, labels, fill=None):
    """ヘッダー行を書き込み"""
    f = fill or header_fill
    for i, label in enumerate(labels):
        c = col_start + i
        cell = ws.cell(row=row, column=c, value=label)
        cell.font = header_font_w
        cell.fill = f
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border


def write_data_row(ws, row, col_start, values, fmt_list=None, is_total=False, is_sub=False):
    """データ行を書き込み"""
    fill = total_fill if is_total else (light_blue_fill if is_sub else None)
    font = bold_font if (is_total or is_sub) else data_font
    for i, val in enumerate(values):
        c = col_start + i
        cell = ws.cell(row=row, column=c, value=val)
        cell.font = font
        cell.border = thin_border
        if fill:
            cell.fill = fill
        if i == 0:
            cell.alignment = Alignment(horizontal='left', vertical='center')
        else:
            cell.alignment = Alignment(horizontal='right', vertical='center')
            if fmt_list and i < len(fmt_list) and fmt_list[i]:
                cell.number_format = fmt_list[i]


def setup_print(ws, title=""):
    """A4横・1ページ収まり・ヘッダー/フッター"""
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(
        fitToPage=True
    )
    ws.oddFooter.center.text = "&P / &N"


# ======================================================================
# 5. Excel出力 — 4シート
# ======================================================================

def build_summary_sheet(wb, ws, agg, grand_total, year, month, history, current_ym):
    """Sheet1: サマリー"""
    ws.title = 'サマリー'
    ws.sheet_properties.tabColor = '2F5496'

    # タイトル
    ws.merge_cells('A1:I1')
    ws['A1'] = f'薬剤月次レポート {month}月度'
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(vertical='center')

    ws.merge_cells('A2:I2')
    ws['A2'] = f'対象期間: {year}年{month}月 | 作成日: {datetime.date.today().strftime("%Y/%m/%d")}'
    ws['A2'].font = Font(size=10, color='666666')

    # KPIボックス
    # 精神科薬剤の合計（抗精神病薬 + 抗うつ薬 + 睡眠薬 + 抗不安薬 + 気分安定薬 + 抗パーキンソン + ADHD + 認知症）
    psych_categories = [
        '抗精神病薬（非定型）', '抗精神病薬（定型）', '抗うつ薬', '睡眠薬',
        '抗不安薬', '気分安定薬・抗てんかん薬', '抗パーキンソン薬', 'ADHD治療薬',
        '認知症治療薬',
    ]
    psych_total = sum(agg.get(c, {}).get('total_amt', 0) for c in psych_categories)
    psych_ratio = psych_total / grand_total if grand_total > 0 else 0

    # 前月データがあれば前月比を計算
    prev_month_total = None
    if history:
        # 1つ前の月を探す
        months_in_history = sorted(set(h['年月'] for h in history))
        months_in_history = [m for m in months_in_history if m != current_ym]
        if months_in_history:
            prev_ym = months_in_history[-1]
            prev_month_total = sum(h['金額'] for h in history if h['年月'] == prev_ym)

    # KPI行
    kpi_row = 4
    kpi_configs = [
        ('A', 'C', '入院薬剤費合計', grand_total, YEN, kpi_blue, kpi_value_font),
        ('D', 'F', '前月比',
         (grand_total - prev_month_total) / prev_month_total if prev_month_total else None,
         '+0.0%;-0.0%',
         kpi_green if (prev_month_total and grand_total >= prev_month_total) else kpi_red,
         kpi_delta_pos if (prev_month_total and grand_total >= prev_month_total) else kpi_delta_neg),
        ('G', 'I', '精神科薬剤比率', psych_ratio, '0.0%', kpi_gray,
         Font(bold=True, size=18, color='2F5496')),
    ]
    for col_s, col_e, label, val, fmt, bg, vfont in kpi_configs:
        ws.merge_cells(f'{col_s}{kpi_row}:{col_e}{kpi_row}')
        cell_label = ws[f'{col_s}{kpi_row}']
        cell_label.value = label
        cell_label.font = kpi_label_font
        cell_label.fill = bg
        cell_label.alignment = Alignment(horizontal='center')

        ws.merge_cells(f'{col_s}{kpi_row+1}:{col_e}{kpi_row+2}')
        cell_val = ws[f'{col_s}{kpi_row+1}']
        if val is not None:
            cell_val.value = val
        else:
            cell_val.value = '—'
        cell_val.font = vfont
        cell_val.fill = bg
        cell_val.alignment = Alignment(horizontal='center', vertical='center')
        if val is not None and fmt:
            cell_val.number_format = fmt

    # 薬効分類別テーブル
    table_start = 8
    ws[f'A{table_start}'] = '■ 薬効分類別 入院薬剤費'
    ws[f'A{table_start}'].font = subtitle_font

    # 推移月ラベル生成
    all_months = sorted(set(h['年月'] for h in history))
    all_months = [m for m in all_months if m != current_ym]
    # 最新5ヶ月（+ 当月で6ヶ月分）
    if len(all_months) > 5:
        all_months = all_months[-5:]

    headers = ['薬効分類', '当月金額', '構成比']
    for ym in all_months:
        # 'YYYY-MM' → 'M月' 表示
        y_str, m_str = ym.split('-')
        headers.append(f'{int(m_str)}月')
    headers.append('前月比')

    h_row = table_start + 1
    write_header_row(ws, h_row, 1, headers)

    # 推移の月別・分類別金額を辞書化
    hist_map = defaultdict(lambda: defaultdict(int))
    for h in history:
        if h['年月'] != current_ym:
            hist_map[h['年月']][h['薬効分類']] += h['金額']

    r = h_row + 1
    for cat in CLASSIFICATION_ORDER:
        cat_amt = agg.get(cat, {}).get('total_amt', 0)
        cat_ratio = cat_amt / grand_total if grand_total > 0 else 0
        row_vals = [cat, cat_amt, cat_ratio]
        fmts = [None, YEN, PCT]
        for ym in all_months:
            row_vals.append(hist_map[ym].get(cat, 0))
            fmts.append(YEN)
        # 前月比
        if all_months:
            prev_val = hist_map[all_months[-1]].get(cat, 0)
            if prev_val > 0:
                row_vals.append((cat_amt - prev_val) / prev_val)
            else:
                row_vals.append(None)
        else:
            row_vals.append(None)
        fmts.append('+0.0%;-0.0%')

        write_data_row(ws, r, 1, row_vals, fmt_list=fmts)
        r += 1

    # 合計行
    total_vals = ['合計', grand_total, 1.0]
    fmts_total = [None, YEN, PCT]
    for ym in all_months:
        month_total = sum(hist_map[ym].values())
        total_vals.append(month_total)
        fmts_total.append(YEN)
    if prev_month_total and prev_month_total > 0:
        total_vals.append((grand_total - prev_month_total) / prev_month_total)
    else:
        total_vals.append(None)
    fmts_total.append('+0.0%;-0.0%')
    write_data_row(ws, r, 1, total_vals, fmt_list=fmts_total, is_total=True)

    num_cols = len(headers)
    set_outer_border(ws, h_row, 1, r, num_cols)

    # 条件付き書式: 前月比列
    last_col_letter = get_column_letter(num_cols)
    pct_range = f'{last_col_letter}{h_row+1}:{last_col_letter}{r}'
    ws.conditional_formatting.add(pct_range,
        CellIsRule(operator='lessThan', formula=['-0.05'],
                   fill=cf_red, font=Font(size=10)))
    ws.conditional_formatting.add(pct_range,
        CellIsRule(operator='greaterThan', formula=['0.05'],
                   fill=cf_blue, font=Font(size=10)))
    ws.conditional_formatting.add(pct_range,
        CellIsRule(operator='lessThan', formula=['-0.1'],
                   fill=cf_red_strong, font=Font(bold=True, size=10)))
    ws.conditional_formatting.add(pct_range,
        CellIsRule(operator='greaterThan', formula=['0.1'],
                   fill=cf_blue_strong, font=Font(bold=True, size=10)))

    # データバー: 当月金額列 (B列)
    amt_range = f'B{h_row+1}:B{r}'
    ws.conditional_formatting.add(amt_range,
        DataBarRule(start_type='min', end_type='max',
                    color='4472C4', showValue=True))

    # 列幅
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 10
    for ci in range(4, num_cols + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 14
    setup_print(ws, f'薬剤月次レポート {month}月度')


def build_detail_sheet(wb, agg, grand_total):
    """Sheet2: 薬効分類別詳細"""
    ws = wb.create_sheet('薬効分類別詳細')
    ws.sheet_properties.tabColor = '4472C4'

    ws.merge_cells('A1:G1')
    ws['A1'] = '薬効分類別 薬剤詳細一覧'
    ws['A1'].font = title_font

    headers = ['診区', 'コード', '薬名', '単価', '入院使用量', '入院金額', '構成比']
    r = 3
    running_total = 0

    for cat in CLASSIFICATION_ORDER:
        data = agg.get(cat, {})
        drugs_in_cat = data.get('drugs', [])
        if not drugs_in_cat:
            continue

        # 分類ヘッダー
        ws.merge_cells(f'A{r}:G{r}')
        # 同一薬名の重複を除いた品目数
        unique_count = len(set(d['name'] for d in drugs_in_cat))
        ws[f'A{r}'] = f'■ {cat}（{unique_count}品目）'
        ws[f'A{r}'].font = section_font
        ws[f'A{r}'].fill = PatternFill(start_color='E8EEF7', end_color='E8EEF7', fill_type='solid')
        ws[f'A{r}'].border = thin_border
        for ci in range(2, 8):
            ws.cell(row=r, column=ci).fill = PatternFill(
                start_color='E8EEF7', end_color='E8EEF7', fill_type='solid')
            ws.cell(row=r, column=ci).border = thin_border
        r += 1

        write_header_row(ws, r, 1, headers, fill=sub_header_fill)
        r += 1

        cat_total = 0
        # 同一薬名を合算（診区21/22等で分かれている場合）
        merged = {}
        for d in drugs_in_cat:
            key = d['name']
            if key in merged:
                merged[key]['inpatient_qty'] += d['inpatient_qty']
                merged[key]['inpatient_amt'] += d['inpatient_amt']
                # 診区は複数あれば結合表示
                if d['shinku'] not in merged[key]['shinku_set']:
                    merged[key]['shinku_set'].add(d['shinku'])
            else:
                merged[key] = {
                    'name': d['name'],
                    'code': d['code'],
                    'unit_price': d['unit_price'],
                    'inpatient_qty': d['inpatient_qty'],
                    'inpatient_amt': d['inpatient_amt'],
                    'shinku_set': {d['shinku']},
                }
        sorted_drugs = sorted(merged.values(), key=lambda x: x['name'])
        for d in sorted_drugs:
            shinku_names = '/'.join(
                SHINKU_MAP.get(s, s) for s in sorted(d['shinku_set'])
            )
            ratio = d['inpatient_amt'] / grand_total if grand_total > 0 else 0
            cat_total += d['inpatient_amt']
            write_data_row(ws, r, 1, [
                shinku_names, d['code'], d['name'], d['unit_price'],
                d['inpatient_qty'], d['inpatient_amt'], ratio,
            ], fmt_list=[None, None, None, '#,##0.00', '#,##0.00', YEN, PCT])
            r += 1

        # 小計行
        running_total += cat_total
        cat_ratio = cat_total / grand_total if grand_total > 0 else 0
        write_data_row(ws, r, 1, [
            '', '', f'{cat} 小計', '', '', cat_total, cat_ratio,
        ], fmt_list=[None, None, None, None, None, YEN, PCT], is_sub=True)
        r += 2  # 空行

    # 総合計行
    write_data_row(ws, r, 1, [
        '', '', '総合計', '', '', grand_total, 1.0,
    ], fmt_list=[None, None, None, None, None, YEN, PCT], is_total=True)

    # 列幅
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 38
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 10
    setup_print(ws, '薬効分類別詳細')
    return ws


def build_trend_sheet(wb, agg, year, month, history, current_ym):
    """Sheet3: 推移グラフ用データ"""
    ws = wb.create_sheet('推移グラフ用データ')
    ws.sheet_properties.tabColor = '00B050'

    ws['A1'] = '年月'
    ws['B1'] = '薬効分類'
    ws['C1'] = '金額'
    for ci in range(1, 4):
        cell = ws.cell(row=1, column=ci)
        cell.font = header_font_w
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # 過去データ（当月除外・最大5ヶ月）
    past = [h for h in history if h['年月'] != current_ym]
    past_months = sorted(set(h['年月'] for h in past))
    if len(past_months) > 5:
        cutoff_months = set(past_months[-5:])
        past = [h for h in past if h['年月'] in cutoff_months]

    r = 2
    for h in sorted(past, key=lambda x: (x['年月'], x['薬効分類'])):
        ws.cell(row=r, column=1, value=h['年月']).font = data_font
        ws.cell(row=r, column=1).border = thin_border
        ws.cell(row=r, column=2, value=h['薬効分類']).font = data_font
        ws.cell(row=r, column=2).border = thin_border
        cell = ws.cell(row=r, column=3, value=h['金額'])
        cell.font = data_font
        cell.number_format = NUM
        cell.border = thin_border
        r += 1

    # 当月データ
    for cat in CLASSIFICATION_ORDER:
        cat_amt = agg.get(cat, {}).get('total_amt', 0)
        ws.cell(row=r, column=1, value=current_ym).font = data_font
        ws.cell(row=r, column=1).border = thin_border
        ws.cell(row=r, column=2, value=cat).font = data_font
        ws.cell(row=r, column=2).border = thin_border
        cell = ws.cell(row=r, column=3, value=cat_amt)
        cell.font = data_font
        cell.number_format = NUM
        cell.border = thin_border
        r += 1

    # オートフィルター
    ws.auto_filter.ref = f'A1:C{r-1}'

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 16
    setup_print(ws, '推移グラフ用データ')
    return ws


def build_settings_sheet(wb, drugs, agg):
    """Sheet4: 設定"""
    ws = wb.create_sheet('設定')
    ws.sheet_properties.tabColor = '808080'

    # マッピング一覧
    ws.merge_cells('A1:C1')
    ws['A1'] = '薬効分類マッピング設定'
    ws['A1'].font = title_font

    ws['A3'] = '■ 分類キーワード一覧'
    ws['A3'].font = subtitle_font

    write_header_row(ws, 4, 1, ['薬名キーワード', '分類先', '備考'])
    r = 5
    for cat in CLASSIFICATION_ORDER:
        if cat == 'その他':
            continue
        keywords = DRUG_CLASSIFICATION.get(cat, [])
        for kw in keywords:
            ws.cell(row=r, column=1, value=kw).font = data_font
            ws.cell(row=r, column=1).border = thin_border
            ws.cell(row=r, column=2, value=cat).font = data_font
            ws.cell(row=r, column=2).border = thin_border
            ws.cell(row=r, column=3).border = thin_border
            r += 1

    r += 2

    # 未分類薬剤リスト
    ws[f'A{r}'] = '■ 未分類薬剤リスト（「その他」に分類された薬剤）'
    ws[f'A{r}'].font = subtitle_font
    r += 1

    write_header_row(ws, r, 1, ['薬名', '診区', '入院金額'])
    r += 1

    other_drugs = agg.get('その他', {}).get('drugs', [])
    other_drugs_sorted = sorted(other_drugs, key=lambda x: x['inpatient_amt'], reverse=True)
    for d in other_drugs_sorted:
        shinku_name = SHINKU_MAP.get(d['shinku'], d['shinku'])
        ws.cell(row=r, column=1, value=d['name']).font = data_font
        ws.cell(row=r, column=1).border = thin_border
        ws.cell(row=r, column=2, value=shinku_name).font = data_font
        ws.cell(row=r, column=2).border = thin_border
        cell = ws.cell(row=r, column=3, value=d['inpatient_amt'])
        cell.font = data_font
        cell.number_format = YEN
        cell.border = thin_border
        r += 1

    ws.column_dimensions['A'].width = 42
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 16
    setup_print(ws, '設定')
    return ws


# ======================================================================
# HTML サマリー生成
# ======================================================================

_COMMON_CSS = """\
@media print{@page{size:A4 portrait;margin:6mm}*{-webkit-print-color-adjust:exact!important;print-color-adjust:exact!important}body{padding:0!important;font-size:7px!important}.no-print{display:none!important}.summary-cards{gap:8px!important;margin-bottom:10px!important}.card{padding:10px!important}.card-value{font-size:18px!important}table{font-size:7.5px!important}th,td{padding:4px 5px!important}}
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:'Inter','Noto Sans JP',sans-serif;font-size:11px;color:#1e293b;background:linear-gradient(135deg,#f8fafc 0%,#e2e8f0 100%);min-height:100vh;padding:20px}
.container{max-width:1000px;margin:0 auto;background:#fff;border-radius:16px;box-shadow:0 4px 24px rgba(0,0,0,0.08);padding:24px}
.header{display:flex;justify-content:space-between;align-items:center;margin-bottom:20px;padding-bottom:16px;border-bottom:1px solid #e2e8f0}
h1{font-size:22px;font-weight:700;color:#0f172a;display:flex;align-items:center;gap:12px}
h1::before{content:'';width:5px;height:28px;background:linear-gradient(180deg,#3b82f6,#1d4ed8);border-radius:3px}
.print-date{font-size:11px;color:#64748b}
.btn-group{display:flex;gap:10px}
.btn{padding:10px 20px;color:#fff;border:none;border-radius:8px;cursor:pointer;font-size:13px;font-weight:600;transition:all 0.2s;display:flex;align-items:center;gap:6px}
.btn-print{background:linear-gradient(135deg,#6366f1,#4f46e5)}
.btn-print:hover{transform:translateY(-2px);box-shadow:0 4px 12px rgba(99,102,241,0.4)}
.btn-save{background:linear-gradient(135deg,#10b981,#059669)}
.btn-save:hover{transform:translateY(-2px);box-shadow:0 4px 12px rgba(16,185,129,0.4)}
.summary-cards{display:grid;grid-template-columns:repeat(4,1fr);gap:16px;margin-bottom:24px}
.card{background:#fff;border-radius:12px;padding:18px;position:relative;overflow:hidden;border:1px solid #e2e8f0;transition:transform 0.2s,box-shadow 0.2s}
.card:hover{transform:translateY(-2px);box-shadow:0 8px 24px rgba(0,0,0,0.1)}
.card::before{content:'';position:absolute;top:0;left:0;right:0;height:4px}
.card-revenue::before{background:linear-gradient(90deg,#3b82f6,#60a5fa)}
.card-expense::before{background:linear-gradient(90deg,#f59e0b,#fbbf24)}
.card-profit::before{background:linear-gradient(90deg,#10b981,#34d399)}
.card-margin::before{background:linear-gradient(90deg,#8b5cf6,#a78bfa)}
.card-label{font-size:11px;color:#64748b;font-weight:500;margin-bottom:8px;display:flex;align-items:center;gap:6px}
.card-icon{width:20px;height:20px;border-radius:6px;display:flex;align-items:center;justify-content:center;font-size:11px}
.card-revenue .card-icon{background:#dbeafe;color:#2563eb}
.card-expense .card-icon{background:#fef3c7;color:#d97706}
.card-profit .card-icon{background:#d1fae5;color:#059669}
.card-margin .card-icon{background:#ede9fe;color:#7c3aed}
.card-value{font-size:26px;font-weight:700;color:#0f172a;font-variant-numeric:tabular-nums;letter-spacing:-0.5px}
.card-sub{display:flex;align-items:center;gap:8px;margin-top:8px;font-size:11px}
.card-period{color:#94a3b8;font-size:10px}
table{width:100%;border-collapse:collapse;font-size:9px;border-radius:12px;overflow:hidden;box-shadow:0 2px 8px rgba(0,0,0,0.06)}
thead th{background:linear-gradient(135deg,#1e3a8a 0%,#1e40af 100%);color:#fff;padding:10px 8px;text-align:center;font-weight:600;letter-spacing:0.3px;position:relative}
thead th::after{content:'';position:absolute;bottom:0;left:0;right:0;height:1px;background:rgba(255,255,255,0.1)}
th:first-child{text-align:left;padding-left:14px;min-width:200px;border-radius:8px 0 0 0}
th:last-child{border-radius:0 8px 0 0}
td{padding:8px 10px;border-bottom:1px solid #f1f5f9;text-align:right;font-family:'Inter',monospace;font-variant-numeric:tabular-nums;color:#334155}
td:first-child{text-align:left;padding-left:14px;font-family:'Noto Sans JP',sans-serif;font-weight:500;color:#1e293b}
tbody tr{background:#fff;transition:background 0.15s}
tbody tr:nth-child(even){background:#fafbfc}
tbody tr:hover{background:#f1f5f9}
.indent{padding-left:28px!important;color:#64748b;font-weight:400!important}
.indent2{padding-left:42px!important;color:#94a3b8;font-weight:400!important;font-size:8.5px}
.section{background:linear-gradient(90deg,#eff6ff,#f8fafc)!important}
.section td{font-weight:600}
.section td:first-child{color:#1d4ed8}
.subtotal{background:#f8fafc!important}
.subtotal td{font-weight:600;color:#1e3a8a}
.grand{background:linear-gradient(90deg,#ecfdf5,#f0fdf4)!important}
.grand td{font-weight:700;color:#047857}
.final{background:linear-gradient(90deg,#eff6ff,#dbeafe)!important}
.final td{font-weight:700;color:#0f172a;font-size:10px}
.negative{color:#dc2626!important}
.pct{color:#94a3b8;font-size:7px;display:block;margin-top:2px}
tbody tr:last-child td:first-child{border-radius:0 0 0 8px}
tbody tr:last-child td:last-child{border-radius:0 0 8px 0}
"""


def _man_yen(yen):
    """1650865 -> '165万円'"""
    return f'{round(yen / 10000):,}万円'


def _yen_row(label, amt, total, cls='', td_cls='', qty=None):
    """円ベースのテーブル行を生成"""
    import html as _html
    pct = f'{amt / total * 100:.1f}%' if total and amt else ''
    tr = f' class="{cls}"' if cls else ''
    td = f' class="{td_cls}"' if td_cls else ''
    qty_cell = f'{qty:,.1f}'.rstrip('0').rstrip('.') if qty else ''
    return f'  <tr{tr}><td{td}>{_html.escape(str(label))}</td><td>{qty_cell}</td><td>{amt:,.0f}<span class="pct">円</span></td><td>{pct}</td></tr>'


def generate_drug_html(agg, drugs, grand_total, year, month, history, current_ym):
    """薬剤月次サマリーHTMLを生成"""
    import html as _html

    report_date = f'令和{year - 2018}年{datetime.date.today().month}月{datetime.date.today().day}日'

    # 精神科薬剤合計
    psych_categories = [
        '抗精神病薬（非定型）', '抗精神病薬（定型）', '抗うつ薬', '睡眠薬',
        '抗不安薬', '気分安定薬・抗てんかん薬', '抗パーキンソン薬', 'ADHD治療薬',
        '認知症治療薬',
    ]
    psych_total = sum(agg.get(c, {}).get('total_amt', 0) for c in psych_categories)
    psych_pct = psych_total / grand_total * 100 if grand_total else 0

    # 品目数
    unique_drugs = len(set(d['name'] for d in drugs))

    # 前月比
    prev_month_total = None
    if history:
        months_in_history = sorted(set(h['年月'] for h in history))
        months_in_history = [m for m in months_in_history if m != current_ym]
        if months_in_history:
            prev_ym = months_in_history[-1]
            prev_month_total = sum(h['金額'] for h in history if h['年月'] == prev_ym)

    if prev_month_total and prev_month_total > 0:
        delta_pct = (grand_total - prev_month_total) / prev_month_total * 100
        delta_sign = '+' if delta_pct >= 0 else ''
        card4_value = f'{delta_sign}{delta_pct:.1f}%'
        card4_sub = f'前月 {prev_month_total:,.0f}円'
    else:
        card4_value = '-'
        card4_sub = '前月データなし'

    # --- テーブル行 ---
    rows = []
    for cat in CLASSIFICATION_ORDER:
        cat_data = agg.get(cat, {})
        if not cat_data or cat_data['total_amt'] == 0:
            continue
        # セクション行
        rows.append(_yen_row(cat, cat_data['total_amt'], grand_total, cls='section'))
        # 全品目（薬剤名順）
        sorted_drugs = sorted(cat_data['drugs'],
                              key=lambda d: d['name'])
        for d in sorted_drugs:
            if d['inpatient_amt'] > 0:
                shinku_label = SHINKU_MAP.get(d['shinku'], d['shinku'])
                display_name = strip_maker_name(d['name'])
                label = f'[{shinku_label}] {display_name}'
                rows.append(_yen_row(label, d['inpatient_amt'], grand_total,
                                     td_cls='indent',
                                     qty=d.get('inpatient_qty', 0)))

    # 合計
    rows.append(f'  <tr class="grand"><td>合計</td><td></td><td>{grand_total:,.0f}'
                f'<span class="pct">円</span></td><td>100.0%</td></tr>')

    table_body = '\n'.join(rows)
    save_name = f'薬剤月次サマリー_{month}月.html'

    html_text = f"""\
<!DOCTYPE html>
<html lang="ja">
<head>
<meta charset="UTF-8">
<title>さきがけホスピタル 薬剤月次サマリー {month}月度</title>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&family=Noto+Sans+JP:wght@400;500;600;700&display=swap" rel="stylesheet">
<style>
{_COMMON_CSS}</style>
</head>
<body>
<div class="container">

<div class="header">
  <h1>さきがけホスピタル 薬剤月次サマリー {month}月度</h1>
  <div class="btn-group no-print">
    <button class="btn btn-save" onclick="saveHTML()">💾 HTMLを保存</button>
    <button class="btn btn-print" onclick="window.print()">🖨️ 印刷 / PDF</button>
  </div>
  <div class="print-date">{report_date}作成</div>
</div>

<div class="summary-cards">
  <div class="card card-revenue">
    <div class="card-label"><span class="card-icon">💊</span>入院薬剤費合計</div>
    <div class="card-value">{_man_yen(grand_total)}</div>
    <div class="card-sub"><span class="card-period">{grand_total:,.0f}円</span></div>
  </div>
  <div class="card card-expense">
    <div class="card-label"><span class="card-icon">🧠</span>精神科薬剤</div>
    <div class="card-value">{_man_yen(psych_total)}</div>
    <div class="card-sub"><span class="card-period">{psych_total:,.0f}円 ／ 構成比 {psych_pct:.1f}%</span></div>
  </div>
  <div class="card card-profit">
    <div class="card-label"><span class="card-icon">💉</span>品目数</div>
    <div class="card-value">{unique_drugs}</div>
    <div class="card-sub"><span class="card-period">全{len(drugs)}レコード</span></div>
  </div>
  <div class="card card-margin">
    <div class="card-label"><span class="card-icon">📊</span>前月比</div>
    <div class="card-value">{card4_value}</div>
    <div class="card-sub"><span class="card-period">{card4_sub}</span></div>
  </div>
</div>

<table>
<thead>
  <tr>
    <th>薬効分類</th>
    <th>使用量</th>
    <th>金額（円）</th>
    <th>構成比</th>
  </tr>
</thead>
<tbody>
{table_body}
</tbody>
</table>

</div>

<script>
function saveHTML(){{
  const h=document.documentElement.outerHTML;
  const b=new Blob([h],{{type:'text/html;charset=utf-8'}});
  const u=URL.createObjectURL(b);
  const a=document.createElement('a');
  a.href=u;
  a.download='{save_name}';
  a.click();
  URL.revokeObjectURL(u);
}}
</script>
</body>
</html>
"""

    month_dir = os.path.join(BASE_DIR, f'{month}月')
    if not os.path.isdir(month_dir):
        month_dir = BASE_DIR
    html_path = os.path.join(month_dir, save_name)
    with open(html_path, 'w', encoding='utf-8') as f:
        f.write(html_text)
    return html_path


# ======================================================================
# JSON サイドカー（前月比用データ保存・読み込み）
# ======================================================================

def save_month_data(month_dir, year, month, agg, grand_total, drugs):
    """月次集計データをJSONに保存（前月比計算用）"""
    data = {
        'year': year,
        'month': month,
        'grand_total': grand_total,
        'record_count': len(drugs),
        'unique_drugs': len(set(d['name'] for d in drugs)),
        'categories': {},
    }
    for cat in CLASSIFICATION_ORDER:
        cat_data = agg.get(cat, {})
        data['categories'][cat] = cat_data.get('total_amt', 0)

    json_path = os.path.join(month_dir, f'薬剤月次データ_{month}月.json')
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    return json_path


def load_month_data(month_dir, month):
    """前月のJSONサイドカーデータを読み込み"""
    json_path = os.path.join(month_dir, f'薬剤月次データ_{month}月.json')
    if not os.path.exists(json_path):
        return None
    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return None


def find_prev_month_data(base_dir, year, month):
    """前月のデータを自動探索して読み込み"""
    if month == 1:
        prev_month = 12
        prev_year = year - 1
    else:
        prev_month = month - 1
        prev_year = year

    prev_dir = os.path.join(base_dir, f'{prev_month}月')
    data = load_month_data(prev_dir, prev_month)
    if data:
        return data

    # Excel推移データからのフォールバック
    history = load_history(os.path.join(base_dir, '薬剤月次レポート.xlsx'))
    if history:
        prev_ym = f'{prev_year}-{prev_month:02d}'
        prev_entries = [h for h in history if h['年月'] == prev_ym]
        if prev_entries:
            cats = {}
            for h in prev_entries:
                cats[h['薬効分類']] = cats.get(h['薬効分類'], 0) + h['金額']
            return {
                'year': prev_year,
                'month': prev_month,
                'grand_total': sum(cats.values()),
                'categories': cats,
            }
    return None


# ======================================================================
# PDFフォルダ自動検出
# ======================================================================

def find_drug_pdfs(base_dir, month_arg=None):
    """薬剤PDFのパスを自動検出する。

    Args:
        base_dir: プロジェクトルート
        month_arg: CLI引数（例: '2月', '2月/2月'）。Noneなら自動検出。

    Returns:
        tuple: (pdf_paths, month_dir) or (None, None) if not found
    """
    if month_arg:
        # 明示指定
        folder = os.path.join(base_dir, month_arg)
        pdf_paths = _find_pdfs_in_folder(folder)
        if pdf_paths:
            return pdf_paths, _get_month_dir(base_dir, month_arg)
        month_dir = _get_month_dir(base_dir, month_arg)
        # 「薬剤（全体）」を含むフォルダを優先的に探索（包括病棟データ含む）
        pdf_paths = _find_zentai_pdfs(folder)
        if pdf_paths:
            return pdf_paths, month_dir
        # その他サブフォルダも探索
        for sub in os.listdir(folder) if os.path.isdir(folder) else []:
            sub_path = os.path.join(folder, sub)
            if os.path.isdir(sub_path) and '薬剤（全体）' not in sub:
                pdf_paths = _find_pdfs_in_folder(sub_path)
                if pdf_paths:
                    return pdf_paths, month_dir
        return None, None

    # 自動検出: 月名フォルダを新しい順に探索
    month_dirs = []
    for name in os.listdir(base_dir):
        full = os.path.join(base_dir, name)
        if os.path.isdir(full) and re.match(r'^\d{1,2}月$', name):
            m = int(re.match(r'^(\d{1,2})月$', name).group(1))
            month_dirs.append((m, name, full))

    for _, name, full in sorted(month_dirs, key=lambda x: x[0], reverse=True):
        # 直下を探索
        pdf_paths = _find_pdfs_in_folder(full)
        if pdf_paths:
            return pdf_paths, full
        # 「薬剤（全体）」を含むフォルダを優先的に探索（包括病棟データ含む）
        pdf_paths = _find_zentai_pdfs(full)
        if pdf_paths:
            return pdf_paths, full
        # その他サブフォルダ（例: 2月/2月/, 1月/元データ/）
        for sub in sorted(os.listdir(full)):
            sub_path = os.path.join(full, sub)
            if os.path.isdir(sub_path) and '薬剤（全体）' not in sub:
                pdf_paths = _find_pdfs_in_folder(sub_path)
                if pdf_paths:
                    return pdf_paths, full
    return None, None


def _find_zentai_pdfs(folder, depth=0):
    """「薬剤（全体）」を含むサブフォルダを再帰的に探索してPDFを返す"""
    if depth > 3 or not os.path.isdir(folder):
        return []
    for sub in os.listdir(folder):
        sub_path = os.path.join(folder, sub)
        if os.path.isdir(sub_path) and '薬剤（全体）' in sub:
            pdf_paths = _find_pdfs_in_folder(sub_path)
            if pdf_paths:
                return pdf_paths
            # さらにネストされている場合（例: R8.1薬剤（全体）/R8.1薬剤（全体）/）
            result = _find_zentai_pdfs(sub_path, depth + 1)
            if result:
                return result
    return []


def _find_pdfs_in_folder(folder):
    """フォルダ内の薬剤PDFを検出"""
    if not os.path.isdir(folder):
        return []
    single = os.path.join(folder, '薬剤.pdf')
    if os.path.exists(single):
        return [single]
    found = sorted(_glob.glob(os.path.join(folder, '薬剤*.pdf')))
    if found:
        return found
    # 「全体薬剤①.pdf」等のパターンにも対応
    found = sorted(_glob.glob(os.path.join(folder, '*薬剤*.pdf')))
    return found if found else []


def _get_month_dir(base_dir, month_arg):
    """月ディレクトリのパスを返す（X月 レベル）"""
    parts = month_arg.replace('\\', '/').split('/')
    return os.path.join(base_dir, parts[0])


def get_month_from_folder(month_dir):
    """フォルダ名から対象月を取得。例: '1月' → 1, '2月' → 2"""
    folder_name = os.path.basename(month_dir)
    m = re.match(r'^(\d{1,2})月$', folder_name)
    if m:
        return int(m.group(1))
    return None


# ======================================================================
# メイン処理
# ======================================================================

def main():
    print('=' * 60)
    print('薬剤月次レポート生成')
    print('=' * 60)

    # 1. PDF検出
    month_arg = sys.argv[1] if len(sys.argv) > 1 else None
    pdf_paths, month_dir = find_drug_pdfs(BASE_DIR, month_arg)

    if not pdf_paths:
        target = month_arg or '(自動検出)'
        print(f'ERROR: 薬剤PDFが見つかりません: {target}')
        return

    print(f'対象フォルダ: {month_dir}')

    # 2. PDF解析
    drugs = []
    for pp in pdf_paths:
        print(f'PDF読み込み: {os.path.basename(pp)}')
        d = extract_drugs_from_pdf(pp)
        print(f'  → {len(d)}件')
        drugs.extend(d)

    # 複数ファイルからの重複除去（ページ境界の重複を排除）
    seen = set()
    unique_drugs = []
    for d in drugs:
        key = (d['shinku'], d['code'])
        if key not in seen:
            seen.add(key)
            unique_drugs.append(d)
    drugs = unique_drugs
    print(f'  抽出薬剤数（重複除去後）: {len(drugs)}件')

    if not drugs:
        print('ERROR: 薬剤データを抽出できませんでした。')
        return

    # 3. 薬効分類
    classify_all_drugs(drugs)

    # 4. 集計
    agg = aggregate_by_classification(drugs)
    grand_total = sum(d['inpatient_amt'] for d in drugs)

    print(f'  入院金額合計: {grand_total:,.0f}円')
    print(f'  分類数: {len(agg)}')
    for cat in CLASSIFICATION_ORDER:
        data = agg.get(cat, {})
        if data:
            print(f'    {cat}: {data["total_amt"]:>12,.0f}円 ({data["count"]}品目)')

    # 対象月（フォルダ名優先、フォールバックはPDF内テキスト）
    folder_month = get_month_from_folder(month_dir)
    pdf_year, pdf_month = get_report_month_from_pdf(pdf_paths[0])
    if folder_month:
        month = folder_month
        year = pdf_year  # 年はPDFから取得（フォルダ名に年情報がないため）
        # PDFの月がフォルダ月+1の場合は請求月のため、年を調整
        if pdf_month == folder_month + 1 or (folder_month == 12 and pdf_month == 1):
            year = pdf_year if pdf_month != 1 else pdf_year - 1
    else:
        year, month = pdf_year, pdf_month
    current_ym = f'{year}-{month:02d}'
    print(f'  対象月: {year}年{month}月')

    # 5. 前月データ読み込み
    prev_data = find_prev_month_data(BASE_DIR, year, month)
    if prev_data:
        print(f'  前月データ: {prev_data["month"]}月 ({prev_data["grand_total"]:,.0f}円)')
    else:
        print('  前月データ: なし')

    # history互換形式に変換（generate_drug_html用）
    history = []
    if prev_data:
        prev_ym = f'{prev_data["year"]}-{prev_data["month"]:02d}'
        for cat, amt in prev_data.get('categories', {}).items():
            history.append({'年月': prev_ym, '薬効分類': cat, '金額': amt})

    # 6. HTML生成
    print('\n薬剤HTML生成中...')
    html_path = generate_drug_html(agg, drugs, grand_total, year, month, history, current_ym)
    print(f'OK: {html_path}')

    # 7. JSONサイドカー保存（次月の前月比用）
    json_path = save_month_data(month_dir, year, month, agg, grand_total, drugs)
    print(f'OK: {json_path}')

    # 検証
    total_classified = sum(
        len(agg.get(c, {}).get('drugs', [])) for c in CLASSIFICATION_ORDER
    )
    print(f'\n--- 検証 ---')
    print(f'  全薬剤数: {len(drugs)}')
    print(f'  分類済み: {total_classified}')
    print(f'  入院金額合計: {grand_total:,.0f}円')
    print(f'  PDFファイル数: {len(pdf_paths)}')


if __name__ == '__main__':
    main()
