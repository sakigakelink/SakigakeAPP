#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""月次入院収益レポート自動生成スクリプト

指定月フォルダ内の診療行為別集計表PDF群から入院データを自動抽出し、
前月比較付きの月次入院収益Excelレポートを生成する。

使用法:
    python generate_monthly_report.py

出力:
    入院収益月次レポート_M月.xlsx
"""

import os
import re
import sys
import glob
import datetime
from collections import defaultdict, Counter

import pdfplumber
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.formatting.rule import CellIsRule, DataBarRule
from openpyxl.utils import get_column_letter

# ======================================================================
# パス定義
# ======================================================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
MONTH_FOLDER = os.path.join(BASE_DIR, '2月', '2月')

# ======================================================================
# スタイル定数
# ======================================================================
thin = Side(style='thin')
thick = Side(style='medium')
thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
sub_header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
light_blue_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
total_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
kpi_green = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
kpi_red = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')
kpi_blue = PatternFill(start_color='DAEEF3', end_color='DAEEF3', fill_type='solid')
kpi_gray = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
toc_fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
cf_red = PatternFill(start_color='FFCCCC', end_color='FFCCCC')
cf_blue = PatternFill(start_color='CCE5FF', end_color='CCE5FF')
cf_red_strong = PatternFill(start_color='FF9999', end_color='FF9999')
cf_blue_strong = PatternFill(start_color='99CCFF', end_color='99CCFF')

title_font = Font(bold=True, size=16, color='2F5496')
subtitle_font = Font(bold=True, size=13, color='2F5496')
section_font = Font(bold=True, size=11, color='2F5496')
header_font_w = Font(bold=True, size=10, color='FFFFFF')
data_font = Font(size=10)
bold_font = Font(bold=True, size=10)
kpi_value_font = Font(bold=True, size=22, color='2F5496')
kpi_label_font = Font(size=9, color='666666')
kpi_delta_pos = Font(bold=True, size=14, color='0066CC')
kpi_delta_neg = Font(bold=True, size=14, color='CC0000')
link_font = Font(size=10, color='0563C1', underline='single')
back_link_font = Font(size=9, color='0563C1', underline='single')

NUM = '#,##0'
PCT = '0.0%'
YEN = '#,##0"円"'
PT = '#,##0"点"'
DPCT = '+0.0%;-0.0%'

FACILITY_NAME = "精神科病院"

# ======================================================================
# 12月参照データ（前月比較用ベースライン、単位: 点）
# ======================================================================
DEC_COMPARE = {
    '診察': 1585,
    '投薬（薬剤+調剤処方）': 166739,
    '注射（手技+薬剤）': 39901,
    '処置': 57710,
    '検査': 67407,
    '画像': 35028,
    '精神科専門': 647370,
    'ベースアップ評価料': 95087,
    'その他・器材': 8940,
}
DEC_WARD = {
    '1病棟（精神療養）': 2156836,
    '2病棟（精神15:1）': 1682299,
    '3病棟（精神15:1）': 2561733,
}
DEC_FOOD = 1040926
DEC_TOTAL = 8561561
SEIKYU_YEN = 78144242  # 2月請求額（円）※総括表（入院）合計

DEC_PSYCH = {
    '入院精神療法': 145970,
    '精神科作業療法': 488840,
    '医療保護入院等': 4500,
    '退院指導料': 2560,
    '治療抵抗性統合失調症': 1000,
    '退院前訪問指導料': 0,
    'その他': 4500,
}

DEC_W2_DETAIL = {
    '入院基本料15:1(期間加算含)': 1359185,
    '看護補助加算': 162284,
    '看護配置加算': 34975,
    '看護補助体制充実': 27980,
    '地域移行実施加算': 27980,
    '重度認知症加算': 44400,
    '身体合併症管理': 15750,
    '隔離室管理加算': 9680,
}

DEC_W3_DETAIL = {
    '入院基本料15:1(期間加算含)': 1535302,
    '看護補助加算': 208916,
    '特殊疾患入院施設管理': 630350,
    '看護配置加算': 45025,
    '看護補助体制充実': 35555,
    '地域移行実施加算': 36020,
    '身体合併症管理': 17550,
    '療養環境加算': 44675,
    '重度認知症加算': 6300,
}

# 参考ファイル準拠の12月値（差異分析シート用）
DEC_REF = {
    '診察_初診': 1585,
    '診察_指導': 0,
    '薬剤_全診区横断': 157298,
    '器材': 0,
    '調剤処方': 26902,
    '注射_手技': 22440,
    '処置': 57710,
    '検査_一般': 55302,
    '検査_心電図': 10855,
    '検査_心理': 1250,
    '画像_XP': 5628,
    '画像_CT': 29400,
    'ベースアップ': 95087,
    'その他': 8940,
}

# 1病棟12月明細（点）
DEC_W1_DETAIL = {
    '精神療養病棟入院料': 1994176,
    '重症者加算1': 106140,
    '重症者加算2': 930,
    '地域移行実施加算': 36000,
    '非定型抗精神病薬': 19590,
}


# ======================================================================
# 1. PDF解析モジュール
# ======================================================================

def _yen2pt(yen):
    """円→点 変換（四捨五入）"""
    return (yen + 5) // 10


def _parse_num(s):
    """カンマ付き数値文字列をintに変換"""
    if s is None:
        return 0
    s = str(s).strip().replace(',', '').replace(' ', '')
    if not s or s in ('-', '―'):
        return 0
    try:
        return int(float(s))
    except ValueError:
        return 0


def extract_grand_total(pdf_path):
    """PDFの総合計行から入院金額（円）を抽出
    テーブル構造: [診区, コード, 名称, 点/単価, 外来数量, 入院数量, 合計数量,
                   外来金額, 入院金額(idx8), 合計金額, 全体比率, グループ比率, ?, ?]
    """
    with pdfplumber.open(pdf_path) as pdf:
        page = pdf.pages[-1]
        # テーブル抽出を試みる
        tables = page.extract_tables({
            'vertical_strategy': 'lines',
            'horizontal_strategy': 'lines',
            'snap_tolerance': 5,
        })
        if not tables:
            tables = page.extract_tables({
                'vertical_strategy': 'text',
                'horizontal_strategy': 'text',
            })
        for table in tables:
            for row in table:
                if row is None or len(row) < 9:
                    continue
                cell2 = str(row[2] or '').strip()
                # 総合計行を検出（改行なしの独立行）
                if '総' in cell2 and '計' in cell2 and '\n' not in cell2:
                    return _parse_num(row[8])
                # 診区別小計が改行で結合されている場合、最終行を確認
                if '\n' in cell2:
                    lines = cell2.split('\n')
                    if any('総' in l and '計' in l for l in lines):
                        idx = next(i for i, l in enumerate(lines) if '総' in l and '計' in l)
                        vals = str(row[8] or '').split('\n')
                        val = _parse_num(vals[idx]) if idx < len(vals) else 0
                        if val > 0:
                            return val
                        # 総合計行のcol8が空の場合はテキストフォールバックへ
                        break

        # テーブルから取れない場合（薬剤・器材等）: テキストベースで抽出
        # 全ページを走査（総合計が最終ページに無い場合もある）
        for page2 in pdf.pages:
            text = page2.extract_text() or ''
            last_subtotal = 0
            for line in text.split('\n'):
                if '総' in line and '計' in line:
                    nums = [_parse_num(p) for p in re.findall(r'[\d,.]+', line)]
                    # テキスト行の数値配列:
                    # [外来数量(0), 入院数量, 合計数量, 外来金額(0), 入院金額, 合計金額, 比率...]
                    # 入院金額は常にindex 4
                    if len(nums) > 4:
                        return nums[4]
                # 総合計が空の場合のフォールバック: 診区別小計の合計を使う
                if '小計' in line:
                    nums = [_parse_num(p) for p in re.findall(r'[\d,.]+', line)]
                    if len(nums) > 4 and nums[4] > 0:
                        last_subtotal += nums[4]
        if last_subtotal > 0:
            return last_subtotal
    return 0


def extract_drug_subtotals(pdf_path):
    """薬剤PDFから診区別小計（円）を抽出
    テーブル構造: [診区, コード, 名称, 単価, 外来数量, 入院数量, 合計数量,
                   外来金額, 入院金額(idx8), 合計金額, 全体比率, ?, ?]
    """
    subtotals = {}
    shinku_codes = {'21', '22', '23', '31', '33', '40'}

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables({
                'vertical_strategy': 'lines',
                'horizontal_strategy': 'lines',
                'snap_tolerance': 5,
            })
            if not tables:
                tables = page.extract_tables({
                    'vertical_strategy': 'text',
                    'horizontal_strategy': 'text',
                })
            current_shinku = None
            for table in tables:
                for row in table:
                    if row is None or len(row) < 9:
                        continue
                    cell0 = str(row[0] or '').strip()
                    cell2 = str(row[2] or '').strip()
                    # 診区番号を追跡
                    if cell0 in shinku_codes:
                        current_shinku = cell0
                    # 改行結合セルの場合
                    if '\n' in cell0:
                        for l in cell0.split('\n'):
                            if l.strip() in shinku_codes:
                                current_shinku = l.strip()
                    # 診区別小計行
                    if '小計' in cell2 and '\n' not in cell2 and current_shinku:
                        subtotals[current_shinku] = _parse_num(row[8])
    return subtotals


def extract_psych_items(pdf_path):
    """精神科専門PDFから主要項目別の入院金額（円）を抽出"""
    items = {}
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ''
            for line in text.split('\n'):
                if '入院精神療法（１）' in line or '入院精神療法(I)' in line or '180018110' in line:
                    nums = [_parse_num(p) for p in re.findall(r'[\d,]+', line) if _parse_num(p) > 100]
                    if nums:
                        items['入院精神療法(I)'] = _find_amount(nums)
                elif '入院精神療法（２）' in line and '６月以内' in line or '180012010' in line:
                    nums = [_parse_num(p) for p in re.findall(r'[\d,]+', line) if _parse_num(p) > 100]
                    if nums:
                        items['入院精神療法(II)(6月以内)'] = _find_amount(nums)
                elif '入院精神療法（２）' in line and '６月超' in line or '180012110' in line:
                    nums = [_parse_num(p) for p in re.findall(r'[\d,]+', line) if _parse_num(p) > 100]
                    if nums:
                        items['入院精神療法(II)(6月超)'] = _find_amount(nums)
    return items


def _find_amount(nums):
    """数値リストから入院金額を推定（2回出現する最大値 or 最大値）
    PDFの行は [数量, 数量, 金額, 金額] のように同値ペアが2組出現する。
    金額ペアは通常、数量ペアより大きいため、最大ペアを返す。
    """
    if not nums:
        return 0
    c = Counter(nums)
    pairs = [val for val, cnt in c.items() if cnt >= 2]
    if pairs:
        return max(pairs)
    return max(nums)


def extract_admission_subtotals(pdf_path):
    """入院料PDFから診区別小計（円）を抽出（単一ファイル版）"""
    return _extract_admission_subtotals_from_paths([pdf_path])


def _extract_admission_subtotals_from_paths(pdf_paths):
    """入院料PDF群から診区別小計（円）を抽出
    テキストベースで解析し、診区番号を追跡して小計行の最大値（＝入院金額）を取得。
    分割PDFの場合、全ファイルを順番に連結して解析する。
    Returns: {'90': int, '92': int, '97': int}
    """
    subtotals = {}
    current_shinku = None

    for pdf_path in pdf_paths:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text() or ''
                for line in text.split('\n'):
                    stripped = line.strip()
                    if stripped.startswith('90 '):
                        current_shinku = '90'
                    elif stripped.startswith('92 '):
                        current_shinku = '92'
                    elif stripped.startswith('97 '):
                        current_shinku = '97'

                    if '小計' in line and current_shinku:
                        nums = [_parse_num(p) for p in re.findall(r'[\d,]+', line)]
                        large = [n for n in nums if n > 1000]
                        if large:
                            subtotals[current_shinku] = max(large)
    return subtotals


def extract_admission_items(pdf_path):
    """入院料PDFからベースアップ評価料の入院金額（円）を抽出"""
    baseup_total = 0
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ''
            for line in text.split('\n'):
                if 'ベースアップ評価料' in line:
                    nums = [_parse_num(p) for p in re.findall(r'[\d,]+', line)]
                    large = [n for n in nums if n > 100]
                    if large:
                        baseup_total += _find_amount(large)
    return baseup_total


def extract_ward_items(pdf_path):
    """病棟別PDFから加算項目別の入院金額（円）を抽出"""
    items = {}
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables({
                'vertical_strategy': 'lines',
                'horizontal_strategy': 'lines',
                'snap_tolerance': 5,
            })
            if not tables:
                tables = page.extract_tables({
                    'vertical_strategy': 'text',
                    'horizontal_strategy': 'text',
                })
            for table in tables:
                for row in table:
                    if row is None or len(row) < 8:
                        continue
                    cell0 = str(row[0] or '').strip()
                    row_text = ' '.join(str(c or '') for c in row)
                    if '診区' in row_text and 'コード' in row_text:
                        continue
                    if '小計' in row_text or '合計' in row_text:
                        continue
                    if '集計表' in row_text or '令和' in row_text:
                        continue
                    if cell0 not in ('90', '92'):
                        continue
                    name = str(row[2] or '').strip()
                    if not name:
                        continue
                    # 入院金額を取得 (index 8)
                    amt = _parse_num(row[8] if len(row) > 8 else row[-3])
                    if amt > 0:
                        items[name] = amt
    return items


def extract_exam_detail(pdf_path):
    """診察PDFから診区別サブカテゴリ（円）を抽出
    診区11=初診, 診区13=指導
    """
    detail = {'初診': 0, '指導': 0}
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ''
            current = None
            for line in text.split('\n'):
                stripped = line.strip()
                if stripped.startswith('11 '):
                    current = '初診'
                elif stripped.startswith('13 '):
                    current = '指導'
                if '小計' in line and current:
                    nums = [_parse_num(p) for p in re.findall(r'[\d,.]+', line)]
                    if len(nums) > 4:
                        detail[current] = nums[4]
    return detail


def extract_test_detail(pdf_path):
    """検査PDFからサブカテゴリ別入院金額（円）を抽出
    心電図・呼吸心拍等: ECG, EEG, 呼吸心拍監視, 脈波
    心理検査: 認知機能検査, 心理検査
    一般: それ以外
    テキスト行の数値配列(ASCII正規表現): idx0=診区, idx1=コード, idx2=点数, ..., idx7=入院金額
    """
    ecg_total = 0
    psych_test_total = 0
    general_total = 0

    ecg_keywords = ['ＥＣＧ', 'ECG', 'ＥＥＧ', 'EEG', '心電図', '呼吸心拍', '脈波', '賦活検査', '脳波']
    psych_keywords = ['認知機能検査', '心理検査', '知能検査', '発達検査']

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ''
            for line in text.split('\n'):
                stripped = line.strip()
                if not stripped.startswith('60 '):
                    continue
                if '小計' in line or '合計' in line or '810000001' in line:
                    continue
                # ASCII数字のみマッチ（全角数字を名前内で拾わないように）
                nums = [_parse_num(p) for p in re.findall(r'[0-9,.]+', line)]
                if len(nums) < 8:
                    continue
                amt = nums[7]  # 入院金額 = idx7
                if amt == 0:
                    continue

                if any(kw in line for kw in ecg_keywords):
                    ecg_total += amt
                elif any(kw in line for kw in psych_keywords):
                    psych_test_total += amt
                else:
                    general_total += amt

    return {'一般': general_total, '心電図': ecg_total, '心理': psych_test_total}


def extract_image_detail(pdf_path):
    """画像PDFからサブカテゴリ別入院金額（円）を抽出
    CT: CT撮影, コンピューター断層
    X-P: それ以外（単純撮影等）
    テキスト行の数値配列(ASCII正規表現): idx7=入院金額
    """
    ct_total = 0
    xp_total = 0
    ct_keywords = ['ＣＴ', 'CT', 'コンピューター断層', 'コンピュータ断層']

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ''
            for line in text.split('\n'):
                stripped = line.strip()
                if not stripped.startswith('70 '):
                    continue
                if '小計' in line or '合計' in line:
                    continue
                nums = [_parse_num(p) for p in re.findall(r'[0-9,.]+', line)]
                if len(nums) < 8:
                    continue
                amt = nums[7]  # 入院金額 = idx7
                if amt == 0:
                    continue

                if any(kw in line for kw in ct_keywords):
                    ct_total += amt
                else:
                    xp_total += amt

    return {'X-P': xp_total, 'CT': ct_total}


def extract_w1_detail(pdf_path):
    """入院料（全体）PDFから1病棟明細（円）を抽出
    診区92: 精神療養病棟入院料, 地域移行実施加算
    診区90: 重症者加算1, 重症者加算2(療養環境), 非定型抗精神病薬
    テキスト行の数値配列(ASCII正規表現): idx7=入院金額
    """
    items = {}

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ''
            current_shinku = None
            for line in text.split('\n'):
                stripped = line.strip()
                # 診区の追跡
                if stripped.startswith('90 '):
                    current_shinku = '90'
                elif stripped.startswith('92 '):
                    current_shinku = '92'
                elif stripped.startswith('97 '):
                    current_shinku = '97'
                elif '小計' in line or '合計' in line:
                    continue

                nums = [_parse_num(p) for p in re.findall(r'[0-9,.]+', line)]
                if len(nums) < 8:
                    continue
                amt = nums[7]  # 入院金額 = idx7
                if amt == 0:
                    continue

                if current_shinku == '92':
                    if '190055010' in line or '精神療養病棟入院料' in line:
                        items['精神療養病棟入院料'] = amt
                    elif '190127810' in line or '地域移行' in line:
                        items['地域移行実施加算'] = amt
                elif current_shinku == '90':
                    # 診区90のアイテムはコードベースで特定
                    # ※190105570は療養環境加算(3病棟)であり重症者加算2ではない
                    if '190151470' in line:
                        items['重症者加算1'] = amt
                    elif '190151270' in line:
                        items['非定型抗精神病薬'] = amt

    return items


def extract_ward_detail(pdf_path):
    """2/3病棟PDFから加算明細（円）を抽出しグループ化して返す。
    入院基本料+期間加算を合算、ベースアップ除外。
    テキスト行の数値配列(ASCII正規表現): idx7=入院金額
    """
    # コード→カテゴリのマッピング
    base_codes = {'190083810'}  # 入院基本料15:1
    period_codes = {'190085210', '190085410', '190085610', '190085710', '190085810'}  # 期間加算
    code_map = {
        '190103970': '看護補助加算',
        '190102070': '看護配置加算',
        '190832970': '看護補助体制充実',
        '190833070': '看護補助体制充実',
        '190118570': '重度認知症加算',
        '190107070': '隔離室管理加算',
        '190127910': '身体合併症管理',
        '190173110': '身体合併症管理',
        '190127810': '地域移行実施加算',
        '190101970': '特殊疾患入院施設管理',
        '190105570': '療養環境加算',
    }
    baseup_code = '180729410'

    result = {}
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ''
            for line in text.split('\n'):
                stripped = line.strip()
                if not stripped.startswith('90 '):
                    continue
                if '小計' in line or '合計' in line:
                    continue
                nums = [_parse_num(p) for p in re.findall(r'[0-9,.]+', line)]
                if len(nums) < 8:
                    continue
                amt = nums[7]
                if amt == 0:
                    continue
                # コードを取得 (nums[1]相当だがテキストから直接取得)
                code = re.findall(r'[0-9]{9}', line)
                if not code:
                    continue
                c = code[0]

                if baseup_code in line:
                    continue  # ベースアップ除外
                elif c in base_codes or c in period_codes:
                    result['入院基本料15:1(期間加算含)'] = result.get('入院基本料15:1(期間加算含)', 0) + amt
                elif c in code_map:
                    key = code_map[c]
                    result[key] = result.get(key, 0) + amt

    return result


def get_report_period(folder_path):
    """フォルダ内のPDFから対象期間を取得"""
    for fname in os.listdir(folder_path):
        if fname.endswith('.pdf'):
            fpath = os.path.join(folder_path, fname)
            try:
                with pdfplumber.open(fpath) as pdf:
                    text = pdf.pages[0].extract_text() or ''
                    m = re.search(r'令和\s*(\d+)\s*年\s*(\d+)\s*月', text)
                    if m:
                        reiwa = int(m.group(1))
                        month = int(m.group(2))
                        year = reiwa + 2018
                        return year, month
            except Exception:
                continue
    return datetime.date.today().year, datetime.date.today().month


# ======================================================================
# 2. データ組み立て
# ======================================================================

def find_pdf(folder, *candidates):
    """フォルダ内からファイル名候補でPDFを探す"""
    for name in candidates:
        path = os.path.join(folder, name)
        if os.path.exists(path):
            return path
    # グロブで探す
    for name in candidates:
        base = name.replace('.pdf', '')
        matches = glob.glob(os.path.join(folder, f'*{base}*'))
        if matches:
            return matches[0]
    return None


def find_pdfs(folder, *candidates):
    """フォルダ内からファイル名候補でPDF群を探す（分割ファイル対応）
    完全一致が1つ見つかればそれだけ返す。なければグロブで全一致を返す。
    """
    for name in candidates:
        path = os.path.join(folder, name)
        if os.path.exists(path):
            return [path]
    for name in candidates:
        base = name.replace('.pdf', '')
        matches = sorted(glob.glob(os.path.join(folder, f'{base}*.pdf')))
        if matches:
            return matches
    return []


def build_data(folder):
    """PDFフォルダから全データを組み立て"""
    data = {}

    # --- 各PDF総合計（円）--- 分割ファイル対応
    pdf_totals = {}
    pdf_map = {
        '診察': ['診察.pdf'],
        '薬剤': ['薬剤.pdf'],
        '調剤処方': ['調剤処方.pdf', '調剤.pdf'],
        '注射手技': ['注射手技.pdf', '注射.pdf'],
        '処置': ['処置.pdf'],
        '検査': ['検査.pdf'],
        '画像': ['画像.pdf'],
        '精神科専門': ['精神科専門.pdf'],
        '器材': ['器材.pdf'],
    }
    for key, fnames in pdf_map.items():
        paths = find_pdfs(folder, *fnames)
        if paths:
            total = sum(extract_grand_total(p) for p in paths)
            pdf_totals[key] = total
            if len(paths) > 1:
                print(f'  {key}: {total:>12,}円 ({len(paths)}ファイル合算)')
            else:
                print(f'  {key}: {total:>12,}円')
        else:
            pdf_totals[key] = 0
            print(f'  {key}: PDF未検出')

    # --- 薬剤 診区別小計（分割ファイル対応） ---
    drug_paths = find_pdfs(folder, '薬剤.pdf')
    drug_subtotals = {}
    for dp in drug_paths:
        subs = extract_drug_subtotals(dp)
        for k, v in subs.items():
            drug_subtotals[k] = drug_subtotals.get(k, 0) + v
    print(f'  薬剤診区別: {drug_subtotals} ({len(drug_paths)}ファイル)')

    # --- 入院料 ---
    # 単体ファイル（従来パターン: 完全一致のみ）
    adm_path = None
    for cand in ['入院料（全体）・食事.pdf', '入院料(全体)・食事.pdf',
                 '入院料（全体）・食事①.pdf', '入院料(全体)・食事①.pdf',
                 '入院料（全体）.pdf']:
        p = os.path.join(folder, cand)
        if os.path.exists(p):
            adm_path = p
            break
    adm_subtotals = {}
    adm_baseup = 0

    if adm_path:
        adm_subtotals = extract_admission_subtotals(adm_path)
        adm_baseup = extract_admission_items(adm_path)
    else:
        # 分割パターン: 入院料①.pdf + 入院料②・食事.pdf 等
        adm_split_paths = sorted(glob.glob(os.path.join(folder, '入院料*.pdf')))
        # 病棟別ファイルを除外
        adm_split_paths = [p for p in adm_split_paths
                           if '病棟' not in os.path.basename(p)]
        if adm_split_paths:
            print(f'  入院料（分割ファイル検出）: {len(adm_split_paths)}件')
            for sp in adm_split_paths:
                print(f'    → {os.path.basename(sp)}')
            # 全ファイルを連結して解析（診区が跨る場合に対応）
            adm_subtotals = _extract_admission_subtotals_from_paths(adm_split_paths)
            for sp in adm_split_paths:
                adm_baseup += extract_admission_items(sp)
            # w1_detail抽出用に最初のファイルをadm_pathに設定
            adm_path = adm_split_paths[0]

    print(f'  入院料診区別: {adm_subtotals}')
    print(f'  入院料ベースアップ: {adm_baseup:,}円')

    # 精神科専門のベースアップ（位置ベース抽出: idx7=入院金額）
    psych_path = find_pdf(folder, '精神科専門.pdf')
    psych_baseup = 0
    if psych_path:
        with pdfplumber.open(psych_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text() or ''
                for line in text.split('\n'):
                    stripped = line.strip()
                    if 'ベースアップ' in line and stripped.startswith('80 '):
                        nums = [_parse_num(p) for p in re.findall(r'[0-9,.]+', line)]
                        if len(nums) >= 8:
                            psych_baseup += nums[7]  # 入院金額 = idx7
    print(f'  精神科ベースアップ: {psych_baseup:,}円')
    total_baseup = adm_baseup + psych_baseup

    # --- 食事②.pdf がある場合の追加取得 ---
    food2_path = find_pdf(folder, '食事②.pdf', '食事2.pdf')
    if food2_path:
        food2_subs = extract_admission_subtotals(food2_path)
        food2_97 = food2_subs.get('97', 0)
        if food2_97 > 0:
            adm_subtotals['97'] = adm_subtotals.get('97', 0) + food2_97
            print(f'  食事②追加: {food2_97:,}円 → 食事合計: {adm_subtotals.get("97", 0):,}円')

    # --- 区分別データ（点） ---
    # 投薬 = 薬剤(内服+屯服+外用) + 調剤処方
    drug_touyaku = (drug_subtotals.get('21', 0) +
                    drug_subtotals.get('22', 0) +
                    drug_subtotals.get('23', 0))
    touyaku = _yen2pt(drug_touyaku + pdf_totals.get('調剤処方', 0))

    # 注射 = 注射手技 + 薬剤(注射+点滴)
    drug_chusha = drug_subtotals.get('31', 0) + drug_subtotals.get('33', 0)
    chusha = _yen2pt(pdf_totals.get('注射手技', 0) + drug_chusha)

    # 処置 = 処置 + 薬剤(処置薬)
    shochi = _yen2pt(pdf_totals.get('処置', 0) + drug_subtotals.get('40', 0))

    # 精神科専門（ベースアップ除く）
    seishin = _yen2pt(pdf_totals.get('精神科専門', 0) - psych_baseup)

    compare_data = {
        '診察': _yen2pt(pdf_totals.get('診察', 0)),
        '投薬（薬剤+調剤処方）': touyaku,
        '注射（手技+薬剤）': chusha,
        '処置': shochi,
        '検査': _yen2pt(pdf_totals.get('検査', 0)),
        '画像': _yen2pt(pdf_totals.get('画像', 0)),
        '精神科専門': seishin,
        'ベースアップ評価料': _yen2pt(total_baseup),
        'その他・器材': _yen2pt(pdf_totals.get('器材', 0)),
    }
    data['compare'] = compare_data

    # --- 病棟データ（点） ---
    adm90 = adm_subtotals.get('90', 0)
    adm92 = adm_subtotals.get('92', 0)
    food = adm_subtotals.get('97', 0)

    # 1病棟 = 診区92 + 診区90中の1病棟分加算
    # （重症者加算、非定型抗精神病薬加算は1病棟に帰属）
    # 概算: 全入院料(90+92) - ベースアップ - 食事 = 病棟合計
    ward_total_yen = adm90 + adm92 - adm_baseup

    # 各病棟の個別PDFから取得を試みる
    w2_path = find_pdf(folder, '入院料(2病棟）.pdf', '入院料（2病棟）.pdf',
                        '入院料(2病棟).pdf')
    w3_path = find_pdf(folder, '入院料（3病棟）.pdf', '入院料(3病棟).pdf')

    w2_total = 0
    w3_total = 0
    w2_items = {}
    w3_items = {}

    if w2_path:
        # 診区90小計のみ取得（食事97を除外）
        w2_subs = extract_admission_subtotals(w2_path)
        w2_total = w2_subs.get('90', 0)
        # ベースアップを除外
        w2_total -= extract_admission_items(w2_path)
        print(f'  2病棟: {w2_total:>12,}円')

    if w3_path:
        # 診区90小計のみ取得（食事97を除外）
        w3_subs = extract_admission_subtotals(w3_path)
        w3_total = w3_subs.get('90', 0)
        # ベースアップを除外
        w3_total -= extract_admission_items(w3_path)
        print(f'  3病棟: {w3_total:>12,}円')

    # 1病棟 = 全体 - 2病棟 - 3病棟
    w1_total = ward_total_yen - w2_total - w3_total
    print(f'  1病棟: {w1_total:>12,}円（逆算）')

    data['ward'] = {
        '1病棟（精神療養）': _yen2pt(w1_total),
        '2病棟（精神15:1）': _yen2pt(w2_total),
        '3病棟（精神15:1）': _yen2pt(w3_total),
    }
    data['food'] = _yen2pt(food)

    # --- 精神科専門 詳細（点） ---
    psych_detail = {}
    if psych_path:
        with pdfplumber.open(psych_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text() or ''
                for line in text.split('\n'):
                    stripped = line.strip()
                    if not stripped.startswith('80 '):
                        continue
                    if '小計' in line or '合計' in line:
                        continue
                    nums = [_parse_num(p) for p in re.findall(r'[0-9,.]+', line)]
                    if len(nums) < 8:
                        continue
                    amt = nums[7]  # 入院金額 = idx7
                    if amt == 0:
                        continue
                    if '入院精神療法' in line and ('（１）' in line or '(I)' in line or '180018110' in line):
                        psych_detail['入院精神療法'] = psych_detail.get('入院精神療法', 0) + amt
                    elif '入院精神療法' in line and ('（２）' in line or '(II)' in line or '180012010' in line or '180012110' in line):
                        psych_detail['入院精神療法'] = psych_detail.get('入院精神療法', 0) + amt
                    elif '精神科作業療法' in line or '180007410' in line:
                        psych_detail['精神科作業療法'] = amt
                    elif '医療保護入院等' in line or '180026410' in line:
                        psych_detail['医療保護入院等'] = amt
                    elif '退院指導' in line and '退院前' not in line:
                        psych_detail['退院指導料'] = amt
                    elif '治療抵抗性' in line:
                        psych_detail['治療抵抗性統合失調症'] = amt
                    elif '退院前訪問' in line:
                        psych_detail['退院前訪問指導料'] = amt
                    elif 'ベースアップ' not in line:
                        if '810000001' not in line and '診区別' not in line:
                            psych_detail.setdefault('その他', 0)
                            psych_detail['その他'] += amt

    # 円→点（四捨五入）
    data['psych_detail'] = {k: _yen2pt(v) for k, v in psych_detail.items()}

    # --- サブカテゴリ詳細データ ---
    # 診察サブカテゴリ
    exam_path = find_pdf(folder, '診察.pdf')
    if exam_path:
        exam_detail_yen = extract_exam_detail(exam_path)
        data['exam_detail'] = {k: _yen2pt(v) for k, v in exam_detail_yen.items()}
    else:
        data['exam_detail'] = {'初診': 0, '指導': 0}

    # 検査サブカテゴリ（分割ファイル対応）
    test_paths = find_pdfs(folder, '検査.pdf')
    if test_paths:
        combined_test = {'一般': 0, '心電図': 0, '心理': 0}
        for tp in test_paths:
            td = extract_test_detail(tp)
            for k, v in td.items():
                combined_test[k] = combined_test.get(k, 0) + v
        data['test_detail'] = {k: _yen2pt(v) for k, v in combined_test.items()}
    else:
        data['test_detail'] = {'一般': 0, '心電図': 0, '心理': 0}

    # 画像サブカテゴリ
    image_path = find_pdf(folder, '画像.pdf')
    if image_path:
        image_detail_yen = extract_image_detail(image_path)
        data['image_detail'] = {k: _yen2pt(v) for k, v in image_detail_yen.items()}
    else:
        data['image_detail'] = {'X-P': 0, 'CT': 0}

    # 薬剤全診区横断（全subtotalの合計）
    data['drug_total'] = _yen2pt(sum(drug_subtotals.values()))

    # 調剤処方（単独）
    data['dispensing'] = _yen2pt(pdf_totals.get('調剤処方', 0))

    # 注射手技のみ
    data['injection_tech'] = _yen2pt(pdf_totals.get('注射手技', 0))

    # 器材
    data['equipment'] = _yen2pt(pdf_totals.get('器材', 0))

    # 処置（処置PDF + 薬剤処置薬）
    data['procedure'] = _yen2pt(pdf_totals.get('処置', 0) + drug_subtotals.get('40', 0))

    # 1病棟明細
    if adm_path:
        w1_detail_yen = extract_w1_detail(adm_path)
        data['w1_detail'] = {k: _yen2pt(v) for k, v in w1_detail_yen.items()}
    else:
        data['w1_detail'] = {}

    # 入院精神療法 内訳（点）
    data['psych_therapy_detail'] = []
    if psych_path:
        with pdfplumber.open(psych_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text() or ''
                for line in text.split('\n'):
                    stripped = line.strip()
                    if not stripped.startswith('80 '):
                        continue
                    nums = [_parse_num(p) for p in re.findall(r'[0-9,.]+', line)]
                    if len(nums) < 8:
                        continue
                    amt = nums[7]  # 入院金額
                    cases = nums[5]  # 入院数量
                    if amt == 0:
                        continue
                    if '180018110' in line:
                        data['psych_therapy_detail'].append(
                            ('入院精神療法(I)', cases, _yen2pt(amt)))
                    elif '180012010' in line:
                        data['psych_therapy_detail'].append(
                            ('入院精神療法(II)(6月以内)', cases, _yen2pt(amt)))
                    elif '180012110' in line:
                        data['psych_therapy_detail'].append(
                            ('入院精神療法(II)(6月超)', cases, _yen2pt(amt)))

    # 2病棟・3病棟 加算明細（点）- PDFから抽出
    if w2_path:
        w2_detail_yen = extract_ward_detail(w2_path)
        data['w2_detail'] = {k: _yen2pt(v) for k, v in w2_detail_yen.items()}
    else:
        data['w2_detail'] = {}
    if w3_path:
        w3_detail_yen = extract_ward_detail(w3_path)
        data['w3_detail'] = {k: _yen2pt(v) for k, v in w3_detail_yen.items()}
    else:
        data['w3_detail'] = {}

    # 合計
    jan_total = (sum(compare_data.values()) +
                 sum(data['ward'].values()) +
                 data['food'])
    data['total'] = jan_total

    return data


# ======================================================================
# 3. Excelヘルパー
# ======================================================================

def set_outer_border(ws, min_row, min_col, max_row, max_col):
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            cell = ws.cell(row=r, column=c)
            l = thick if c == min_col else thin
            ri = thick if c == max_col else thin
            t = thick if r == min_row else thin
            b = thick if r == max_row else thin
            cell.border = Border(left=l, right=ri, top=t, bottom=b)


def write_header_row(ws, row, col_start, labels, fill=None):
    f = fill or header_fill
    for i, label in enumerate(labels):
        c = col_start + i
        cell = ws.cell(row=row, column=c, value=label)
        cell.font = header_font_w
        cell.fill = f
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border


def write_data_row(ws, row, col_start, values, fmt_list=None,
                   is_total=False, is_sub=False):
    fill = total_fill if is_total else (light_blue_fill if is_sub else None)
    font = bold_font if (is_total or is_sub) else data_font
    for i, val in enumerate(values):
        c = col_start + i
        cell = ws.cell(row=row, column=c, value=val)
        cell.font = font
        cell.border = thin_border
        if fill:
            cell.fill = fill
        if i == 0:
            cell.alignment = Alignment(horizontal='left', vertical='center')
        else:
            cell.alignment = Alignment(horizontal='right', vertical='center')
            if fmt_list and i < len(fmt_list) and fmt_list[i]:
                cell.number_format = fmt_list[i]


def add_back_link(ws, row=1, col=1):
    cell = ws.cell(row=row, column=col, value="< 目次に戻る")
    cell.hyperlink = "#目次!A1"
    cell.font = back_link_font


def setup_print(ws, title=""):
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(
        fitToPage=True)
    ws.oddHeader.left.text = FACILITY_NAME
    ws.oddHeader.center.text = title
    ws.oddFooter.center.text = "&P / &N"


def add_cf(ws, pct_range, delta_range):
    ws.conditional_formatting.add(pct_range,
        CellIsRule(operator='lessThan', formula=['-0.05'],
                   fill=cf_red, font=Font(size=10)))
    ws.conditional_formatting.add(pct_range,
        CellIsRule(operator='greaterThan', formula=['0.05'],
                   fill=cf_blue, font=Font(size=10)))
    ws.conditional_formatting.add(pct_range,
        CellIsRule(operator='lessThan', formula=['-0.1'],
                   fill=cf_red_strong, font=Font(bold=True, size=10)))
    ws.conditional_formatting.add(pct_range,
        CellIsRule(operator='greaterThan', formula=['0.1'],
                   fill=cf_blue_strong, font=Font(bold=True, size=10)))
    ws.conditional_formatting.add(delta_range,
        DataBarRule(start_type='min', end_type='max',
                    color='4472C4', showValue=True))


# ======================================================================
# 4. シート構築
# ======================================================================

def build_toc(wb, ws, month, report_date):
    """目次シート"""
    ws.title = '目次'
    ws.sheet_properties.tabColor = '2F5496'

    ws.merge_cells('B2:F2')
    ws['B2'] = f'入院収益月次レポート {month}月度'
    ws['B2'].font = Font(bold=True, size=20, color='2F5496')

    ws.merge_cells('B3:F3')
    ws['B3'] = f'{FACILITY_NAME}　{report_date}作成　（単位：点、1点=10円）'
    ws['B3'].font = Font(size=11, color='666666')

    sheets = [
        ('区分別内訳', f'{month}月 診療区分別・病棟別の詳細内訳'),
        ('ダッシュボード', f'{month}月 KPI・区分別・病棟別・精神科専門療法の経営概況'),
    ]

    write_header_row(ws, 5, 2, ['No.', 'シート名', '内容'])
    for i, (name, desc) in enumerate(sheets):
        r = 6 + i
        ws.cell(row=r, column=2, value=i + 1).font = data_font
        ws.cell(row=r, column=2).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=2).border = thin_border
        cell = ws.cell(row=r, column=3, value=name)
        cell.hyperlink = f"#'{name}'!A1"
        cell.font = link_font
        cell.border = thin_border
        ws.cell(row=r, column=4, value=desc).font = data_font
        ws.cell(row=r, column=4).border = thin_border

    set_outer_border(ws, 5, 2, 5 + len(sheets), 4)
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 6
    ws.column_dimensions['C'].width = 24
    ws.column_dimensions['D'].width = 50
    setup_print(ws, '目次')


def build_dashboard(wb, data, month):
    """ダッシュボードシート"""
    ws = wb.create_sheet('ダッシュボード')
    ws.sheet_properties.tabColor = '00B050'
    add_back_link(ws)

    ws.merge_cells('A2:E2')
    ws['A2'] = f'入院収益ダッシュボード {month}月度'
    ws['A2'].font = title_font

    jan = data['total']

    # KPIボックス（月合計のみ）
    ws.merge_cells('A4:C4')
    c = ws['A4']
    c.value = f'{month}月 入院合計'
    c.font = kpi_label_font
    c.fill = kpi_blue
    c.alignment = Alignment(horizontal='center')
    ws.merge_cells('A5:C6')
    c = ws['A5']
    c.value = jan
    c.font = kpi_value_font
    c.fill = kpi_blue
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.number_format = PT

    # 円換算
    ws.merge_cells('D4:E4')
    c = ws['D4']
    c.value = f'{month}月 円換算'
    c.font = kpi_label_font
    c.fill = kpi_gray
    c.alignment = Alignment(horizontal='center')
    ws.merge_cells('D5:E6')
    c = ws['D5']
    c.value = jan * 10
    c.font = kpi_value_font
    c.fill = kpi_gray
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.number_format = YEN

    # 区分別内訳
    ws.merge_cells('A8:B8')
    ws['A8'] = f'■ 区分別 {month}月（点）'
    ws['A8'].font = subtitle_font

    write_header_row(ws, 9, 1, ['区分', f'{month}月(点)'])
    r = 10
    cat_order = list(data['compare'].keys())
    cat_sum = 0
    for cat in cat_order:
        j = data['compare'].get(cat, 0)
        cat_sum += j
        write_data_row(ws, r, 1, [cat, j], fmt_list=[None, NUM])
        r += 1
    write_data_row(ws, r, 1, ['区分別 小計', cat_sum],
                   fmt_list=[None, NUM], is_sub=True)
    r += 1

    # 入院料・食事
    ward_j = sum(data['ward'].values())
    write_data_row(ws, r, 1, ['入院料 小計', ward_j],
                   fmt_list=[None, NUM], is_sub=True)
    r += 1
    fj = data['food']
    write_data_row(ws, r, 1, ['食事 小計', fj],
                   fmt_list=[None, NUM], is_sub=True)
    r += 1
    write_data_row(ws, r, 1, ['合計', jan],
                   fmt_list=[None, NUM], is_total=True)
    set_outer_border(ws, 9, 1, r, 2)

    # --- 1病棟（精神療養）内訳 ---
    r += 2
    ws[f'A{r}'] = '■ 1病棟（精神療養）内訳'
    ws[f'A{r}'].font = subtitle_font
    r += 1
    write_header_row(ws, r, 1, ['項目', f'{month}月(点)'])
    r += 1
    w1s = r
    w1_order = ['精神療養病棟入院料', '重症者加算1', '重症者加算2',
                '地域移行実施加算', '非定型抗精神病薬']
    for name in w1_order:
        j = data['w1_detail'].get(name, 0)
        write_data_row(ws, r, 1, [name, j], fmt_list=[None, NUM])
        r += 1
    j_w1 = sum(data['w1_detail'].values())
    write_data_row(ws, r, 1, ['小計', j_w1], fmt_list=[None, NUM], is_total=True)
    set_outer_border(ws, w1s - 1, 1, r, 2)

    # --- 2病棟（15対1）内訳 ---
    r += 2
    ws[f'A{r}'] = '■ 2病棟（15対1）内訳'
    ws[f'A{r}'].font = subtitle_font
    r += 1
    write_header_row(ws, r, 1, ['項目', f'{month}月(点)'])
    r += 1
    w2s = r
    for name in data['w2_detail']:
        j = data['w2_detail'][name]
        write_data_row(ws, r, 1, [name, j], fmt_list=[None, NUM])
        r += 1
    j_w2 = sum(data['w2_detail'].values())
    write_data_row(ws, r, 1, ['小計', j_w2], fmt_list=[None, NUM], is_total=True)
    set_outer_border(ws, w2s - 1, 1, r, 2)

    # --- 3病棟（15対1）内訳 ---
    r += 2
    ws[f'A{r}'] = '■ 3病棟（15対1）内訳'
    ws[f'A{r}'].font = subtitle_font
    r += 1
    write_header_row(ws, r, 1, ['項目', f'{month}月(点)'])
    r += 1
    w3s = r
    for name in data['w3_detail']:
        j = data['w3_detail'][name]
        write_data_row(ws, r, 1, [name, j], fmt_list=[None, NUM])
        r += 1
    j_w3 = sum(data['w3_detail'].values())
    write_data_row(ws, r, 1, ['小計', j_w3], fmt_list=[None, NUM], is_total=True)
    set_outer_border(ws, w3s - 1, 1, r, 2)

    # --- 入院料 小計 ---
    r += 1
    write_data_row(ws, r, 1, ['入院料 小計', ward_j],
                   fmt_list=[None, NUM], is_sub=True)
    set_outer_border(ws, r, 1, r, 2)
    r += 1

    # --- 食事 小計 ---
    write_data_row(ws, r, 1, ['食事 小計', data['food']],
                   fmt_list=[None, NUM], is_sub=True)
    set_outer_border(ws, r, 1, r, 2)
    r += 1

    # --- 合計 ---
    write_data_row(ws, r, 1, ['合計', jan],
                   fmt_list=[None, NUM], is_total=True)
    set_outer_border(ws, r, 1, r, 2)
    r += 1

    # --- 精神科専門療法 内訳 ---
    r += 1
    ws[f'A{r}'] = f'■ 精神科専門療法 内訳（{month}月）'
    ws[f'A{r}'].font = subtitle_font
    r += 1
    write_header_row(ws, r, 1, ['項目', f'{month}月(点)'])
    r += 1
    ps_start = r
    psych_order = ['入院精神療法', '精神科作業療法', '医療保護入院等',
                   '退院指導料', '治療抵抗性統合失調症', '退院前訪問指導料', 'その他']
    for name in psych_order:
        j = data['psych_detail'].get(name, 0)
        write_data_row(ws, r, 1, [name, j], fmt_list=[None, NUM])
        r += 1
    jan_psych_total = sum(data['psych_detail'].get(k, 0) for k in psych_order)
    write_data_row(ws, r, 1, ['合計（ベースアップ除く）', jan_psych_total],
                   fmt_list=[None, NUM], is_total=True)
    set_outer_border(ws, ps_start - 1, 1, r, 2)

    # 入院精神療法 内訳
    if data.get('psych_therapy_detail'):
        r += 2
        ws[f'A{r}'] = f'■ 入院精神療法 内訳（{month}月）'
        ws[f'A{r}'].font = subtitle_font
        r += 1
        write_header_row(ws, r, 1, ['項目', '件数', '金額(点)'])
        r += 1
        pt_start = r
        total_cases = 0
        total_amt = 0
        for name, cases, amt in data['psych_therapy_detail']:
            write_data_row(ws, r, 1, [name, cases, amt], fmt_list=[None, NUM, NUM])
            total_cases += cases
            total_amt += amt
            r += 1
        write_data_row(ws, r, 1, ['合計', total_cases, total_amt],
                       fmt_list=[None, NUM, NUM], is_total=True)
        set_outer_border(ws, pt_start - 1, 1, r, 3)

    ws.column_dimensions['A'].width = 30
    for c in 'BCDE':
        ws.column_dimensions[c].width = 16
    setup_print(ws, 'ダッシュボード')


def build_comparison(wb, data, month):
    """区分別比較シート"""
    ws = wb.create_sheet('区分別比較')
    ws.sheet_properties.tabColor = '4472C4'
    add_back_link(ws)

    ws.merge_cells('A2:E2')
    ws['A2'] = f'入院収益 区分別比較（単位：点）'
    ws['A2'].font = title_font

    write_header_row(ws, 4, 1, ['区分', '12月(点)', f'{month}月(点)', '増減(点)', '増減率'])
    r = 5
    for cat in DEC_COMPARE:
        d = DEC_COMPARE[cat]
        j = data['compare'].get(cat, 0)
        diff = j - d
        pct = diff / d if d else None
        write_data_row(ws, r, 1, [cat, d, j, diff, pct],
                       fmt_list=[None, NUM, NUM, NUM, PCT])
        r += 1

    ward_d = sum(DEC_WARD.values())
    ward_j = sum(data['ward'].values())
    write_data_row(ws, r, 1,
        ['入院料 小計', ward_d, ward_j, ward_j - ward_d,
         (ward_j - ward_d) / ward_d if ward_d else None],
        fmt_list=[None, NUM, NUM, NUM, PCT], is_sub=True)
    r += 1
    fd, fj = DEC_FOOD, data['food']
    write_data_row(ws, r, 1,
        ['食事 小計', fd, fj, fj - fd,
         (fj - fd) / fd if fd else None],
        fmt_list=[None, NUM, NUM, NUM, PCT], is_sub=True)
    r += 1
    dec, jan = DEC_TOTAL, data['total']
    write_data_row(ws, r, 1,
        ['合計', dec, jan, jan - dec,
         (jan - dec) / dec if dec else None],
        fmt_list=[None, NUM, NUM, NUM, PCT], is_total=True)
    set_outer_border(ws, 4, 1, r, 5)
    add_cf(ws, f'E5:E{r}', f'D5:D{r}')

    ws.column_dimensions['A'].width = 28
    for c in 'BCDE':
        ws.column_dimensions[c].width = 16
    setup_print(ws, '区分別比較')


def build_ward(wb, data, month):
    """病棟別内訳シート"""
    ws = wb.create_sheet('病棟別内訳')
    ws.sheet_properties.tabColor = 'ED7D31'
    add_back_link(ws)

    ws.merge_cells('A2:B2')
    ws['A2'] = f'病棟別 入院料内訳（ベースアップ除く、単位：点）  {month}月'
    ws['A2'].font = title_font

    # 総括
    ws['A4'] = '■ 総括'
    ws['A4'].font = subtitle_font
    write_header_row(ws, 5, 1, ['病棟', f'{month}月(点)'])
    r = 6
    for ward_name in data['ward']:
        j = data['ward'][ward_name]
        write_data_row(ws, r, 1, [ward_name, j], fmt_list=[None, NUM])
        r += 1
    wj = sum(data['ward'].values())
    write_data_row(ws, r, 1, ['合計', wj], fmt_list=[None, NUM], is_total=True)
    set_outer_border(ws, 5, 1, r, 2)

    # 1病棟
    r += 2
    ws[f'A{r}'] = '■ 1病棟（精神療養）主要加算'
    ws[f'A{r}'].font = subtitle_font
    r += 1
    write_header_row(ws, r, 1, ['項目', f'{month}月(点)'])
    r += 1
    w1s = r
    w1_order = ['精神療養病棟入院料', '重症者加算1', '重症者加算2',
                '地域移行実施加算', '非定型抗精神病薬']
    for name in w1_order:
        j = data['w1_detail'].get(name, 0)
        write_data_row(ws, r, 1, [name, j], fmt_list=[None, NUM])
        r += 1
    j_w1 = sum(data['w1_detail'].values())
    write_data_row(ws, r, 1, ['小計', j_w1], fmt_list=[None, NUM], is_total=True)
    set_outer_border(ws, w1s - 1, 1, r, 2)

    # 2病棟
    r += 2
    ws[f'A{r}'] = '■ 2病棟 主要加算'
    ws[f'A{r}'].font = subtitle_font
    r += 1
    write_header_row(ws, r, 1, ['項目', f'{month}月(点)'])
    r += 1
    w2s = r
    for name in data['w2_detail']:
        j = data['w2_detail'][name]
        write_data_row(ws, r, 1, [name, j], fmt_list=[None, NUM])
        r += 1
    set_outer_border(ws, w2s - 1, 1, r - 1, 2)

    # 3病棟
    r += 1
    ws[f'A{r}'] = '■ 3病棟 主要加算'
    ws[f'A{r}'].font = subtitle_font
    r += 1
    write_header_row(ws, r, 1, ['項目', f'{month}月(点)'])
    r += 1
    w3s = r
    for name in data['w3_detail']:
        j = data['w3_detail'][name]
        write_data_row(ws, r, 1, [name, j], fmt_list=[None, NUM])
        r += 1
    set_outer_border(ws, w3s - 1, 1, r - 1, 2)

    ws.column_dimensions['A'].width = 36
    ws.column_dimensions['B'].width = 16
    setup_print(ws, '病棟別内訳')


def build_psych(wb, data, month):
    """精神科専門療法シート"""
    ws = wb.create_sheet('精神科専門療法')
    ws.sheet_properties.tabColor = '7030A0'
    add_back_link(ws)

    ws.merge_cells('A2:B2')
    ws['A2'] = f'精神科専門療法 内訳（単位：点）  {month}月'
    ws['A2'].font = title_font

    write_header_row(ws, 4, 1, ['項目', f'{month}月(点)'])
    r = 5
    psych_order = ['入院精神療法', '精神科作業療法', '医療保護入院等',
                   '退院指導料', '治療抵抗性統合失調症', '退院前訪問指導料', 'その他']
    for name in psych_order:
        j = data['psych_detail'].get(name, 0)
        write_data_row(ws, r, 1, [name, j], fmt_list=[None, NUM])
        r += 1

    jan_psych_total = sum(data['psych_detail'].get(k, 0) for k in psych_order)
    write_data_row(ws, r, 1,
        ['合計（ベースアップ除く）', jan_psych_total],
        fmt_list=[None, NUM], is_total=True)
    set_outer_border(ws, 4, 1, r, 2)

    # 入院精神療法内訳
    if data.get('psych_therapy_detail'):
        r += 2
        ws[f'A{r}'] = f'■ 入院精神療法 内訳（{month}月）'
        ws[f'A{r}'].font = subtitle_font
        r += 1
        write_header_row(ws, r, 1, ['項目', '件数', '金額(点)'])
        r += 1
        total_cases = 0
        total_amt = 0
        for name, cases, amt in data['psych_therapy_detail']:
            write_data_row(ws, r, 1, [name, cases, amt], fmt_list=[None, NUM, NUM])
            total_cases += cases
            total_amt += amt
            r += 1
        write_data_row(ws, r, 1, ['合計', total_cases, total_amt],
                       fmt_list=[None, NUM, NUM], is_total=True)

    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 16
    setup_print(ws, '精神科専門療法')


def build_factors(wb, data, month):
    """変動要因まとめシート"""
    ws = wb.create_sheet('変動要因まとめ')
    ws.sheet_properties.tabColor = 'FF0000'
    add_back_link(ws)

    ws.merge_cells('A2:C2')
    ws['A2'] = f'主な変動要因まとめ（12月 → {month}月）'
    ws['A2'].font = title_font

    # 各カテゴリの増減を計算
    changes = []
    for cat in DEC_COMPARE:
        d = DEC_COMPARE[cat]
        j = data['compare'].get(cat, 0)
        changes.append((cat, j - d))
    for ward_name in DEC_WARD:
        d = DEC_WARD[ward_name]
        j = data['ward'].get(ward_name, 0)
        changes.append((ward_name, j - d))
    fd, fj = DEC_FOOD, data['food']
    changes.append(('食事', fj - fd))

    decreases = sorted([(n, v) for n, v in changes if v < 0], key=lambda x: x[1])
    increases = sorted([(n, v) for n, v in changes if v > 0], key=lambda x: x[1], reverse=True)
    dec_sum = sum(v for _, v in decreases)
    inc_sum = sum(v for _, v in increases)

    # 減収
    ws['A4'] = f'■ 減収要因（計 約 {dec_sum:+,}点）'
    ws['A4'].font = Font(bold=True, size=12, color='CC0000')
    write_header_row(ws, 5, 1, ['要因', '影響額(点)', '前月比'])
    r = 6
    for name, val in decreases:
        d_val = DEC_COMPARE.get(name, DEC_WARD.get(name, DEC_FOOD if name == '食事' else 0))
        pct = val / d_val if d_val else None
        write_data_row(ws, r, 1, [name, val, pct], fmt_list=[None, NUM, PCT])
        ws.cell(row=r, column=2).font = Font(size=10, color='CC0000')
        r += 1
    set_outer_border(ws, 5, 1, r - 1, 3)

    # 増収
    r += 1
    ws[f'A{r}'] = f'■ 増収要因（計 約 {inc_sum:+,}点）'
    ws[f'A{r}'].font = Font(bold=True, size=12, color='0066CC')
    r += 1
    write_header_row(ws, r, 1, ['要因', '影響額(点)', '前月比'])
    r += 1
    inc_start = r
    for name, val in increases:
        d_val = DEC_COMPARE.get(name, DEC_WARD.get(name, DEC_FOOD if name == '食事' else 0))
        pct = val / d_val if d_val else None
        write_data_row(ws, r, 1, [name, val, pct], fmt_list=[None, NUM, PCT])
        ws.cell(row=r, column=2).font = Font(size=10, color='0066CC')
        r += 1
    set_outer_border(ws, inc_start - 1, 1, r - 1, 3)

    # 総括
    jan = data['total']
    dec = DEC_TOTAL
    delta = jan - dec
    delta_yen = delta * 10
    delta_pct = delta / dec * 100 if dec else 0

    r += 1
    ws[f'A{r}'] = '■ 総括'
    ws[f'A{r}'].font = subtitle_font
    r += 1
    ws.merge_cells(f'A{r}:C{r + 3}')
    summary = (
        f"{month}月は12月比で {delta:+,}点（約{delta_yen // 10000:+,}万円、{delta_pct:+.1f}%）の"
        f"{'減収' if delta < 0 else '増収'}。\n"
    )
    # 主な減収要因
    if decreases:
        top_dec = decreases[0]
        summary += f"主因は{top_dec[0]}（{top_dec[1]:+,}点）"
        if len(decreases) > 1:
            summary += f"と{decreases[1][0]}（{decreases[1][1]:+,}点）"
        summary += "。\n"
    # 主な増収要因
    if increases:
        top_inc = increases[0]
        summary += f"{top_inc[0]}は{top_inc[1]:+,}点と堅調。"

    ws[f'A{r}'] = summary
    ws[f'A{r}'].alignment = Alignment(wrap_text=True, vertical='top')
    ws[f'A{r}'].font = Font(size=11)

    ws.column_dimensions['A'].width = 26
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 14
    setup_print(ws, '変動要因')


def build_analysis(wb, data, month):
    """区分別内訳シートを作成"""
    ws = wb.create_sheet('区分別内訳', 1)  # 目次の次
    ws.sheet_properties.tabColor = '2F5496'

    # 列幅（4列: 区分, 細目, N月(点), 備考）
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 30

    # スタイル定義
    section_font_a = Font(name='游ゴシック', bold=True, color='FFFFFF', size=11)
    section_fill_a = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    subtotal_font_a = Font(name='游ゴシック', bold=True, size=10)
    subtotal_fill_a = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    total_font_a = Font(name='游ゴシック', bold=True, size=11)
    normal_font_a = Font(name='游ゴシック', size=10)
    header_font_a = Font(name='游ゴシック', bold=True, size=10, color='FFFFFF')
    num_fmt = '#,##0'

    NCOLS = 4  # 列数

    def write_section_a(row, title):
        """セクションヘッダーを書く"""
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
        cell = ws.cell(row=row, column=1, value=f'\u25a0 {title}')
        cell.font = section_font_a
        cell.fill = section_fill_a
        for c in range(1, NCOLS + 1):
            ws.cell(row=row, column=c).fill = section_fill_a
            ws.cell(row=row, column=c).border = thin_border
        return row + 1

    def write_item_a(row, cat, detail, val, note=''):
        """明細行を書く"""
        ws.cell(row=row, column=1, value=cat).font = normal_font_a
        ws.cell(row=row, column=2, value=detail).font = normal_font_a
        ws.cell(row=row, column=3, value=val).font = normal_font_a
        ws.cell(row=row, column=3).number_format = num_fmt
        if note:
            ws.cell(row=row, column=4, value=note).font = normal_font_a
        for c in range(1, NCOLS + 1):
            ws.cell(row=row, column=c).border = thin_border
        return row + 1

    def write_subtotal_a(row, val):
        """小計行を書く"""
        ws.cell(row=row, column=2, value='\u2605\u5c0f\u8a08').font = subtotal_font_a
        ws.cell(row=row, column=3, value=val).font = subtotal_font_a
        ws.cell(row=row, column=3).number_format = num_fmt
        for c in range(1, NCOLS + 1):
            ws.cell(row=row, column=c).fill = subtotal_fill_a
            ws.cell(row=row, column=c).border = thin_border
        return row + 1

    # === タイトル ===
    ws.merge_cells('A1:D1')
    ws.cell(row=1, column=1, value=f'さきがけ病院 入院診療報酬 {month}月 区分別内訳')
    ws.cell(row=1, column=1).font = Font(name='游ゴシック', bold=True, size=14)
    ws.merge_cells('A2:D2')
    ws.cell(row=2, column=1, value='単位:点数 ／ 薬剤=全診区横断、食事=除外')
    ws.cell(row=2, column=1).font = Font(name='游ゴシック', size=9, color='666666')

    # === ヘッダー ===
    headers = ['区分', '細目', f'{month}月(点)', '備考']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font = header_font_a
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    row = 5

    # === 診察 ===
    exam = data.get('exam_detail', {})
    row = write_section_a(row, '診察')
    jan_init = exam.get('初診', 0)
    jan_guid = exam.get('指導', 0)
    row = write_item_a(row, '診察', '初診', jan_init)
    row = write_item_a(row, '診察', '指導(特定薬剤管理等)', jan_guid)
    jan_exam = jan_init + jan_guid
    row = write_subtotal_a(row, jan_exam)

    # === 投薬（薬剤+調剤処方） ===
    row = write_section_a(row, '投薬（薬剤+調剤処方）')
    jan_drug = data.get('drug_total', 0)
    jan_disp = data.get('dispensing', 0)
    jan_equip = data.get('equipment', 0)
    row = write_item_a(row, '投薬', '薬剤（全診区横断）', jan_drug)
    row = write_item_a(row, '投薬', '調剤・処方・調基', jan_disp)
    row = write_item_a(row, '投薬', '器材・材料', jan_equip)
    jan_touyaku = data['compare'].get('投薬（薬剤+調剤処方）', 0)
    row = write_subtotal_a(row, jan_touyaku)

    # === 注射（手技+薬剤） ===
    row = write_section_a(row, '注射（手技+薬剤）')
    jan_inj_tech = data.get('injection_tech', 0)
    row = write_item_a(row, '注射', '手技', jan_inj_tech)
    jan_chusha = data['compare'].get('注射（手技+薬剤）', 0)
    jan_inj_drug = jan_chusha - jan_inj_tech
    row = write_item_a(row, '注射', '薬剤(注射+点滴)', jan_inj_drug)
    row = write_subtotal_a(row, jan_chusha)

    # === 処置 ===
    row = write_section_a(row, '処置')
    jan_proc = data.get('procedure', 0)
    row = write_item_a(row, '処置', '処置(+薬剤)', jan_proc)
    row = write_subtotal_a(row, jan_proc)

    # === 検査 ===
    test = data.get('test_detail', {})
    row = write_section_a(row, '検査')
    jan_gen = test.get('一般', 0)
    jan_ecg = test.get('心電図', 0)
    jan_psy = test.get('心理', 0)
    row = write_item_a(row, '検査', '一般', jan_gen)
    row = write_item_a(row, '検査', '心電図・呼吸心拍等', jan_ecg)
    row = write_item_a(row, '検査', '心理検査・他', jan_psy)
    jan_test = jan_gen + jan_ecg + jan_psy
    row = write_subtotal_a(row, jan_test)

    # === 画像 ===
    img = data.get('image_detail', {})
    row = write_section_a(row, '画像')
    jan_xp = img.get('X-P', 0)
    jan_ct = img.get('CT', 0)
    row = write_item_a(row, '画像', 'X-P', jan_xp)
    row = write_item_a(row, '画像', 'CT', jan_ct)
    jan_img = jan_xp + jan_ct
    row = write_subtotal_a(row, jan_img)

    # === 精神科専門 ===
    psych = data.get('psych_detail', {})
    row = write_section_a(row, '精神科専門')
    psych_items = [
        ('入院精神療法', '入院精神療法'),
        ('精神科作業療法', '精神科作業療法'),
        ('治療抵抗性統合失調症', '治療抵抗性統合失調症'),
        ('医療保護入院等', '医療保護入院等'),
        ('精神科退院指導料', '退院指導料'),
        ('その他', 'その他'),
    ]
    jan_psych_total = 0
    for display_name, data_key in psych_items:
        jan_v = psych.get(data_key, 0)
        jan_psych_total += jan_v
        row = write_item_a(row, '精神科専門', display_name, jan_v)
    row = write_subtotal_a(row, jan_psych_total)

    # === ベースアップ評価料 ===
    row = write_section_a(row, 'ベースアップ評価料')
    jan_bu = data['compare'].get('ベースアップ評価料', 0)
    row = write_item_a(row, 'ベースUp', '', jan_bu)
    row = write_subtotal_a(row, jan_bu)

    # === その他・器材 ===
    row = write_section_a(row, 'その他・器材')
    jan_sonota = data['compare'].get('その他・器材', 0)
    row = write_item_a(row, 'その他', '', jan_sonota)
    row = write_subtotal_a(row, jan_sonota)

    # === ★ 区分別 小計 ===
    cat_sum = sum(data['compare'].values())
    ws.cell(row=row, column=1, value='区分別 小計').font = total_font_a
    ws.cell(row=row, column=3, value=cat_sum).font = total_font_a
    ws.cell(row=row, column=3).number_format = num_fmt
    for c in range(1, NCOLS + 1):
        ws.cell(row=row, column=c).fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
        ws.cell(row=row, column=c).border = thin_border
    row += 1

    # === 入院料(1病棟: 精神療養) ===
    w1d = data.get('w1_detail', {})
    row = write_section_a(row, '入院料(1病棟: 精神療養)')
    w1_items = ['精神療養病棟入院料', '重症者加算1', '重症者加算2', '地域移行実施加算', '非定型抗精神病薬']
    jan_w1_total = 0
    for item in w1_items:
        jan_v = w1d.get(item, 0)
        jan_w1_total += jan_v
        row = write_item_a(row, '1病棟', item, jan_v)
    row = write_subtotal_a(row, jan_w1_total)

    # === 入院料(2病棟: 15対1) ===
    w2d = data.get('w2_detail', {})
    row = write_section_a(row, '入院料(2病棟: 15対1)')
    w2_order = ['入院基本料15:1(期間加算含)', '看護補助加算', '看護配置加算',
                '看護補助体制充実', '地域移行実施加算', '隔離室管理加算',
                '身体合併症管理', '重度認知症加算']
    jan_w2_total = 0
    for item in w2_order:
        jan_v = w2d.get(item, 0)
        jan_w2_total += jan_v
        row = write_item_a(row, '2病棟', item, jan_v)
    row = write_subtotal_a(row, jan_w2_total)

    # === 入院料(3病棟: 15対1) ===
    w3d = data.get('w3_detail', {})
    row = write_section_a(row, '入院料(3病棟: 15対1)')
    w3_order = ['入院基本料15:1(期間加算含)', '看護補助加算', '特殊疾患入院施設管理',
                '看護配置加算', '看護補助体制充実', '地域移行実施加算',
                '身体合併症管理', '療養環境加算', '重度認知症加算']
    jan_w3_total = 0
    for item in w3_order:
        jan_v = w3d.get(item, 0)
        jan_w3_total += jan_v
        row = write_item_a(row, '3病棟', item, jan_v)
    row = write_subtotal_a(row, jan_w3_total)

    # === ★ 入院料 小計 ===
    ward_sum = sum(data['ward'].values())
    ws.cell(row=row, column=1, value='入院料 小計').font = total_font_a
    ws.cell(row=row, column=3, value=ward_sum).font = total_font_a
    ws.cell(row=row, column=3).number_format = num_fmt
    for c in range(1, NCOLS + 1):
        ws.cell(row=row, column=c).fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
        ws.cell(row=row, column=c).border = thin_border
    row += 1

    # === ★ 食事 小計 ===
    jan_food = data['food']
    ws.cell(row=row, column=1, value='食事 小計').font = total_font_a
    ws.cell(row=row, column=3, value=jan_food).font = total_font_a
    ws.cell(row=row, column=3).number_format = num_fmt
    for c in range(1, NCOLS + 1):
        ws.cell(row=row, column=c).fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
        ws.cell(row=row, column=c).border = thin_border
    row += 1

    # === 合計 ===
    jan_grand = cat_sum + ward_sum + jan_food
    ws.cell(row=row, column=1, value='合計').font = total_font_a
    ws.cell(row=row, column=3, value=jan_grand).font = total_font_a
    ws.cell(row=row, column=3).number_format = num_fmt
    for c in range(1, NCOLS + 1):
        ws.cell(row=row, column=c).fill = total_fill
        ws.cell(row=row, column=c).border = thin_border
    row += 1

    # 円換算行
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
    ws.cell(row=row, column=1,
            value=f'【円換算】 {month}月: {jan_grand * 10:,}円')
    ws.cell(row=row, column=1).font = Font(name='游ゴシック', size=9, color='666666')

    # === 精神科専門療法 内訳 ===
    row += 2
    row = write_section_a(row, f'精神科専門療法 内訳（{month}月）')
    psych = data.get('psych_detail', {})
    psych_order_a = ['入院精神療法', '精神科作業療法', '医療保護入院等',
                     '退院指導料', '治療抵抗性統合失調症', '退院前訪問指導料', 'その他']
    jan_pt = 0
    for name in psych_order_a:
        jan_v = psych.get(name, 0)
        jan_pt += jan_v
        row = write_item_a(row, '精神科専門', name, jan_v)
    row = write_subtotal_a(row, jan_pt)

    # === 入院精神療法 内訳 ===
    if data.get('psych_therapy_detail'):
        row += 1
        row = write_section_a(row, f'入院精神療法 内訳（{month}月）')
        pt_total_cases = 0
        pt_total_amt = 0
        for name, cases, amt in data['psych_therapy_detail']:
            row = write_item_a(row, '入院精神療法', f'{name}（{cases}件）', amt)
            pt_total_cases += cases
            pt_total_amt += amt
        row = write_subtotal_a(row, pt_total_amt)

    # === 最終請求額 ===
    calc_yen = jan_grand * 10
    diff_yen = SEIKYU_YEN - calc_yen  # 減算分（負の数）

    row += 2
    blue_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')

    # 集計額（円）
    ws.cell(row=row, column=1, value='集計額').font = total_font_a
    ws.cell(row=row, column=3, value=calc_yen).font = total_font_a
    ws.cell(row=row, column=3).number_format = '#,##0"円"'
    for c in range(1, NCOLS + 1):
        ws.cell(row=row, column=c).border = thin_border
    row += 1

    # 減算分
    ws.cell(row=row, column=1, value='減算分').font = Font(name='游ゴシック', bold=True, size=11, color='CC0000')
    ws.cell(row=row, column=3, value=diff_yen).font = Font(name='游ゴシック', bold=True, size=11, color='CC0000')
    ws.cell(row=row, column=3).number_format = '#,##0"円"'
    for c in range(1, NCOLS + 1):
        ws.cell(row=row, column=c).border = thin_border
    row += 1

    # 最終請求額
    ws.cell(row=row, column=1, value='最終請求額').font = Font(name='游ゴシック', bold=True, size=12)
    ws.cell(row=row, column=3, value=SEIKYU_YEN).font = Font(name='游ゴシック', bold=True, size=12)
    ws.cell(row=row, column=3).number_format = '#,##0"円"'
    for c in range(1, NCOLS + 1):
        ws.cell(row=row, column=c).fill = blue_fill
        ws.cell(row=row, column=c).border = thin_border

    # 印刷設定
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.print_title_rows = '4:4'


# ======================================================================
# 5. HTML サマリー生成
# ======================================================================

_COMMON_CSS = """\
@media print{@page{size:A4 portrait;margin:6mm}*{-webkit-print-color-adjust:exact!important;print-color-adjust:exact!important}body{padding:0!important;font-size:7px!important}.no-print{display:none!important}.summary-cards{gap:8px!important;margin-bottom:10px!important}.card{padding:10px!important}.card-value{font-size:18px!important}table{font-size:7.5px!important}th,td{padding:4px 5px!important}}
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:'Inter','Noto Sans JP',sans-serif;font-size:11px;color:#1e293b;background:linear-gradient(135deg,#f8fafc 0%,#e2e8f0 100%);min-height:100vh;padding:20px}
.container{max-width:1000px;margin:0 auto;background:#fff;border-radius:16px;box-shadow:0 4px 24px rgba(0,0,0,0.08);padding:24px}
.header{display:flex;justify-content:space-between;align-items:center;margin-bottom:20px;padding-bottom:16px;border-bottom:1px solid #e2e8f0}
h1{font-size:22px;font-weight:700;color:#0f172a;display:flex;align-items:center;gap:12px}
h1::before{content:'';width:5px;height:28px;background:linear-gradient(180deg,#3b82f6,#1d4ed8);border-radius:3px}
.print-date{font-size:11px;color:#64748b}
.btn-group{display:flex;gap:10px}
.btn{padding:10px 20px;color:#fff;border:none;border-radius:8px;cursor:pointer;font-size:13px;font-weight:600;transition:all 0.2s;display:flex;align-items:center;gap:6px}
.btn-print{background:linear-gradient(135deg,#6366f1,#4f46e5)}
.btn-print:hover{transform:translateY(-2px);box-shadow:0 4px 12px rgba(99,102,241,0.4)}
.btn-save{background:linear-gradient(135deg,#10b981,#059669)}
.btn-save:hover{transform:translateY(-2px);box-shadow:0 4px 12px rgba(16,185,129,0.4)}
.summary-cards{display:grid;grid-template-columns:repeat(4,1fr);gap:16px;margin-bottom:24px}
.card{background:#fff;border-radius:12px;padding:18px;position:relative;overflow:hidden;border:1px solid #e2e8f0;transition:transform 0.2s,box-shadow 0.2s}
.card:hover{transform:translateY(-2px);box-shadow:0 8px 24px rgba(0,0,0,0.1)}
.card::before{content:'';position:absolute;top:0;left:0;right:0;height:4px}
.card-revenue::before{background:linear-gradient(90deg,#3b82f6,#60a5fa)}
.card-expense::before{background:linear-gradient(90deg,#f59e0b,#fbbf24)}
.card-profit::before{background:linear-gradient(90deg,#10b981,#34d399)}
.card-margin::before{background:linear-gradient(90deg,#8b5cf6,#a78bfa)}
.card-label{font-size:11px;color:#64748b;font-weight:500;margin-bottom:8px;display:flex;align-items:center;gap:6px}
.card-icon{width:20px;height:20px;border-radius:6px;display:flex;align-items:center;justify-content:center;font-size:11px}
.card-revenue .card-icon{background:#dbeafe;color:#2563eb}
.card-expense .card-icon{background:#fef3c7;color:#d97706}
.card-profit .card-icon{background:#d1fae5;color:#059669}
.card-margin .card-icon{background:#ede9fe;color:#7c3aed}
.card-value{font-size:26px;font-weight:700;color:#0f172a;font-variant-numeric:tabular-nums;letter-spacing:-0.5px}
.card-sub{display:flex;align-items:center;gap:8px;margin-top:8px;font-size:11px}
.card-period{color:#94a3b8;font-size:10px}
table{width:100%;border-collapse:collapse;font-size:9px;border-radius:12px;overflow:hidden;box-shadow:0 2px 8px rgba(0,0,0,0.06)}
thead th{background:linear-gradient(135deg,#1e3a8a 0%,#1e40af 100%);color:#fff;padding:10px 8px;text-align:center;font-weight:600;letter-spacing:0.3px;position:relative}
thead th::after{content:'';position:absolute;bottom:0;left:0;right:0;height:1px;background:rgba(255,255,255,0.1)}
th:first-child{text-align:left;padding-left:14px;min-width:200px;border-radius:8px 0 0 0}
th:last-child{border-radius:0 8px 0 0}
td{padding:8px 10px;border-bottom:1px solid #f1f5f9;text-align:right;font-family:'Inter',monospace;font-variant-numeric:tabular-nums;color:#334155}
td:first-child{text-align:left;padding-left:14px;font-family:'Noto Sans JP',sans-serif;font-weight:500;color:#1e293b}
tbody tr{background:#fff;transition:background 0.15s}
tbody tr:nth-child(even){background:#fafbfc}
tbody tr:hover{background:#f1f5f9}
.indent{padding-left:28px!important;color:#64748b;font-weight:400!important}
.indent2{padding-left:42px!important;color:#94a3b8;font-weight:400!important;font-size:8.5px}
.section{background:linear-gradient(90deg,#eff6ff,#f8fafc)!important}
.section td{font-weight:600}
.section td:first-child{color:#1d4ed8}
.subtotal{background:#f8fafc!important}
.subtotal td{font-weight:600;color:#1e3a8a}
.grand{background:linear-gradient(90deg,#ecfdf5,#f0fdf4)!important}
.grand td{font-weight:700;color:#047857}
.final{background:linear-gradient(90deg,#eff6ff,#dbeafe)!important}
.final td{font-weight:700;color:#0f172a;font-size:10px}
.negative{color:#dc2626!important}
.pct{color:#94a3b8;font-size:7px;display:block;margin-top:2px}
tbody tr:last-child td:first-child{border-radius:0 0 0 8px}
tbody tr:last-child td:last-child{border-radius:0 0 8px 0}
"""


def _man_yen(yen):
    """83595442 -> '8,360万円'"""
    return f'{round(yen / 10000):,}万円'


def _pt_row(label, points, total, cls='', td_cls=''):
    """点数ベースのテーブル行を生成"""
    pct = f'{points / total * 100:.1f}%' if total and points else ''
    yen_span = f'<span class="pct">{points * 10:,}円</span>' if points >= 100 else ''
    tr = f' class="{cls}"' if cls else ''
    td = f' class="{td_cls}"' if td_cls else ''
    return f'  <tr{tr}><td{td}>{label}</td><td>{points:,}{yen_span}</td><td>{pct}</td></tr>'


def generate_html_summary(data, year, month, report_date, folder):
    """入院収益月次サマリーHTMLを生成"""
    import html as _html

    total = data['total']
    calc_yen = total * 10
    seikyu = SEIKYU_YEN if SEIKYU_YEN > 0 else calc_yen
    diff_yen = seikyu - calc_yen
    ward_sum = sum(data['ward'].values())
    ward_pct = ward_sum / total * 100 if total else 0
    psych = data['compare'].get('精神科専門', 0)
    psych_pct = psych / total * 100 if total else 0
    has_seikyu = SEIKYU_YEN > 0

    # --- テーブル行を組み立て ---
    rows = []

    # 区分別（簡易表示名 → compare キー）
    simple_cats = [
        ('診察', '診察'),
        ('投薬', '投薬（薬剤+調剤処方）'),
        ('注射', '注射（手技+薬剤）'),
        ('処置', '処置'),
        ('検査', '検査'),
        ('画像', '画像'),
    ]
    for disp, key in simple_cats:
        v = data['compare'].get(key, 0)
        rows.append(_pt_row(disp, v, total))

    # 精神科専門セクション
    rows.append(_pt_row('精神科専門', psych, total, cls='section'))
    for name, val in data.get('psych_detail', {}).items():
        rows.append(_pt_row(name, val, total, td_cls='indent'))
        # 入院精神療法の内訳（indent2）
        if name == '入院精神療法' and data.get('psych_therapy_detail'):
            for tname, cases, amt in data['psych_therapy_detail']:
                short = tname.replace('入院精神療法', '')
                label = f'{short} {cases}件'
                rows.append(f'  <tr><td class="indent2">{_html.escape(label)}</td>'
                            f'<td>{amt:,}</td><td></td></tr>')

    # ベースアップ・その他
    bu = data['compare'].get('ベースアップ評価料', 0)
    rows.append(_pt_row('ベースアップ評価料', bu, total))
    etc = data['compare'].get('その他・器材', 0)
    rows.append(_pt_row('その他・器材', etc, total))

    # 入院料 各病棟
    ward_names = [
        ('1病棟（精神療養）', '入院料（1病棟：精神療養）', 'w1_detail'),
        ('2病棟（精神15:1）', '入院料（2病棟：15対1）', 'w2_detail'),
        ('3病棟（精神15:1）', '入院料（3病棟：15対1）', 'w3_detail'),
    ]
    for wkey, wlabel, detail_key in ward_names:
        wval = data['ward'].get(wkey, 0)
        rows.append(_pt_row(wlabel, wval, total, cls='section'))
        for iname, ival in data.get(detail_key, {}).items():
            rows.append(_pt_row(iname, ival, total, td_cls='indent'))

    # 小計・合計
    rows.append(_pt_row('入院料 小計', ward_sum, total, cls='subtotal'))
    rows.append(_pt_row('食事 小計', data['food'], total, cls='subtotal'))
    rows.append(_pt_row('合計', total, total, cls='grand'))

    # 最終集計（円ベース）
    rows.append(f'  <tr><td>集計額</td><td>{calc_yen:,}<span class="pct">円</span></td><td></td></tr>')
    if has_seikyu:
        rows.append(f'  <tr><td class="negative">減算分</td><td class="negative">{diff_yen:,}<span class="pct">円</span></td><td></td></tr>')
        rows.append(f'  <tr class="final"><td>最終請求額</td><td>{seikyu:,}<span class="pct">円</span></td><td></td></tr>')
    else:
        rows.append(f'  <tr class="final"><td>集計合計</td><td>{calc_yen:,}<span class="pct">円</span></td><td></td></tr>')

    table_body = '\n'.join(rows)
    save_name = f'入院収益月次サマリー_{month}月.html'

    html_text = f"""\
<!DOCTYPE html>
<html lang="ja">
<head>
<meta charset="UTF-8">
<title>さきがけホスピタル 入院収益月次サマリー {month}月度</title>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&family=Noto+Sans+JP:wght@400;500;600;700&display=swap" rel="stylesheet">
<style>
{_COMMON_CSS}</style>
</head>
<body>
<div class="container">

<div class="header">
  <h1>さきがけホスピタル 入院収益月次サマリー {month}月度</h1>
  <div class="btn-group no-print">
    <button class="btn btn-save" onclick="saveHTML()">💾 HTMLを保存</button>
    <button class="btn btn-print" onclick="window.print()">🖨️ 印刷 / PDF</button>
  </div>
  <div class="print-date">{report_date}作成</div>
</div>

<div class="summary-cards">
  <div class="card card-revenue">
    <div class="card-label"><span class="card-icon">💰</span>{'最終請求額' if has_seikyu else '集計額'}</div>
    <div class="card-value">{_man_yen(seikyu)}</div>
    <div class="card-sub"><span class="card-period">{seikyu:,}円{'（減算後）' if has_seikyu else '（点数×10円）'}</span></div>
  </div>
  <div class="card card-expense">
    <div class="card-label"><span class="card-icon">🏥</span>入院料 小計</div>
    <div class="card-value">{_man_yen(ward_sum * 10)}</div>
    <div class="card-sub"><span class="card-period">{ward_sum:,}点 ／ 構成比 {ward_pct:.1f}%</span></div>
  </div>
  <div class="card card-profit">
    <div class="card-label"><span class="card-icon">🧠</span>精神科専門 小計</div>
    <div class="card-value">{_man_yen(psych * 10)}</div>
    <div class="card-sub"><span class="card-period">{psych:,}点 ／ 構成比 {psych_pct:.1f}%</span></div>
  </div>
  <div class="card card-margin">
    <div class="card-label"><span class="card-icon">📊</span>{month}月 合計</div>
    <div class="card-value">{_man_yen(seikyu)}</div>
    <div class="card-sub"><span class="card-period">{'最終請求額 ' + f'{seikyu:,}' + '円（減算 ▲' + f'{abs(diff_yen):,}' + '円）' if has_seikyu else f'集計額 {calc_yen:,}円'}</span></div>
  </div>
</div>

<table>
<thead>
  <tr>
    <th>項目</th>
    <th>{month}月（点）</th>
    <th>構成比</th>
  </tr>
</thead>
<tbody>
{table_body}
</tbody>
</table>

</div>

<script>
function saveHTML(){{
  const h=document.documentElement.outerHTML;
  const b=new Blob([h],{{type:'text/html;charset=utf-8'}});
  const u=URL.createObjectURL(b);
  const a=document.createElement('a');
  a.href=u;
  a.download='{save_name}';
  a.click();
  URL.revokeObjectURL(u);
}}
</script>
</body>
</html>
"""

    html_path = os.path.join(folder, save_name)
    with open(html_path, 'w', encoding='utf-8') as f:
        f.write(html_text)
    return html_path


# ======================================================================
# 6. メイン処理
# ======================================================================

def main():
    print('=' * 60)
    print('入院収益月次レポート自動生成')
    print('=' * 60)

    folder = MONTH_FOLDER
    if len(sys.argv) > 1:
        folder = os.path.join(BASE_DIR, sys.argv[1])

    if not os.path.isdir(folder):
        print(f'ERROR: フォルダが見つかりません: {folder}')
        return

    print(f'対象フォルダ: {folder}')

    # 対象期間を取得
    year, month = get_report_period(folder)
    print(f'対象期間: {year}年{month}月')
    report_date = f'令和{year - 2018}年{datetime.date.today().month}月{datetime.date.today().day}日'

    # データ抽出
    print('\nPDF解析中...')
    data = build_data(folder)

    jan = data['total']
    print(f'\n--- 集計結果 ---')
    print(f'  {month}月 入院合計: {jan:>10,}点 ({jan * 10:>12,}円)')

    # HTML生成（月フォルダに出力）
    print('\nHTML生成中...')
    html_dir = os.path.join(BASE_DIR, f'{month}月')
    if not os.path.isdir(html_dir):
        html_dir = BASE_DIR
    html_path = generate_html_summary(data, year, month, report_date, html_dir)
    print(f'OK: {html_path}')

    # 検証
    print(f'\n--- 検証 ---')
    for cat, val in data['compare'].items():
        print(f'  {cat:20s}: {val:>10,}点')


if __name__ == '__main__':
    main()
